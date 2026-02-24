"""
BTR Prospecting System - Backend Server
Flask API with Claude AI integration for automated prospect discovery
"""

from flask import Flask, request, jsonify, send_from_directory, send_file, g, make_response
from flask_cors import CORS
from functools import wraps
import os
import secrets
import hashlib
from datetime import datetime, timedelta
import json
import time
import csv
import io
import anthropic
from dotenv import load_dotenv
import sqlite3
from db import get_db as _get_db_conn, is_postgres as _is_postgres, IntegrityError as _IntegrityError
import re
import threading
import traceback
import uuid
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import pytz
import bcrypt

# Load environment variables
load_dotenv()

SESSION_SECRET = os.getenv('SESSION_SECRET', secrets.token_hex(32))
COOKIE_SECURE = os.getenv('COOKIE_SECURE', 'false').lower() == 'true'
SESSION_DURATION_HOURS = 720  # 30 days
SUPER_ADMIN_EMAIL = os.getenv('SUPER_ADMIN_EMAIL', 'mjhopkins88@gmail.com').strip().lower()

app = Flask(__name__, static_folder='static')
CORS(app, supports_credentials=True)

# --- Login rate limiter (in-memory) ---
_login_attempts = {}  # email -> { count, locked_until }
_LOGIN_MAX_ATTEMPTS = 10
_LOGIN_LOCKOUT_MINUTES = 10

# Initialize Claude client
client = anthropic.Anthropic(api_key=os.getenv('ANTHROPIC_API_KEY'))

# Simple in-memory cache to avoid redundant API calls
# Key: "city|limit", Value: {"prospects": [...], "timestamp": datetime}
_search_cache = {}
CACHE_TTL_SECONDS = 900  # 15 minutes

# Daily Discovery configuration
DISCOVERY_CONFIG = {
    'cities': [
        {'city': 'Phoenix', 'state': 'AZ'},
        {'city': 'Dallas', 'state': 'TX'},
        {'city': 'Atlanta', 'state': 'GA'},
        {'city': 'Charlotte', 'state': 'NC'},
    ],
    'icp_keywords': [
        'build to rent developer',
        'single family rental',
        'residential land developer',
        'homebuilder',
        'horizontal multifamily',
        'BTR community builder',
    ],
    'target_sources': [
        'bisnow.com',
        'multihousingnews.com',
        'credaily.com',
        'commercialobserver.com',
        'bizjournals.com',
        'linkedin.com',
        'sec.gov EDGAR',
    ],
    'monitor_operators': [
        'Invitation Homes',
        'American Homes 4 Rent',
        'Tricon Residential',
        'Progress Residential',
    ],
    'max_signals_per_day': 10,
    'schedule_hour': 7,
    'schedule_minute': 0,
    'timezone': 'America/Los_Angeles',
    'delivery_method': 'in_app',  # 'in_app' or 'webhook'
    'webhook_url': '',
}

# Track whether a discovery run is in progress
_discovery_running = False

# Database setup
_IS_PRODUCTION = os.getenv('RAILWAY_ENVIRONMENT', '') != '' or os.getenv('RAILWAY_PROJECT_ID', '') != ''

def _adapt_schema_sql(sql):
    """Adapt CREATE TABLE SQL for PostgreSQL when needed."""
    if not _is_postgres():
        return sql
    # INTEGER PRIMARY KEY AUTOINCREMENT -> SERIAL PRIMARY KEY
    sql = sql.replace('INTEGER PRIMARY KEY AUTOINCREMENT', 'SERIAL PRIMARY KEY')
    # BOOLEAN DEFAULT 0/1 -> BOOLEAN DEFAULT FALSE/TRUE (only for BOOLEAN columns)
    sql = re.sub(r'BOOLEAN\s+(NOT\s+NULL\s+)?DEFAULT\s+0', r'BOOLEAN \1DEFAULT FALSE', sql)
    sql = re.sub(r'BOOLEAN\s+(NOT\s+NULL\s+)?DEFAULT\s+1', r'BOOLEAN \1DEFAULT TRUE', sql)
    return sql

def _safe_add_column(cursor, table, column, definition):
    """Add a column if it doesn't exist. Works on both SQLite and PostgreSQL."""
    try:
        if _is_postgres():
            cursor.execute(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {column} {definition}")
        else:
            cursor.execute(f'ALTER TABLE {table} ADD COLUMN {column} {definition}')
    except Exception:
        pass  # column already exists

def init_db():
    """Initialize database schema (non-destructive, additive only).
    All CREATE TABLE statements are automatically adapted for PostgreSQL.
    Uses IF NOT EXISTS to be safely re-runnable (migration-safe).
    """
    conn = _get_db_conn()
    _real_cursor = conn.cursor()

    class _SchemaExecProxy:
        """Proxy that auto-adapts schema SQL for the target database engine."""
        def execute(self, sql, params=None):
            adapted = _adapt_schema_sql(sql)
            if params:
                _real_cursor.execute(adapted, params)
            else:
                _real_cursor.execute(adapted)
        def fetchone(self):
            return _real_cursor.fetchone()
        def fetchall(self):
            return _real_cursor.fetchall()
    c = _SchemaExecProxy()

    c.execute('''
        CREATE TABLE IF NOT EXISTS prospects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company TEXT NOT NULL,
            executive TEXT,
            title TEXT,
            linkedin TEXT,
            email TEXT,
            phone TEXT,
            city TEXT,
            state TEXT,
            score INTEGER,
            tiv TEXT,
            units TEXT,
            project_name TEXT,
            project_status TEXT,
            signals TEXT,
            why_now TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(company, project_name)
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS discovery_seen (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            business_name TEXT NOT NULL,
            city TEXT NOT NULL,
            state TEXT NOT NULL,
            address TEXT,
            phone TEXT,
            website TEXT,
            rating REAL,
            review_count INTEGER,
            category TEXT,
            icp_keyword TEXT,
            first_seen TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(business_name, city, state)
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS discovery_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            results_json TEXT,
            digest_text TEXT,
            city_count INTEGER DEFAULT 0,
            total_new INTEGER DEFAULT 0,
            status TEXT DEFAULT 'completed',
            adapter_stats TEXT
        )
    ''')
    # Add adapter_stats column if missing (migration for existing DBs)
    _safe_add_column(_real_cursor if _is_postgres() else c, 'discovery_runs', 'adapter_stats', 'TEXT')
    c.execute('''
        CREATE TABLE IF NOT EXISTS prospecting_runs (
            id TEXT PRIMARY KEY,
            status TEXT NOT NULL DEFAULT 'pending',
            search_params TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP,
            total_prospects INTEGER DEFAULT 0,
            error TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS run_prospects (
            id TEXT PRIMARY KEY,
            run_id TEXT NOT NULL REFERENCES prospecting_runs(id),
            company_name TEXT NOT NULL,
            city TEXT,
            state TEXT,
            score INTEGER DEFAULT 0,
            tiv_estimate TEXT,
            deal_status TEXT,
            signals TEXT,
            why_call_now TEXT,
            executive TEXT,
            title TEXT,
            linkedin TEXT,
            units TEXT,
            project_name TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            score_meta TEXT
        )
    ''')
    # Add score_meta column if missing (migration for existing DBs)
    _safe_add_column(_real_cursor if _is_postgres() else c, 'run_prospects', 'score_meta', 'TEXT')
    c.execute('CREATE INDEX IF NOT EXISTS idx_run_prospects_score ON run_prospects(run_id, score DESC)')
    c.execute('''
        CREATE TABLE IF NOT EXISTS discovery_signal_seen (
            fingerprint TEXT PRIMARY KEY,
            first_seen_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS search_cache (
            cache_key TEXT PRIMARY KEY,
            created_at TIMESTAMP,
            expires_at TIMESTAMP,
            payload_json TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS discovery_source_refresh (
            source_type TEXT PRIMARY KEY,
            last_refreshed_at TIMESTAMP,
            items_found INTEGER DEFAULT 0
        )
    ''')
    # --- Auth & CRM Tables ---
    c.execute('''
        CREATE TABLE IF NOT EXISTS workspaces (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY,
            workspace_id TEXT NOT NULL REFERENCES workspaces(id),
            name TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'producer',
            is_super_admin BOOLEAN NOT NULL DEFAULT 0,
            is_disabled BOOLEAN NOT NULL DEFAULT 0,
            last_login_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    # Migration: add columns if missing on existing databases
    _col_target = _real_cursor if _is_postgres() else c
    for col, defn in [('is_super_admin', 'BOOLEAN NOT NULL DEFAULT FALSE' if _is_postgres() else 'BOOLEAN NOT NULL DEFAULT 0'),
                      ('is_disabled', 'BOOLEAN NOT NULL DEFAULT FALSE' if _is_postgres() else 'BOOLEAN NOT NULL DEFAULT 0'),
                      ('last_login_at', 'TIMESTAMP')]:
        _safe_add_column(_col_target, 'users', col, defn)
    c.execute('''
        CREATE TABLE IF NOT EXISTS sessions (
            id TEXT PRIMARY KEY,
            user_id TEXT NOT NULL REFERENCES users(id),
            session_token TEXT NOT NULL UNIQUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            expires_at TIMESTAMP NOT NULL,
            last_seen_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS admin_events (
            id TEXT PRIMARY KEY,
            workspace_id TEXT NOT NULL,
            actor_user_id TEXT NOT NULL,
            action TEXT NOT NULL,
            target_user_id TEXT,
            details TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute('CREATE INDEX IF NOT EXISTS idx_admin_events_ws ON admin_events(workspace_id, created_at DESC)')
    c.execute('''
        CREATE TABLE IF NOT EXISTS crm_companies (
            id TEXT PRIMARY KEY,
            workspace_id TEXT NOT NULL REFERENCES workspaces(id),
            prospect_key TEXT NOT NULL,
            company_name TEXT NOT NULL,
            website TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(workspace_id, prospect_key)
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS crm_leads (
            id TEXT PRIMARY KEY,
            workspace_id TEXT NOT NULL REFERENCES workspaces(id),
            company_id TEXT NOT NULL REFERENCES crm_companies(id),
            owner_user_id TEXT REFERENCES users(id),
            status TEXT NOT NULL DEFAULT 'New',
            last_touch_at TIMESTAMP,
            next_followup_at TIMESTAMP,
            priority INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS crm_touchpoints (
            id TEXT PRIMARY KEY,
            workspace_id TEXT NOT NULL REFERENCES workspaces(id),
            lead_id TEXT NOT NULL REFERENCES crm_leads(id),
            user_id TEXT NOT NULL REFERENCES users(id),
            type TEXT NOT NULL,
            outcome TEXT,
            notes TEXT,
            occurred_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            next_followup_at TIMESTAMP
        )
    ''')
    c.execute('CREATE INDEX IF NOT EXISTS idx_crm_leads_owner ON crm_leads(workspace_id, owner_user_id, status)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_crm_leads_followup ON crm_leads(workspace_id, next_followup_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_crm_companies_key ON crm_companies(workspace_id, prospect_key)')

    # --- Lead Activity Log ---
    c.execute('''
        CREATE TABLE IF NOT EXISTS lead_activity (
            id TEXT PRIMARY KEY,
            lead_id TEXT NOT NULL REFERENCES crm_leads(id),
            actor_user_id TEXT NOT NULL REFERENCES users(id),
            action_type TEXT NOT NULL,
            old_value TEXT,
            new_value TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute('CREATE INDEX IF NOT EXISTS idx_lead_activity_lead ON lead_activity(lead_id, created_at DESC)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_lead_activity_actor ON lead_activity(actor_user_id, created_at DESC)')

    # --- Trend Detection & Weekly Briefs Tables ---
    c.execute('''
        CREATE TABLE IF NOT EXISTS trend_signals (
            id TEXT PRIMARY KEY,
            state TEXT NOT NULL,
            city TEXT NOT NULL,
            topic TEXT NOT NULL,
            count_7d INTEGER NOT NULL,
            count_30d INTEGER NOT NULL,
            trend_ratio REAL NOT NULL,
            classification TEXT NOT NULL,
            computed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(state, city, topic, computed_at)
        )
    ''')
    c.execute('CREATE INDEX IF NOT EXISTS idx_trend_signals_date ON trend_signals(computed_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_trend_signals_state ON trend_signals(state, computed_at)')
    c.execute('''
        CREATE TABLE IF NOT EXISTS weekly_briefs (
            id TEXT PRIMARY KEY,
            brief_json TEXT NOT NULL,
            generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            week_start TEXT NOT NULL,
            week_end TEXT NOT NULL
        )
    ''')
    # --- Weighted Signals (materialized from discovery_runs + search_cache) ---
    c.execute('''
        CREATE TABLE IF NOT EXISTS weighted_signals (
            id TEXT PRIMARY KEY,
            state TEXT NOT NULL,
            city TEXT NOT NULL,
            topic TEXT NOT NULL,
            signal_weight INTEGER NOT NULL DEFAULT 1,
            entity_name TEXT,
            title TEXT,
            summary TEXT,
            source TEXT,
            source_type TEXT,
            confidence TEXT,
            published_at TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute('CREATE INDEX IF NOT EXISTS idx_weighted_signals_state ON weighted_signals(state, city)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_weighted_signals_date ON weighted_signals(published_at)')
    # --- Market Momentum ---
    c.execute('''
        CREATE TABLE IF NOT EXISTS market_momentum (
            id TEXT PRIMARY KEY,
            state TEXT NOT NULL,
            city TEXT NOT NULL,
            window_end_date TEXT NOT NULL,
            signals_7d INTEGER DEFAULT 0,
            signals_14d INTEGER DEFAULT 0,
            signals_30d INTEGER DEFAULT 0,
            weighted_signals_7d INTEGER DEFAULT 0,
            weighted_signals_14d INTEGER DEFAULT 0,
            weighted_signals_30d INTEGER DEFAULT 0,
            momentum_score REAL DEFAULT 0,
            momentum_label TEXT DEFAULT 'Stable',
            computed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(state, city, window_end_date)
        )
    ''')
    c.execute('CREATE INDEX IF NOT EXISTS idx_market_momentum_state ON market_momentum(state, city, window_end_date)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_market_momentum_date ON market_momentum(window_end_date)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_market_momentum_st ON market_momentum(state)')
    # --- Sunbelt Sparknotes Summaries (cached LLM output) ---
    c.execute('''
        CREATE TABLE IF NOT EXISTS sunbelt_summaries (
            id TEXT PRIMARY KEY,
            tab TEXT NOT NULL,
            window_days INTEGER NOT NULL,
            date_bucket TEXT NOT NULL,
            payload_json TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(tab, window_days, date_bucket)
        )
    ''')
    c.execute('CREATE INDEX IF NOT EXISTS idx_sunbelt_summaries_lookup ON sunbelt_summaries(tab, window_days, date_bucket)')
    # --- Lead Timing Scores ---
    c.execute('''
        CREATE TABLE IF NOT EXISTS lead_timing_scores (
            id TEXT PRIMARY KEY,
            workspace_id TEXT,
            prospect_key TEXT NOT NULL,
            company_name TEXT NOT NULL,
            state TEXT,
            city TEXT,
            trigger_severity INTEGER DEFAULT 0,
            swim_lane_fit INTEGER,
            engagement_score INTEGER,
            market_momentum_score REAL DEFAULT 50,
            freshness_score INTEGER DEFAULT 30,
            call_timing_score REAL DEFAULT 0,
            timing_label TEXT DEFAULT 'Watch',
            reasons TEXT,
            computed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(workspace_id, prospect_key)
        )
    ''')
    c.execute('CREATE INDEX IF NOT EXISTS idx_lead_timing_key ON lead_timing_scores(workspace_id, prospect_key)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_lead_timing_score ON lead_timing_scores(call_timing_score DESC)')

    # ---- Quoting tables ----
    c.execute('''
        CREATE TABLE IF NOT EXISTS rate_groups (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL UNIQUE,
            rate REAL NOT NULL,
            rate_x100 REAL NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS county_group_map (
            id TEXT PRIMARY KEY,
            state TEXT NOT NULL,
            county_name TEXT NOT NULL,
            group_name TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(state, county_name)
        )
    ''')
    c.execute('CREATE INDEX IF NOT EXISTS idx_county_map_lookup ON county_group_map(state, county_name)')
    c.execute('''
        CREATE TABLE IF NOT EXISTS quote_requests (
            id TEXT PRIMARY KEY,
            workspace_id TEXT,
            created_by_user_id TEXT,
            sqft INTEGER NOT NULL,
            rc_per_sf REAL NOT NULL,
            loss_rents REAL NOT NULL,
            city TEXT,
            state TEXT,
            county TEXT,
            grouping_name TEXT,
            rate_x100 REAL,
            aop_buydown BOOLEAN DEFAULT 0,
            replacement_cost REAL,
            total_tiv REAL,
            base_premium REAL,
            taxes REAL,
            total_premium REAL,
            warnings TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS county_cache (
            id TEXT PRIMARY KEY,
            cache_key TEXT NOT NULL UNIQUE,
            county TEXT,
            candidates TEXT,
            confidence TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Seed rate_groups if empty
    c.execute('SELECT COUNT(*) FROM rate_groups')
    if c.fetchone()[0] == 0:
        _seed_rates = [
            ('AOP Buydown',                          0.000345, 0.0345),
            ('All Other Locations',                   0.001739, 0.1739),
            ('All Other Tier 1',                      0.003659, 0.3659),
            ('Florida Group 1',                       0.007592, 0.7592),
            ('Florida Group 2',                       0.005827, 0.5827),
            ('Florida Group 3',                       0.005141, 0.5141),
            ('Florida Group 4',                       0.004841, 0.4841),
            ('Tier 1 Louisiana',                      0.004113, 0.4113),
            ('Tier 1 Atlantic (GA, SC, NC, VA)',      0.003108, 0.3108),
            ('Texas North',                           0.003337, 0.3337),
            ('Texas - Southern Non-Coastal',          0.002096, 0.2096),
        ]
        for name, rate, rate_x100 in _seed_rates:
            c.execute('INSERT INTO rate_groups (id, name, rate, rate_x100) VALUES (?, ?, ?, ?)',
                      (str(uuid.uuid4()), name, rate, rate_x100))

    # Seed county_group_map if empty
    c.execute('SELECT COUNT(*) FROM county_group_map')
    if c.fetchone()[0] == 0:
        _county_seeds = {
            'Florida Group 1': {
                'FL': ['Monroe', 'Miami-Dade', 'Broward', 'Palm Beach'],
            },
            'Florida Group 2': {
                'FL': ['Hillsborough', 'Pinellas'],
            },
            'Florida Group 3': {
                'FL': ['Escambia', 'Santa Rosa', 'Okaloosa', 'Walton', 'Bay', 'Gulf', 'Franklin',
                       'Wakulla', 'Jefferson', 'Taylor', 'Dixie', 'Levy', 'Citrus', 'Hernando',
                       'Pasco', 'Manatee', 'Sarasota', 'Charlotte', 'Lee', 'Collier', 'Martin',
                       'St. Lucie', 'Indian River', 'Brevard', 'Volusia', 'Flagler', 'St. Johns',
                       'Duval', 'Nassau'],
            },
            'Florida Group 4': {
                'FL': ['Washington', 'Holmes', 'Jackson', 'Calhoun', 'Liberty', 'Gadsden', 'Leon',
                       'Madison', 'Hamilton', 'Lafayette', 'Suwannee', 'Columbia', 'Gilchrist',
                       'Baker', 'Union', 'Bradford', 'Clay', 'Alachua', 'Putnam', 'Marion', 'Lake',
                       'Seminole', 'Sumter', 'Polk', 'Orange', 'Osceola', 'Okeechobee', 'Hardee',
                       'Desoto', 'Highlands', 'Glades', 'Hendry'],
            },
            'Tier 1 Louisiana': {
                'LA': ['Assumption', 'Calcasieu', 'Cameron', 'Iberia', 'Jefferson', 'Lafourche',
                       'Livingston', 'Orleans', 'Plaquemines', 'St. Bernard', 'St. Charles',
                       'St. James', 'St. John The Baptist', 'St. Martin (South)', 'St. Mary',
                       'St. Tammany', 'Tangipahoa', 'Terrebonne', 'Vermilion'],
            },
            'Tier 1 Atlantic (GA, SC, NC, VA)': {
                'GA': ['Bryan', 'Camden', 'Chatham', 'Glynn', 'Liberty', 'Mcintosh'],
                'SC': ['Beaufort', 'Berkeley', 'Charleston', 'Colleton', 'Dorchester',
                       'Georgetown', 'Hampton', 'Horry', 'Jasper'],
                'NC': ['Beaufort', 'Bertie', 'Brunswick', 'Camden', 'Carteret', 'Chowan',
                       'Columbus', 'Craven', 'Currituck', 'Dare', 'Hyde', 'Jones',
                       'New Hanover', 'Onslow', 'Pamlico', 'Pasquotank', 'Pender',
                       'Perquimans', 'Tyrrell', 'Washington'],
            },
            'All Other Tier 1': {
                'AL': ['Baldwin', 'Mobile'],
                'TX': ['Aransas', 'Brazoria', 'Calhoun', 'Cameron', 'Chambers', 'Galveston',
                       'Harris', 'Jackson', 'Jefferson', 'Kenedy', 'Kleberg', 'Liberty',
                       'Matagorda', 'Newton', 'Nueces', 'Orange', 'Refugio', 'San Patricio',
                       'Victoria', 'Willacy'],
                'MS': ['Hancock', 'Harrison', 'Jackson'],
            },
            'Texas - Southern Non-Coastal': {
                'TX': ['Atascosa', 'Bee', 'Brooks', 'Dewitt', 'Dimmit', 'Duval', 'Frio',
                       'Goliad', 'Hidalgo', 'Jim Hogg', 'Jim Wells', 'Karnes', 'La Salle',
                       'Live Oak', 'Maverick', 'Mcmullen', 'Starr', 'Webb', 'Wilson',
                       'Zapata', 'Zavala'],
            },
        }
        for group_name, states in _county_seeds.items():
            for st, counties in states.items():
                for county in counties:
                    c.execute('INSERT INTO county_group_map (id, state, county_name, group_name) VALUES (?, ?, ?, ?)',
                              (str(uuid.uuid4()), st, county, group_name))

    # government_signals table for Phase 1 gov enrichment
    c.execute('''
        CREATE TABLE IF NOT EXISTS government_signals (
            id TEXT PRIMARY KEY,
            city TEXT NOT NULL,
            state TEXT NOT NULL,
            signal_type TEXT NOT NULL,
            operator_name TEXT,
            operator_aliases TEXT,
            project_name TEXT,
            amount REAL,
            filing_date TEXT NOT NULL,
            source_url TEXT NOT NULL,
            source_name TEXT NOT NULL,
            summary TEXT NOT NULL,
            raw_payload TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    try:
        c.execute('CREATE INDEX IF NOT EXISTS idx_gov_signals_city_date ON government_signals (state, city, filing_date DESC)')
    except Exception:
        pass
    try:
        c.execute('CREATE INDEX IF NOT EXISTS idx_gov_signals_type_date ON government_signals (signal_type, filing_date DESC)')
    except Exception:
        pass

    # --- Broker Saved Items (Deal Board) ---
    c.execute('''
        CREATE TABLE IF NOT EXISTS broker_saved (
            id TEXT PRIMARY KEY,
            user_id TEXT NOT NULL,
            prospect_id INTEGER,
            company TEXT,
            notes TEXT DEFAULT '',
            status TEXT DEFAULT 'saved',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    try:
        c.execute('CREATE INDEX IF NOT EXISTS idx_broker_saved_user ON broker_saved(user_id)')
    except Exception:
        pass

    # --- Underwriting Communities + Rows ---
    c.execute('''
        CREATE TABLE IF NOT EXISTS underwriting_communities (
            id TEXT PRIMARY KEY,
            community_key TEXT UNIQUE,
            location_name TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    # Build underwriting_rows table with all canonical columns
    from underwriting_columns import COLUMN_KEYS
    _uw_data_cols = ', '.join(f'{k} TEXT' for k in COLUMN_KEYS)
    c.execute(f'''
        CREATE TABLE IF NOT EXISTS underwriting_rows (
            id TEXT PRIMARY KEY,
            community_id TEXT NOT NULL,
            row_version INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            {_uw_data_cols}
        )
    ''')
    try:
        c.execute('CREATE INDEX IF NOT EXISTS idx_uw_rows_community ON underwriting_rows(community_id)')
    except Exception:
        pass

    conn.commit()
    conn.close()

init_db()


# ===================================================================
# AUTH HELPERS & MIDDLEWARE
# ===================================================================

def _hash_password(password):
    """Hash a password using bcrypt."""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def _check_password(password, password_hash):
    """Verify a password against its bcrypt hash."""
    return bcrypt.checkpw(password.encode('utf-8'), password_hash.encode('utf-8'))

def _get_session_user(conn=None):
    """Look up the current user from the session cookie. Returns (user_dict, workspace_id) or (None, None)."""
    token = request.cookies.get('session_token')
    if not token:
        return None, None
    close_conn = False
    if conn is None:
        conn = _get_db_conn()
        close_conn = True
    c = conn.cursor()
    c.execute('''
        SELECT s.id, s.user_id, s.expires_at,
               u.id, u.workspace_id, u.name, u.email, u.role, u.is_super_admin, u.is_disabled
        FROM sessions s JOIN users u ON s.user_id = u.id
        WHERE s.session_token = ? AND s.expires_at > ?
    ''', (token, datetime.utcnow().isoformat()))
    row = c.fetchone()
    if not row:
        if close_conn:
            conn.close()
        return None, None
    # Deny disabled users
    if row[9]:
        if close_conn:
            conn.close()
        return None, None
    # Update last_seen_at
    c.execute('UPDATE sessions SET last_seen_at = ? WHERE id = ?', (datetime.utcnow().isoformat(), row[0]))
    conn.commit()
    user = {
        'id': row[3],
        'workspace_id': row[4],
        'name': row[5],
        'email': row[6],
        'role': row[7],
        'is_super_admin': bool(row[8]),
    }
    if close_conn:
        conn.close()
    return user, row[4]

def _has_users():
    """Check if any users exist in the database."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM users')
    count = c.fetchone()[0]
    conn.close()
    return count > 0

def require_auth(f):
    """Decorator: require a valid session. Sets g.user and g.workspace_id."""
    @wraps(f)
    def decorated(*args, **kwargs):
        # If no users exist yet, allow unauthenticated access (bootstrap mode)
        if not _has_users():
            g.user = None
            g.workspace_id = None
            return f(*args, **kwargs)
        user, workspace_id = _get_session_user()
        if not user:
            return jsonify({'success': False, 'message': 'Authentication required', 'auth_required': True}), 401
        g.user = user
        g.workspace_id = workspace_id
        return f(*args, **kwargs)
    return decorated

def require_role(role):
    """Decorator: require a specific role (must be used after require_auth)."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if g.user and g.user.get('role') != role:
                return jsonify({'success': False, 'message': 'Insufficient permissions'}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator

_VALID_ROLES = ('admin', 'producer', 'broker')

def require_any_role(*roles):
    """Decorator: require user has one of the specified roles (must be used after require_auth)."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if g.user and g.user.get('role') not in roles:
                return jsonify({'success': False, 'message': 'Insufficient permissions'}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator

def require_super_admin(f):
    """Decorator: require super admin (must be used after require_auth).
    Checks BOTH is_super_admin flag AND email matches SUPER_ADMIN_EMAIL."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not g.user:
            return jsonify({'success': False, 'message': 'Authentication required'}), 401
        if g.user.get('role') != 'admin' or not g.user.get('is_super_admin') or g.user.get('email') != SUPER_ADMIN_EMAIL:
            return jsonify({'success': False, 'message': 'Super admin access required'}), 403
        return f(*args, **kwargs)
    return decorated

def _check_login_rate(email):
    """Returns True if login is allowed, False if locked out."""
    info = _login_attempts.get(email)
    if not info:
        return True
    if info.get('locked_until') and datetime.utcnow() < info['locked_until']:
        return False
    if info.get('locked_until') and datetime.utcnow() >= info['locked_until']:
        # Lockout expired, reset
        _login_attempts.pop(email, None)
        return True
    return True

def _record_login_failure(email):
    """Record a failed login attempt."""
    info = _login_attempts.get(email, {'count': 0})
    info['count'] = info.get('count', 0) + 1
    if info['count'] >= _LOGIN_MAX_ATTEMPTS:
        info['locked_until'] = datetime.utcnow() + timedelta(minutes=_LOGIN_LOCKOUT_MINUTES)
    _login_attempts[email] = info

def _clear_login_failures(email):
    """Clear login failure count on success."""
    _login_attempts.pop(email, None)

def _make_prospect_key(company_name, website=None, city=None, state=None):
    """Generate a stable prospect_key for CRM matching."""
    norm_company = re.sub(r'[^a-z0-9]', '', (company_name or '').lower())
    if website:
        norm_domain = re.sub(r'^https?://(www\.)?', '', (website or '').lower()).rstrip('/')
        return f"{norm_company}|{norm_domain}"
    norm_city = re.sub(r'[^a-z0-9]', '', (city or '').lower())
    norm_state = re.sub(r'[^a-z0-9]', '', (state or '').lower())
    return f"{norm_company}|{norm_city}|{norm_state}"


# ===================================================================
# AUTH API ROUTES
# ===================================================================

def _cleanup_expired_sessions():
    """Remove expired sessions to prevent DB bloat. Called during login."""
    try:
        conn = _get_db_conn()
        c = conn.cursor()
        c.execute('DELETE FROM sessions WHERE expires_at < ?', (datetime.utcnow().isoformat(),))
        conn.commit()
        conn.close()
    except Exception:
        pass

@app.route('/api/auth/bootstrap', methods=['POST'])
def api_auth_bootstrap():
    """Create the first admin user + workspace. Only works if zero users exist."""
    if _has_users():
        return jsonify({'success': False, 'message': 'Bootstrap not available. Users already exist.'}), 403
    data = request.json or {}
    workspace_name = data.get('workspace_name', '').strip()
    name = data.get('name', '').strip()
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')

    if not workspace_name or not name or not email or not password:
        return jsonify({'success': False, 'message': 'All fields are required: workspace_name, name, email, password'}), 400
    if len(password) < 8:
        return jsonify({'success': False, 'message': 'Password must be at least 8 characters'}), 400

    workspace_id = str(uuid.uuid4())
    user_id = str(uuid.uuid4())
    password_hash = _hash_password(password)

    is_super = (email == SUPER_ADMIN_EMAIL)

    conn = _get_db_conn()
    c = conn.cursor()
    now = datetime.utcnow().isoformat()
    try:
        c.execute('INSERT INTO workspaces (id, name) VALUES (?, ?)', (workspace_id, workspace_name))
        c.execute('''INSERT INTO users (id, workspace_id, name, email, password_hash, role, is_super_admin, last_login_at)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                  (user_id, workspace_id, name, email, password_hash, 'admin', is_super, now))
        conn.commit()
    except _IntegrityError:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': 'Email already exists'}), 400
    except Exception as e:
        app.logger.error(f'[Bootstrap] DB error creating workspace/user: {e}')
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': f'Database error: {e}'}), 500

    # Auto-login: create session
    try:
        session_token = secrets.token_urlsafe(48)
        session_id = str(uuid.uuid4())
        expires_at = (datetime.utcnow() + timedelta(hours=SESSION_DURATION_HOURS)).isoformat()
        c.execute('INSERT INTO sessions (id, user_id, session_token, expires_at) VALUES (?, ?, ?, ?)',
                  (session_id, user_id, session_token, expires_at))
        conn.commit()
    except Exception as e:
        app.logger.error(f'[Bootstrap] DB error creating session: {e}')
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': f'User created but session failed: {e}'}), 500
    conn.close()

    resp = make_response(jsonify({
        'success': True,
        'user': {'id': user_id, 'name': name, 'email': email, 'role': 'admin',
                 'workspace_id': workspace_id, 'is_super_admin': bool(is_super)}
    }))
    resp.set_cookie('session_token', session_token, httponly=True, samesite='Lax',
                    secure=COOKIE_SECURE, path='/', max_age=SESSION_DURATION_HOURS * 3600)
    return resp


@app.route('/api/auth/login', methods=['POST'])
def api_auth_login():
    """Login with email + password."""
    _cleanup_expired_sessions()
    data = request.json or {}
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')

    if not email or not password:
        return jsonify({'success': False, 'message': 'Email and password are required'}), 400

    # Rate limit check
    if not _check_login_rate(email):
        return jsonify({'success': False, 'message': 'Too many login attempts. Try again in 10 minutes.'}), 429

    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT id, workspace_id, name, email, password_hash, role, is_super_admin, is_disabled FROM users WHERE email = ?', (email,))
    row = c.fetchone()

    if not row or not _check_password(password, row[4]):
        conn.close()
        _record_login_failure(email)
        return jsonify({'success': False, 'message': 'Invalid email or password'}), 401

    # Check if user is disabled
    if row[7]:
        conn.close()
        return jsonify({'success': False, 'message': 'Account is disabled. Contact your administrator.'}), 403

    _clear_login_failures(email)
    user_id, workspace_id, name, user_email, _, role, is_super, _ = row

    # Update last_login_at
    now = datetime.utcnow().isoformat()
    c.execute('UPDATE users SET last_login_at = ? WHERE id = ?', (now, user_id))

    # Create session
    session_token = secrets.token_urlsafe(48)
    session_id = str(uuid.uuid4())
    expires_at = (datetime.utcnow() + timedelta(hours=SESSION_DURATION_HOURS)).isoformat()
    c.execute('INSERT INTO sessions (id, user_id, session_token, expires_at) VALUES (?, ?, ?, ?)',
              (session_id, user_id, session_token, expires_at))
    conn.commit()
    conn.close()

    resp = make_response(jsonify({
        'success': True,
        'user': {'id': user_id, 'name': name, 'email': user_email, 'role': role,
                 'workspace_id': workspace_id, 'is_super_admin': bool(is_super)}
    }))
    resp.set_cookie('session_token', session_token, httponly=True, samesite='Lax',
                    secure=COOKIE_SECURE, path='/', max_age=SESSION_DURATION_HOURS * 3600)
    return resp


@app.route('/api/auth/logout', methods=['POST'])
def api_auth_logout():
    """Logout: invalidate session."""
    token = request.cookies.get('session_token')
    if token:
        conn = _get_db_conn()
        c = conn.cursor()
        c.execute('DELETE FROM sessions WHERE session_token = ?', (token,))
        conn.commit()
        conn.close()
    resp = make_response(jsonify({'success': True}))
    resp.delete_cookie('session_token', path='/')
    return resp


@app.route('/api/auth/me', methods=['GET'])
def api_auth_me():
    """Get current authenticated user."""
    has_users = _has_users()
    if not has_users:
        return jsonify({'success': True, 'user': None, 'needs_bootstrap': True})
    user, workspace_id = _get_session_user()
    if not user:
        return jsonify({'success': False, 'user': None, 'auth_required': True}), 401
    return jsonify({'success': True, 'user': user})


@app.route('/api/auth/has-users', methods=['GET'])
def api_auth_has_users():
    """Check if any users exist (determines bootstrap vs login)."""
    return jsonify({'has_users': _has_users()})


@app.route('/api/auth/users', methods=['GET'])
@require_auth
@require_role('admin')
def api_auth_list_users():
    """Admin: list all users in workspace."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT id, name, email, role, created_at FROM users WHERE workspace_id = ?', (g.workspace_id,))
    users = [{'id': r[0], 'name': r[1], 'email': r[2], 'role': r[3], 'created_at': r[4]} for r in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'users': users})


@app.route('/api/auth/users', methods=['POST'])
@require_auth
@require_role('admin')
def api_auth_create_user():
    """Admin: create a new user in the workspace."""
    data = request.json or {}
    name = data.get('name', '').strip()
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')
    role = data.get('role', 'producer')

    if not name or not email or not password:
        return jsonify({'success': False, 'message': 'name, email, and password are required'}), 400
    if len(password) < 8:
        return jsonify({'success': False, 'message': 'Password must be at least 8 characters'}), 400
    if role not in _VALID_ROLES:
        return jsonify({'success': False, 'message': f'Role must be one of: {", ".join(_VALID_ROLES)}'}), 400

    user_id = str(uuid.uuid4())
    password_hash = _hash_password(password)

    conn = _get_db_conn()
    c = conn.cursor()
    try:
        c.execute('INSERT INTO users (id, workspace_id, name, email, password_hash, role) VALUES (?, ?, ?, ?, ?, ?)',
                  (user_id, g.workspace_id, name, email, password_hash, role))
        conn.commit()
    except _IntegrityError:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': 'Email already exists'}), 400
    conn.close()
    return jsonify({'success': True, 'user': {'id': user_id, 'name': name, 'email': email, 'role': role}})


def _log_admin_event(conn, workspace_id, actor_id, action, target_id=None, details=None):
    """Record an admin audit event."""
    c = conn.cursor()
    c.execute('''INSERT INTO admin_events (id, workspace_id, actor_user_id, action, target_user_id, details)
                 VALUES (?, ?, ?, ?, ?, ?)''',
              (str(uuid.uuid4()), workspace_id, actor_id, action, target_id, details))


@app.route('/api/admin/users', methods=['GET'])
@require_auth
@require_super_admin
def api_admin_list_users():
    """Super admin: list all users in workspace with full details."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('''SELECT id, name, email, role, is_super_admin, is_disabled, last_login_at, created_at
                 FROM users WHERE workspace_id = ? ORDER BY created_at''', (g.workspace_id,))
    users = []
    for r in c.fetchall():
        users.append({
            'id': r[0], 'name': r[1], 'email': r[2], 'role': r[3],
            'is_super_admin': bool(r[4]), 'is_disabled': bool(r[5]),
            'last_login_at': r[6], 'created_at': r[7]
        })
    conn.close()
    return jsonify({'success': True, 'users': users})


@app.route('/api/admin/users', methods=['POST'])
@require_auth
@require_super_admin
def api_admin_create_user():
    """Super admin: create a new user in the workspace."""
    data = request.json or {}
    name = data.get('name', '').strip()
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')
    role = data.get('role', 'producer')

    if not name or not email or not password:
        return jsonify({'success': False, 'message': 'name, email, and password are required'}), 400
    if len(password) < 8:
        return jsonify({'success': False, 'message': 'Password must be at least 8 characters'}), 400
    if role not in _VALID_ROLES:
        return jsonify({'success': False, 'message': f'Role must be one of: {", ".join(_VALID_ROLES)}'}), 400

    user_id = str(uuid.uuid4())
    password_hash = _hash_password(password)

    conn = _get_db_conn()
    c = conn.cursor()
    try:
        c.execute('''INSERT INTO users (id, workspace_id, name, email, password_hash, role, is_super_admin)
                     VALUES (?, ?, ?, ?, ?, ?, ?)''',
                  (user_id, g.workspace_id, name, email, password_hash, role, False))
        _log_admin_event(conn, g.workspace_id, g.user['id'], 'create_user', user_id,
                         json.dumps({'name': name, 'email': email, 'role': role}))
        conn.commit()
    except _IntegrityError:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': 'Email already exists'}), 400
    conn.close()
    return jsonify({'success': True, 'user': {'id': user_id, 'name': name, 'email': email, 'role': role}})


@app.route('/api/admin/users/<user_id>/disable', methods=['POST'])
@require_auth
@require_super_admin
def api_admin_disable_user(user_id):
    """Super admin: disable or enable a user account."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT id, email, is_super_admin, is_disabled FROM users WHERE id = ? AND workspace_id = ?',
              (user_id, g.workspace_id))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': 'User not found'}), 404
    if row[2]:  # is_super_admin
        conn.close()
        return jsonify({'success': False, 'message': 'Cannot disable the super admin account'}), 403
    if user_id == g.user['id']:
        conn.close()
        return jsonify({'success': False, 'message': 'Cannot disable your own account'}), 403

    new_state = not bool(row[3])
    c.execute('UPDATE users SET is_disabled = ? WHERE id = ?', (new_state, user_id))
    if new_state:
        # Revoke all sessions for this user
        c.execute('DELETE FROM sessions WHERE user_id = ?', (user_id,))
    action = 'disable_user' if new_state else 'enable_user'
    _log_admin_event(conn, g.workspace_id, g.user['id'], action, user_id)
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'is_disabled': new_state})


@app.route('/api/admin/users/<user_id>', methods=['DELETE'])
@require_auth
@require_super_admin
def api_admin_delete_user(user_id):
    """Super admin: hard-delete a user account."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT id, email, is_super_admin FROM users WHERE id = ? AND workspace_id = ?',
              (user_id, g.workspace_id))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': 'User not found'}), 404
    if row[2]:  # is_super_admin
        conn.close()
        return jsonify({'success': False, 'message': 'Cannot delete the super admin account'}), 403
    if user_id == g.user['id']:
        conn.close()
        return jsonify({'success': False, 'message': 'Cannot delete your own account'}), 403

    # Revoke sessions then delete user
    c.execute('DELETE FROM sessions WHERE user_id = ?', (user_id,))
    c.execute('DELETE FROM users WHERE id = ?', (user_id,))
    _log_admin_event(conn, g.workspace_id, g.user['id'], 'delete_user', user_id,
                     json.dumps({'email': row[1]}))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/admin/users/<user_id>/reset-password', methods=['POST'])
@require_auth
@require_super_admin
def api_admin_reset_password(user_id):
    """Super admin: reset a user's password."""
    data = request.json or {}
    new_password = data.get('password', '')
    if len(new_password) < 8:
        return jsonify({'success': False, 'message': 'Password must be at least 8 characters'}), 400

    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT id, email FROM users WHERE id = ? AND workspace_id = ?', (user_id, g.workspace_id))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': 'User not found'}), 404

    password_hash = _hash_password(new_password)
    c.execute('UPDATE users SET password_hash = ? WHERE id = ?', (password_hash, user_id))
    # Revoke all existing sessions so user must re-login with new password
    c.execute('DELETE FROM sessions WHERE user_id = ?', (user_id,))
    _log_admin_event(conn, g.workspace_id, g.user['id'], 'reset_password', user_id)
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ===================================================================
# QUOTING: County Lookup, Grouping, Calculation, and API
# ===================================================================

QUOTE_TAX_RATE = 0.06
AOP_BUYDOWN_RATE_X100 = 0.0345
_COUNTY_CACHE_DAYS = 30

# State fallback mapping for grouping resolution
_STATE_FALLBACK_GROUPS = {
    'GA': 'Tier 1 Atlantic (GA, SC, NC, VA)',
    'SC': 'Tier 1 Atlantic (GA, SC, NC, VA)',
    'NC': 'Tier 1 Atlantic (GA, SC, NC, VA)',
    'VA': 'Tier 1 Atlantic (GA, SC, NC, VA)',
    'LA': 'Tier 1 Louisiana',
    'TX': 'Texas North',       # Counties not mapped to Tier 1 or Southern default to North
    'FL': 'Florida Group 4',   # Unmapped FL counties fallback to Group 4
}


def _normalize_county(name):
    """Normalize county name: strip ' County', titlecase."""
    if not name:
        return name
    n = name.strip()
    for suffix in (' County', ' Parish', ' Borough', ' Census Area', ' Municipality'):
        if n.lower().endswith(suffix.lower()):
            n = n[:len(n) - len(suffix)]
    return n.strip().title()


def resolve_county(city, state, street=None, zip_code=None):
    """Resolve county from city/state via Nominatim + FCC Area API. Returns dict with county, confidence, candidates."""
    import requests as req

    # Build cache key
    parts = [
        (street or '').strip().lower(),
        city.strip().lower(),
        state.strip().upper(),
        (zip_code or '').strip(),
    ]
    cache_key = '|'.join(parts)

    # Check cache
    conn = _get_db_conn()
    c = conn.cursor()
    cutoff = (datetime.utcnow() - timedelta(days=_COUNTY_CACHE_DAYS)).isoformat()
    c.execute('SELECT county, candidates, confidence FROM county_cache WHERE cache_key = ? AND created_at > ?',
              (cache_key, cutoff))
    cached = c.fetchone()
    if cached:
        conn.close()
        candidates = json.loads(cached[1]) if cached[1] else []
        return {'county': cached[0], 'confidence': cached[2], 'candidates': candidates}

    # Build geocode query
    query_parts = []
    if street:
        query_parts.append(street.strip())
    query_parts.append(city.strip())
    query_parts.append(state.strip())
    if zip_code:
        query_parts.append(zip_code.strip())
    query_parts.append('USA')
    query = ', '.join(query_parts)

    county = None
    confidence = 'low'
    candidates = []

    try:
        # Step 1: Geocode with Nominatim
        geo_resp = req.get('https://nominatim.openstreetmap.org/search', params={
            'q': query, 'format': 'json', 'limit': 3, 'addressdetails': 1, 'countrycodes': 'us'
        }, headers={'User-Agent': 'BTR-Prospecting/1.0'}, timeout=8)

        if geo_resp.status_code == 200:
            results = geo_resp.json()
            if results:
                # Try to get county directly from address details first
                for r in results:
                    addr = r.get('address', {})
                    c_name = addr.get('county', '')
                    if c_name:
                        norm = _normalize_county(c_name)
                        if norm and norm not in candidates:
                            candidates.append(norm)

                # Also try FCC API with lat/lon from first result
                lat, lon = results[0].get('lat'), results[0].get('lon')
                if lat and lon:
                    try:
                        fcc_resp = req.get('https://geo.fcc.gov/api/census/area', params={
                            'lat': lat, 'lon': lon, 'format': 'json'
                        }, timeout=8)
                        if fcc_resp.status_code == 200:
                            fcc_data = fcc_resp.json()
                            fcc_results = fcc_data.get('results', [])
                            if fcc_results:
                                fcc_county = fcc_results[0].get('county_name', '')
                                if fcc_county:
                                    norm_fcc = _normalize_county(fcc_county)
                                    if norm_fcc and norm_fcc not in candidates:
                                        candidates.insert(0, norm_fcc)
                    except Exception:
                        pass

        # Determine best county and confidence
        if candidates:
            county = candidates[0]
            if len(candidates) == 1:
                confidence = 'high' if street else 'medium'
            else:
                # Multiple candidates — check if they agree
                if all(c == candidates[0] for c in candidates):
                    confidence = 'high'
                else:
                    confidence = 'low'
    except Exception as e:
        app.logger.warning(f'County lookup failed: {e}')

    # Cache result
    try:
        c2 = conn.cursor()
        c2.execute('''INSERT INTO county_cache (id, cache_key, county, candidates, confidence) VALUES (?, ?, ?, ?, ?)
                      ON CONFLICT (cache_key) DO UPDATE SET id = EXCLUDED.id, county = EXCLUDED.county, candidates = EXCLUDED.candidates, confidence = EXCLUDED.confidence''',
                   (str(uuid.uuid4()), cache_key, county, json.dumps(candidates), confidence))
        conn.commit()
    except Exception:
        pass
    conn.close()

    return {'county': county, 'confidence': confidence, 'candidates': candidates}


def resolve_grouping(state, county):
    """Map state+county to a rate group name. Returns (group_name, is_fallback)."""
    st = (state or '').strip().upper()
    cn = _normalize_county(county) if county else None

    if cn:
        conn = _get_db_conn()
        c = conn.cursor()
        c.execute('SELECT group_name FROM county_group_map WHERE state = ? AND county_name = ?', (st, cn))
        row = c.fetchone()
        conn.close()
        if row:
            return row[0], False

    # Fallback
    fallback = _STATE_FALLBACK_GROUPS.get(st, 'All Other Locations')
    return fallback, True


def calculate_quote(sqft, rc_per_sf, loss_rents, group_rate_x100, aop_buydown):
    """Pure function: compute quote breakdown. Returns dict."""
    replacement_cost = sqft * rc_per_sf
    total_tiv = replacement_cost + loss_rents
    effective_rate_x100 = group_rate_x100 + (AOP_BUYDOWN_RATE_X100 if aop_buydown else 0)
    base_premium = total_tiv * (effective_rate_x100 / 100)
    taxes = base_premium * QUOTE_TAX_RATE
    total_premium = base_premium + taxes
    return {
        'replacement_cost': round(replacement_cost, 2),
        'total_tiv': round(total_tiv, 2),
        'effective_rate_x100': round(effective_rate_x100, 4),
        'base_premium': round(base_premium, 2),
        'taxes': round(taxes, 2),
        'total_premium': round(total_premium, 2),
    }


@app.route('/api/quotes/rates', methods=['GET'])
@require_auth
@require_role('admin')
def api_quotes_rates():
    """Return all rate groups."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT id, name, rate, rate_x100 FROM rate_groups ORDER BY name')
    groups = [{'id': r[0], 'name': r[1], 'rate': r[2], 'rate_x100': r[3]} for r in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'groups': groups})


@app.route('/api/quotes/property', methods=['POST'])
@require_auth
@require_role('admin')
def api_quotes_property():
    """Generate a property insurance quote."""
    data = request.json or {}
    sqft = data.get('sqft')
    loss_rents = data.get('loss_rents')
    city = (data.get('city') or '').strip()
    state = (data.get('state') or '').strip().upper()

    if not sqft or not loss_rents or not city or not state:
        return jsonify({'success': False, 'message': 'sqft, loss_rents, city, and state are required'}), 400
    try:
        sqft = int(sqft)
        loss_rents = float(loss_rents)
    except (ValueError, TypeError):
        return jsonify({'success': False, 'message': 'sqft must be integer, loss_rents must be numeric'}), 400

    rc_per_sf = float(data.get('rc_per_sf', 120))
    aop_buydown = bool(data.get('aop_buydown', False))
    street = data.get('street')
    zip_code = data.get('zip')
    county_override = data.get('county_override')

    warnings = []

    # Resolve county
    if county_override:
        county = _normalize_county(county_override)
        confidence = 'override'
        candidates = []
    else:
        result = resolve_county(city, state, street, zip_code)
        county = result['county']
        confidence = result['confidence']
        candidates = result['candidates']
        if not county:
            return jsonify({
                'success': False,
                'message': 'Could not determine county. Please provide a street address or select a county.',
                'candidates': candidates,
                'needs_county': True,
            }), 200
        if confidence == 'low' and len(candidates) > 1:
            return jsonify({
                'success': True,
                'needs_county_selection': True,
                'county': county,
                'candidates': candidates,
                'message': 'Multiple counties possible. Please confirm or select the correct county.',
            }), 200

    # Resolve grouping
    group_name, is_fallback = resolve_grouping(state, county)
    if is_fallback:
        warnings.append(f'No county mapping found for {county}, {state}. Using fallback: {group_name}')

    # Look up rate
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT rate_x100 FROM rate_groups WHERE name = ?', (group_name,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': f'Rate group "{group_name}" not found in database'}), 500
    group_rate_x100 = row[0]

    # Calculate
    quote = calculate_quote(sqft, rc_per_sf, loss_rents, group_rate_x100, aop_buydown)

    # Store
    quote_id = str(uuid.uuid4())
    ws = g.workspace_id if g.user else None
    uid = g.user['id'] if g.user else None
    c.execute('''INSERT INTO quote_requests
                 (id, workspace_id, created_by_user_id, sqft, rc_per_sf, loss_rents, city, state, county,
                  grouping_name, rate_x100, aop_buydown, replacement_cost, total_tiv, base_premium, taxes, total_premium, warnings)
                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
              (quote_id, ws, uid, sqft, rc_per_sf, loss_rents, city, state, county,
               group_name, quote['effective_rate_x100'], 1 if aop_buydown else 0,
               quote['replacement_cost'], quote['total_tiv'], quote['base_premium'],
               quote['taxes'], quote['total_premium'], json.dumps(warnings)))
    conn.commit()
    conn.close()

    return jsonify({
        'success': True,
        'quote_id': quote_id,
        'county': county,
        'county_confidence': confidence,
        'grouping': group_name,
        'group_rate_x100': group_rate_x100,
        **quote,
        'aop_buydown': aop_buydown,
        'warnings': warnings,
    })


@app.route('/api/quotes/history', methods=['GET'])
@require_auth
@require_role('admin')
def api_quotes_history():
    """Return recent quotes for the current user."""
    conn = _get_db_conn()
    c = conn.cursor()
    uid = g.user['id'] if g.user else None
    c.execute('''SELECT id, sqft, rc_per_sf, loss_rents, city, state, county, grouping_name,
                        rate_x100, aop_buydown, replacement_cost, total_tiv, base_premium, taxes,
                        total_premium, warnings, created_at
                 FROM quote_requests WHERE created_by_user_id = ? ORDER BY created_at DESC LIMIT 20''', (uid,))
    quotes = []
    for r in c.fetchall():
        quotes.append({
            'id': r[0], 'sqft': r[1], 'rc_per_sf': r[2], 'loss_rents': r[3],
            'city': r[4], 'state': r[5], 'county': r[6], 'grouping': r[7],
            'rate_x100': r[8], 'aop_buydown': bool(r[9]),
            'replacement_cost': r[10], 'total_tiv': r[11],
            'base_premium': r[12], 'taxes': r[13], 'total_premium': r[14],
            'warnings': json.loads(r[15]) if r[15] else [], 'created_at': r[16],
        })
    conn.close()
    return jsonify({'success': True, 'quotes': quotes})


# ---- Admin: Rate Groups management ----
@app.route('/api/admin/rate-groups', methods=['GET'])
@require_auth
@require_super_admin
def api_admin_rate_groups_list():
    """Super admin: list all rate groups."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT id, name, rate, rate_x100, created_at FROM rate_groups ORDER BY name')
    groups = [{'id': r[0], 'name': r[1], 'rate': r[2], 'rate_x100': r[3], 'created_at': r[4]} for r in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'groups': groups})


@app.route('/api/admin/rate-groups', methods=['POST'])
@require_auth
@require_super_admin
def api_admin_rate_groups_create():
    """Super admin: create a rate group."""
    data = request.json or {}
    name = (data.get('name') or '').strip()
    rate = data.get('rate')
    rate_x100 = data.get('rate_x100')
    if not name or rate is None or rate_x100 is None:
        return jsonify({'success': False, 'message': 'name, rate, and rate_x100 are required'}), 400
    conn = _get_db_conn()
    c = conn.cursor()
    gid = str(uuid.uuid4())
    try:
        c.execute('INSERT INTO rate_groups (id, name, rate, rate_x100) VALUES (?, ?, ?, ?)',
                  (gid, name, float(rate), float(rate_x100)))
        conn.commit()
    except _IntegrityError:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': 'Rate group name already exists'}), 400
    conn.close()
    return jsonify({'success': True, 'id': gid})


@app.route('/api/admin/rate-groups/<group_id>', methods=['PATCH'])
@require_auth
@require_super_admin
def api_admin_rate_groups_update(group_id):
    """Super admin: update a rate group."""
    data = request.json or {}
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT id FROM rate_groups WHERE id = ?', (group_id,))
    if not c.fetchone():
        conn.close()
        return jsonify({'success': False, 'message': 'Rate group not found'}), 404
    updates = []
    params = []
    for field in ('name', 'rate', 'rate_x100'):
        if field in data:
            updates.append(f'{field} = ?')
            params.append(data[field])
    if not updates:
        conn.close()
        return jsonify({'success': False, 'message': 'No fields to update'}), 400
    params.append(group_id)
    try:
        c.execute(f'UPDATE rate_groups SET {", ".join(updates)} WHERE id = ?', params)
        conn.commit()
    except _IntegrityError:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': 'Name conflict'}), 400
    conn.close()
    return jsonify({'success': True})


# ---- Admin: County mapping management ----
@app.route('/api/admin/county-mapping', methods=['GET'])
@require_auth
@require_super_admin
def api_admin_county_mapping_list():
    """Super admin: list county-to-group mappings."""
    state_filter = request.args.get('state', '').strip().upper()
    conn = _get_db_conn()
    c = conn.cursor()
    if state_filter:
        c.execute('SELECT id, state, county_name, group_name, created_at FROM county_group_map WHERE state = ? ORDER BY county_name', (state_filter,))
    else:
        c.execute('SELECT id, state, county_name, group_name, created_at FROM county_group_map ORDER BY state, county_name')
    mappings = [{'id': r[0], 'state': r[1], 'county_name': r[2], 'group_name': r[3], 'created_at': r[4]} for r in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'mappings': mappings})


@app.route('/api/admin/county-mapping', methods=['POST'])
@require_auth
@require_super_admin
def api_admin_county_mapping_create():
    """Super admin: create or update a county mapping."""
    data = request.json or {}
    state = (data.get('state') or '').strip().upper()
    county_name = _normalize_county(data.get('county_name') or '')
    group_name = (data.get('group_name') or '').strip()
    if not state or not county_name or not group_name:
        return jsonify({'success': False, 'message': 'state, county_name, and group_name are required'}), 400
    conn = _get_db_conn()
    c = conn.cursor()
    # Verify group exists
    c.execute('SELECT id FROM rate_groups WHERE name = ?', (group_name,))
    if not c.fetchone():
        conn.close()
        return jsonify({'success': False, 'message': f'Rate group "{group_name}" not found'}), 400
    mid = str(uuid.uuid4())
    c.execute('''INSERT INTO county_group_map (id, state, county_name, group_name) VALUES (?, ?, ?, ?)
                  ON CONFLICT (state, county_name) DO UPDATE SET group_name = EXCLUDED.group_name''',
              (mid, state, county_name, group_name))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/admin/county-mapping/<mapping_id>', methods=['DELETE'])
@require_auth
@require_super_admin
def api_admin_county_mapping_delete(mapping_id):
    """Super admin: delete a county mapping."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('DELETE FROM county_group_map WHERE id = ?', (mapping_id,))
    conn.commit()
    deleted = c.rowcount > 0
    conn.close()
    if not deleted:
        return jsonify({'success': False, 'message': 'Mapping not found'}), 404
    return jsonify({'success': True})


# ===================================================================
# CRM API ROUTES
# ===================================================================

def _log_lead_activity(cursor, lead_id, actor_user_id, action_type, old_value=None, new_value=None):
    """Write an activity log entry. Must be called within an open transaction."""
    cursor.execute(
        'INSERT INTO lead_activity (id, lead_id, actor_user_id, action_type, old_value, new_value) VALUES (?, ?, ?, ?, ?, ?)',
        (str(uuid.uuid4()), lead_id, actor_user_id, action_type,
         json.dumps(old_value) if old_value is not None else None,
         json.dumps(new_value) if new_value is not None else None)
    )

@app.route('/api/crm/lead/upsert', methods=['POST'])
@require_auth
@require_any_role('admin', 'producer')
def api_crm_upsert_lead():
    """Create or find a CRM lead by prospect_key."""
    if not g.user:
        return jsonify({'success': False, 'message': 'Auth required for CRM'}), 401
    data = request.json or {}
    prospect_key = data.get('prospect_key', '').strip()
    company_name = data.get('company_name', '').strip()
    website = data.get('website', '').strip() or None

    if not prospect_key or not company_name:
        return jsonify({'success': False, 'message': 'prospect_key and company_name are required'}), 400

    conn = _get_db_conn()
    c = conn.cursor()
    ws = g.workspace_id

    # Find or create crm_company
    c.execute('SELECT id FROM crm_companies WHERE workspace_id = ? AND prospect_key = ?', (ws, prospect_key))
    row = c.fetchone()
    if row:
        company_id = row[0]
    else:
        company_id = str(uuid.uuid4())
        c.execute('INSERT INTO crm_companies (id, workspace_id, prospect_key, company_name, website) VALUES (?, ?, ?, ?, ?)',
                  (company_id, ws, prospect_key, company_name, website))

    # Find or create crm_lead
    c.execute('SELECT id, status, owner_user_id, next_followup_at, priority FROM crm_leads WHERE workspace_id = ? AND company_id = ?', (ws, company_id))
    lead_row = c.fetchone()
    if lead_row:
        lead = {
            'id': lead_row[0], 'status': lead_row[1], 'owner_user_id': lead_row[2],
            'next_followup_at': lead_row[3], 'priority': lead_row[4], 'company_id': company_id,
            'company_name': company_name, 'prospect_key': prospect_key,
        }
        # If lead exists but unowned, auto-assign to current user (producer self-assignment)
        if not lead_row[2]:
            c.execute('UPDATE crm_leads SET owner_user_id = ? WHERE id = ?', (g.user['id'], lead_row[0]))
            lead['owner_user_id'] = g.user['id']
            _log_lead_activity(c, lead_row[0], g.user['id'], 'OWNER_ASSIGNED', None, g.user['id'])
    else:
        lead_id = str(uuid.uuid4())
        # Auto-assign owner to current user on create
        c.execute('INSERT INTO crm_leads (id, workspace_id, company_id, owner_user_id, status) VALUES (?, ?, ?, ?, ?)',
                  (lead_id, ws, company_id, g.user['id'], 'New'))
        lead = {
            'id': lead_id, 'status': 'New', 'owner_user_id': g.user['id'],
            'next_followup_at': None, 'priority': None, 'company_id': company_id,
            'company_name': company_name, 'prospect_key': prospect_key,
        }
        _log_lead_activity(c, lead_id, g.user['id'], 'SAVED')
        _log_lead_activity(c, lead_id, g.user['id'], 'OWNER_ASSIGNED', None, g.user['id'])

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'lead': lead})


@app.route('/api/crm/leads', methods=['GET'])
@require_auth
@require_any_role('admin', 'producer')
def api_crm_list_leads():
    """List CRM leads with optional filters: owner=me, status=, due=1"""
    if not g.user:
        return jsonify({'success': True, 'leads': []})
    ws = g.workspace_id
    conn = _get_db_conn()
    c = conn.cursor()

    query = '''
        SELECT l.id, l.status, l.owner_user_id, l.last_touch_at, l.next_followup_at, l.priority, l.created_at,
               co.company_name, co.prospect_key, co.website,
               u.name as owner_name,
               la.action_type as last_action_type, la.created_at as last_activity_at
        FROM crm_leads l
        JOIN crm_companies co ON l.company_id = co.id
        LEFT JOIN users u ON l.owner_user_id = u.id
        LEFT JOIN lead_activity la ON la.id = (
            SELECT la2.id FROM lead_activity la2 WHERE la2.lead_id = l.id ORDER BY la2.created_at DESC LIMIT 1
        )
        WHERE l.workspace_id = ?
    '''
    params = [ws]

    owner = request.args.get('owner')
    if owner == 'me':
        query += ' AND l.owner_user_id = ?'
        params.append(g.user['id'])
    elif owner == 'unassigned':
        query += ' AND l.owner_user_id IS NULL'

    status = request.args.get('status')
    if status:
        query += ' AND l.status = ?'
        params.append(status)

    due = request.args.get('due')
    if due == '1':
        query += ' AND l.next_followup_at IS NOT NULL AND l.next_followup_at <= ?'
        params.append(datetime.utcnow().isoformat())

    query += ' ORDER BY CASE WHEN l.next_followup_at IS NOT NULL THEN 0 ELSE 1 END, l.next_followup_at ASC, l.created_at DESC'

    c.execute(query, params)
    leads = []
    for r in c.fetchall():
        leads.append({
            'id': r[0], 'status': r[1], 'owner_user_id': r[2],
            'last_touch_at': r[3], 'next_followup_at': r[4], 'priority': r[5],
            'created_at': r[6], 'company_name': r[7], 'prospect_key': r[8],
            'website': r[9], 'owner_name': r[10],
            'last_action_type': r[11], 'last_activity_at': r[12],
        })
    conn.close()
    return jsonify({'success': True, 'leads': leads})


@app.route('/api/crm/leads/<lead_id>', methods=['PATCH'])
@require_auth
@require_any_role('admin', 'producer')
def api_crm_update_lead(lead_id):
    """Update a CRM lead (status, owner, followup, priority)."""
    if not g.user:
        return jsonify({'success': False, 'message': 'Auth required'}), 401
    data = request.json or {}
    conn = _get_db_conn()
    c = conn.cursor()

    # Verify lead belongs to workspace — fetch current values for activity log
    c.execute('SELECT id, owner_user_id, status, next_followup_at, priority FROM crm_leads WHERE id = ? AND workspace_id = ?',
              (lead_id, g.workspace_id))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': 'Lead not found'}), 404

    old_owner = row[1]
    old_status = row[2]
    old_followup = row[3]

    # Producers can only update their own leads; admins can update any
    if g.user['role'] != 'admin' and old_owner and old_owner != g.user['id']:
        conn.close()
        return jsonify({'success': False, 'message': 'Cannot update another user\'s lead'}), 403

    # Non-admins cannot change owner_user_id (server-side enforced)
    if 'owner_user_id' in data and g.user['role'] != 'admin':
        conn.close()
        return jsonify({'success': False, 'message': 'Only admins can reassign leads'}), 403

    updates = []
    params = []
    for field in ('status', 'owner_user_id', 'next_followup_at', 'priority'):
        if field in data:
            updates.append(f'{field} = ?')
            params.append(data[field])

    if updates:
        params.append(lead_id)
        c.execute(f'UPDATE crm_leads SET {", ".join(updates)} WHERE id = ?', params)

        # Log activity for each changed field
        if 'status' in data and data['status'] != old_status:
            _log_lead_activity(c, lead_id, g.user['id'], 'STATUS_CHANGED', old_status, data['status'])
        if 'next_followup_at' in data:
            new_followup = data['next_followup_at']
            if new_followup and new_followup != old_followup:
                _log_lead_activity(c, lead_id, g.user['id'], 'FOLLOWUP_SET', old_followup, new_followup)
            elif not new_followup and old_followup:
                _log_lead_activity(c, lead_id, g.user['id'], 'FOLLOWUP_CLEARED', old_followup, None)
        if 'owner_user_id' in data and data['owner_user_id'] != old_owner:
            if data['owner_user_id']:
                _log_lead_activity(c, lead_id, g.user['id'], 'OWNER_ASSIGNED', old_owner, data['owner_user_id'])
            else:
                _log_lead_activity(c, lead_id, g.user['id'], 'OWNER_CLEARED', old_owner, None)

        conn.commit()

    conn.close()
    return jsonify({'success': True})


@app.route('/api/crm/leads/<lead_id>/touchpoints', methods=['POST'])
@require_auth
@require_any_role('admin', 'producer')
def api_crm_add_touchpoint(lead_id):
    """Log a CRM touchpoint."""
    if not g.user:
        return jsonify({'success': False, 'message': 'Auth required'}), 401
    data = request.json or {}
    touch_type = data.get('type', '').strip()
    if not touch_type:
        return jsonify({'success': False, 'message': 'type is required'}), 400

    conn = _get_db_conn()
    c = conn.cursor()

    # Verify lead belongs to workspace and check ownership
    c.execute('SELECT id, owner_user_id FROM crm_leads WHERE id = ? AND workspace_id = ?', (lead_id, g.workspace_id))
    lead_row = c.fetchone()
    if not lead_row:
        conn.close()
        return jsonify({'success': False, 'message': 'Lead not found'}), 404

    # Non-admins can only log touchpoints on their own leads
    if g.user['role'] != 'admin' and lead_row[1] and lead_row[1] != g.user['id']:
        conn.close()
        return jsonify({'success': False, 'message': 'Cannot modify another user\'s lead'}), 403

    tp_id = str(uuid.uuid4())
    occurred_at = data.get('occurred_at', datetime.utcnow().isoformat())
    next_followup = data.get('next_followup_at')

    c.execute('''
        INSERT INTO crm_touchpoints (id, workspace_id, lead_id, user_id, type, outcome, notes, occurred_at, next_followup_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (tp_id, g.workspace_id, lead_id, g.user['id'], touch_type,
          data.get('outcome'), data.get('notes'), occurred_at, next_followup))

    # Update lead timestamps
    update_parts = ['last_touch_at = ?']
    update_params = [occurred_at]
    if next_followup:
        update_parts.append('next_followup_at = ?')
        update_params.append(next_followup)
    update_params.append(lead_id)
    c.execute(f'UPDATE crm_leads SET {", ".join(update_parts)} WHERE id = ?', update_params)

    # Log activity
    _log_lead_activity(c, lead_id, g.user['id'], 'NOTE_ADDED', None,
                       {'type': touch_type, 'outcome': data.get('outcome'), 'notes': data.get('notes')})
    if next_followup:
        _log_lead_activity(c, lead_id, g.user['id'], 'FOLLOWUP_SET', None, next_followup)

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'touchpoint_id': tp_id})


@app.route('/api/crm/leads/<lead_id>/touchpoints', methods=['GET'])
@require_auth
@require_any_role('admin', 'producer')
def api_crm_list_touchpoints(lead_id):
    """Get all touchpoints for a lead."""
    if not g.user:
        return jsonify({'success': True, 'touchpoints': []})
    conn = _get_db_conn()
    c = conn.cursor()

    c.execute('''
        SELECT tp.id, tp.type, tp.outcome, tp.notes, tp.occurred_at, tp.next_followup_at, u.name
        FROM crm_touchpoints tp
        JOIN users u ON tp.user_id = u.id
        WHERE tp.lead_id = ? AND tp.workspace_id = ?
        ORDER BY tp.occurred_at DESC
    ''', (lead_id, g.workspace_id))

    touchpoints = []
    for r in c.fetchall():
        touchpoints.append({
            'id': r[0], 'type': r[1], 'outcome': r[2], 'notes': r[3],
            'occurred_at': r[4], 'next_followup_at': r[5], 'user_name': r[6],
        })
    conn.close()
    return jsonify({'success': True, 'touchpoints': touchpoints})


@app.route('/api/crm/leads/<lead_id>/assign', methods=['PATCH'])
@require_auth
@require_role('admin')
def api_crm_assign_lead(lead_id):
    """Admin only: reassign a lead to a different user or unassign."""
    if not g.user:
        return jsonify({'success': False, 'message': 'Auth required'}), 401
    data = request.json or {}
    new_owner_id = data.get('owner_id')  # null = unassign

    conn = _get_db_conn()
    c = conn.cursor()

    # Verify lead belongs to workspace and get current owner
    c.execute('SELECT id, owner_user_id FROM crm_leads WHERE id = ? AND workspace_id = ?', (lead_id, g.workspace_id))
    lead_row = c.fetchone()
    if not lead_row:
        conn.close()
        return jsonify({'success': False, 'message': 'Lead not found'}), 404
    old_owner_id = lead_row[1]

    # If assigning to a user, verify that user exists in workspace
    if new_owner_id:
        c.execute('SELECT id, name FROM users WHERE id = ? AND workspace_id = ?', (new_owner_id, g.workspace_id))
        target = c.fetchone()
        if not target:
            conn.close()
            return jsonify({'success': False, 'message': 'Target user not found'}), 400
        owner_name = target[1]
    else:
        owner_name = None

    c.execute('UPDATE crm_leads SET owner_user_id = ? WHERE id = ?', (new_owner_id, lead_id))

    # Log activity
    if new_owner_id != old_owner_id:
        action = 'OWNER_ASSIGNED' if new_owner_id else 'OWNER_CLEARED'
        _log_lead_activity(c, lead_id, g.user['id'], action, old_owner_id, new_owner_id)

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'owner_user_id': new_owner_id, 'owner_name': owner_name})


@app.route('/api/crm/workspace-users', methods=['GET'])
@require_auth
@require_any_role('admin', 'producer')
def api_crm_workspace_users():
    """List users in the current workspace (for assignment dropdowns)."""
    if not g.user:
        return jsonify({'success': True, 'users': []})
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT id, name, role FROM users WHERE workspace_id = ? AND is_disabled = FALSE ORDER BY name',
              (g.workspace_id,))
    users = [{'id': r[0], 'name': r[1], 'role': r[2]} for r in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'users': users})


@app.route('/api/crm/leads/<lead_id>/activity', methods=['GET'])
@require_auth
@require_any_role('admin', 'producer')
def api_crm_lead_activity(lead_id):
    """Get activity timeline for a lead. Admin sees any; producer only own leads."""
    if not g.user:
        return jsonify({'success': False, 'message': 'Auth required'}), 401
    conn = _get_db_conn()
    c = conn.cursor()

    # Verify lead exists in workspace and check ownership
    c.execute('SELECT id, owner_user_id FROM crm_leads WHERE id = ? AND workspace_id = ?', (lead_id, g.workspace_id))
    lead_row = c.fetchone()
    if not lead_row:
        conn.close()
        return jsonify({'success': False, 'message': 'Lead not found'}), 404
    if g.user['role'] != 'admin' and lead_row[1] and lead_row[1] != g.user['id']:
        conn.close()
        return jsonify({'success': False, 'message': 'Access denied'}), 403

    c.execute('''
        SELECT a.id, a.action_type, a.old_value, a.new_value, a.created_at,
               u.id, u.name, u.email, u.role
        FROM lead_activity a
        JOIN users u ON a.actor_user_id = u.id
        WHERE a.lead_id = ?
        ORDER BY a.created_at DESC
    ''', (lead_id,))

    activities = []
    for r in c.fetchall():
        old_val = json.loads(r[2]) if r[2] else None
        new_val = json.loads(r[3]) if r[3] else None
        activities.append({
            'id': r[0], 'action_type': r[1], 'old_value': old_val,
            'new_value': new_val, 'created_at': r[4],
            'actor': {'id': r[5], 'name': r[6], 'email': r[7], 'role': r[8]},
        })
    conn.close()
    return jsonify({'success': True, 'activities': activities})


@app.route('/api/activity', methods=['GET'])
@require_auth
@require_role('admin')
def api_admin_activity_overview():
    """Admin only: query activity logs across all leads with filters."""
    if not g.user:
        return jsonify({'success': False, 'message': 'Auth required'}), 401

    owner_id = request.args.get('owner_id')
    actor_id = request.args.get('actor_id')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    limit_val = min(int(request.args.get('limit', 50)), 200)

    conn = _get_db_conn()
    c = conn.cursor()

    query = '''
        SELECT a.id, a.lead_id, a.action_type, a.old_value, a.new_value, a.created_at,
               u.id, u.name, u.email, u.role,
               co.company_name, co.prospect_key
        FROM lead_activity a
        JOIN users u ON a.actor_user_id = u.id
        JOIN crm_leads l ON a.lead_id = l.id
        JOIN crm_companies co ON l.company_id = co.id
        WHERE l.workspace_id = ?
    '''
    params = [g.workspace_id]

    if owner_id:
        query += ' AND l.owner_user_id = ?'
        params.append(owner_id)
    if actor_id:
        query += ' AND a.actor_user_id = ?'
        params.append(actor_id)
    if date_from:
        query += ' AND a.created_at >= ?'
        params.append(date_from)
    if date_to:
        query += ' AND a.created_at <= ?'
        params.append(date_to)

    query += ' ORDER BY a.created_at DESC LIMIT ?'
    params.append(limit_val)

    c.execute(query, params)
    activities = []
    for r in c.fetchall():
        old_val = json.loads(r[3]) if r[3] else None
        new_val = json.loads(r[4]) if r[4] else None
        activities.append({
            'id': r[0], 'lead_id': r[1], 'action_type': r[2],
            'old_value': old_val, 'new_value': new_val, 'created_at': r[5],
            'actor': {'id': r[6], 'name': r[7], 'email': r[8], 'role': r[9]},
            'company_name': r[10], 'prospect_key': r[11],
        })
    conn.close()
    return jsonify({'success': True, 'activities': activities})


@app.route('/api/crm/leads/bulk-status', methods=['GET'])
@require_auth
@require_any_role('admin', 'producer')
def api_crm_bulk_status():
    """Get CRM status for multiple prospect_keys at once (for card overlays)."""
    if not g.user:
        return jsonify({'success': True, 'statuses': {}})
    keys_param = request.args.get('keys', '')
    if not keys_param:
        return jsonify({'success': True, 'statuses': {}})
    keys = [k.strip() for k in keys_param.split(',') if k.strip()]
    if not keys:
        return jsonify({'success': True, 'statuses': {}})

    conn = _get_db_conn()
    c = conn.cursor()
    placeholders = ','.join('?' * len(keys))
    c.execute(f'''
        SELECT co.prospect_key, l.id, l.status, l.owner_user_id, l.next_followup_at, u.name as owner_name
        FROM crm_leads l
        JOIN crm_companies co ON l.company_id = co.id
        LEFT JOIN users u ON l.owner_user_id = u.id
        WHERE l.workspace_id = ? AND co.prospect_key IN ({placeholders})
    ''', [g.workspace_id] + keys)

    statuses = {}
    for r in c.fetchall():
        statuses[r[0]] = {
            'lead_id': r[1], 'status': r[2], 'owner_user_id': r[3],
            'next_followup_at': r[4], 'owner_name': r[5],
        }
    conn.close()
    return jsonify({'success': True, 'statuses': statuses})


def search_btr_prospects(city="Texas", limit=10):
    """
    Use SerpAPI (Stage A) + Claude extraction (Stage B) to find BTR prospects.
    Returns (prospects_list, error_message) tuple.
    """
    from serpapi_client import cached_serpapi_search, SerpAPIError

    # Check API keys first
    api_key = os.getenv('ANTHROPIC_API_KEY')
    if not api_key or api_key == 'your_anthropic_api_key_here':
        return [], "ANTHROPIC_API_KEY is not set. Add it to your .env file or Railway environment variables."

    serp_key = os.getenv('SERPAPI_API_KEY')
    use_serpapi = bool(serp_key)

    # Check in-memory cache first
    cache_key = f"{city.lower().strip()}|{limit}"
    if cache_key in _search_cache:
        cached = _search_cache[cache_key]
        age = (datetime.now() - cached['timestamp']).total_seconds()
        if age < CACHE_TTL_SECONDS:
            print(f"Returning cached results for {city} ({int(age)}s old)")
            return cached['prospects'], None
        else:
            del _search_cache[cache_key]

    try:
        existing = get_existing_companies(city)
        exclude_clause = ""
        if existing:
            names = ", ".join(existing[:20])
            exclude_clause = f"\n\nIMPORTANT: I already have these companies in my database, so DO NOT include them. Find DIFFERENT companies:\n{names}\n"

        today = datetime.now().strftime('%B %d, %Y')
        ninety_days_ago = (datetime.now() - timedelta(days=90)).strftime('%B %Y')
        ask_count = min(limit, 10)

        if use_serpapi:
            # --- Path A: SerpAPI candidate retrieval + Claude extraction ---
            from serpapi_client import cached_serpapi_search, SerpAPIError

            queries = [
                f'site:bisnow.com ("build to rent" OR BTR) "{city}"',
                f'site:multihousingnews.com ("build to rent" OR BTR) "{city}"',
                f'site:bizjournals.com ("build to rent" OR BTR) "{city}"',
                f'("build to rent" OR "single family rental community") "{city}" (acquires OR acquisition OR sells OR sale OR groundbreaking OR "under construction")',
            ]

            all_candidates = []
            seen_links = set()

            print(f"Fetching SerpAPI candidates for {city}...")

            for query in queries:
                try:
                    results = cached_serpapi_search(
                        query, num=5, feature='prospect', city=city, state=''
                    )
                    for r in results:
                        link = r.get('link', '')
                        if link and link not in seen_links:
                            seen_links.add(link)
                            all_candidates.append(r)
                except SerpAPIError as e:
                    return [], f"Search temporarily rate limited: {e}"
                except Exception as e:
                    print(f"SerpAPI query error: {e}")

                if len(all_candidates) >= 20:
                    break

            if not all_candidates:
                return [], "No search results found for this area. Try a different city or state."

            print(f"SerpAPI returned {len(all_candidates)} candidates for {city}")

            candidates_json = json.dumps([{
                'title': c['title'],
                'url': c['link'],
                'snippet': c.get('snippet', ''),
                'source': c.get('source', ''),
                'date': c.get('date', ''),
            } for c in all_candidates[:20]], indent=2)

            extraction_prompt = f"""You are a real estate intelligence researcher. Today's date is {today}.

I have search results about Build-to-Rent (BTR) / Single-Family Rental (SFR) activity in {city}.
Analyze these search results and extract BTR developer prospects.

SEARCH RESULTS:
{candidates_json}

FOCUS ON RECENT ACTIVITY from the last 90 days (since {ninety_days_ago}). Look for groundbreakings, land acquisitions, construction starts, capital raises, new projects.
{exclude_clause}
Find up to {ask_count} companies. For each extract:
- Company name
- CEO/key executive name (if mentioned)
- LinkedIn profile (if mentioned)
- City location
- Recent project details (include dates when available)
- Active signals (financing, construction, sales, expansion)
- Total Investment Value estimate (if mentioned)
- Why to call them NOW (specific trigger from the search result)
- Score 0-100 (the total of the four sub-scores below)
- Score breakdown (must sum to the total score):
  - capital_event (0-40): recapitalizations, credit facilities, JV capital, acquisitions, institutional partners
  - construction_stage (0-25): under construction, groundbreaking, permitting, first deliveries
  - expansion_velocity (0-20): multi-market growth, pipeline mentions, "expanding into", multiple projects
  - freshness (0-15): activity within last 14 days=15, last 30 days=10, last 90 days=5
- Score explanation: 2-4 bullet points explaining why it scored that way
- Insurance triggers: 0-4 labels from this list that apply:
  "Builder's Risk → Property conversion", "New lender covenants / insurance requirements",
  "Portfolio scale / blanket limits", "New state expansion", "JV / institutional capital event",
  "Refinance window / debt facility", "Lease-up stabilization shift"
- unit_band: classify the operator's typical project size as one of: "<40", "40-150", "150-400", "400-1000", "1000+"
- active_project_count_estimate: estimated number of active projects (integer, best guess from context)
- markets_active_estimate: estimated number of metros/markets they operate in (integer, best guess)
- swim_lane_fit_score (0-100): Start at 50, then apply modifiers:
  +25 if unit_band "40-150", +20 if "150-400", -15 if "<40", -10 if "400-1000", -25 if "1000+"
  +15 if active_project_count 2-5, +5 if 1, -10 if >10
  +10 if markets_active 2-4, +5 if 1, -5 if 5+
  Clamp result to 0-100
- competitive_difficulty: "Low" (niche/emerging operator), "Medium" (established regional), "High" (national/institutional)

Return ONLY valid JSON in this exact format:

{{
  "prospects": [
    {{
      "company": "Company Name",
      "executive": "Executive Name",
      "title": "CEO",
      "linkedin": "linkedin.com/in/profile",
      "city": "City",
      "state": "TX",
      "score": 85,
      "score_breakdown": {{
        "capital_event": 35,
        "construction_stage": 20,
        "expansion_velocity": 18,
        "freshness": 12
      }},
      "score_explanation": [
        "Recent $200M credit facility with institutional lender",
        "3 communities under construction across TX",
        "Expanding into AZ and FL markets",
        "Activity reported within last 2 weeks"
      ],
      "insurance_triggers": [
        "Builder's Risk → Property conversion",
        "JV / institutional capital event"
      ],
      "unit_band": "40-150",
      "active_project_count_estimate": 3,
      "markets_active_estimate": 2,
      "swim_lane_fit_score": 85,
      "competitive_difficulty": "Low",
      "tiv": "$50M-200M",
      "units": "200-500 units",
      "projectName": "Project Name",
      "projectStatus": "Under construction / Pre-leasing / Recently opened",
      "signals": ["Signal 1", "Signal 2", "Signal 3"],
      "whyNow": "Why call this prospect now"
    }}
  ]
}}

CRITICAL: Return ONLY the JSON object, no other text. Only include companies clearly related to BTR/SFR development."""

            print(f"Calling Claude API for {city} extraction...")

            message = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4096,
                messages=[{"role": "user", "content": extraction_prompt}]
            )

        else:
            # --- Path B: Claude web_search fallback (no SerpAPI key) ---
            print(f"No SERPAPI_API_KEY set, using Claude web_search for {city}...")

            search_prompt = f"""You are a real estate intelligence researcher. Today's date is {today}.

Search for Build-to-Rent (BTR) and Single-Family Rental (SFR) development activity in {city}.

Search for recent news about:
- BTR groundbreakings, land acquisitions, construction starts in {city}
- Single family rental community developments in {city}
- Build to rent developers active in {city}

FOCUS ON RECENT ACTIVITY from the last 90 days (since {ninety_days_ago}).
{exclude_clause}
Find up to {ask_count} companies actively developing BTR/SFR projects. For each extract:
- Company name
- CEO/key executive name (if mentioned)
- LinkedIn profile (if mentioned)
- City location
- Recent project details (include dates when available)
- Active signals (financing, construction, sales, expansion)
- Total Investment Value estimate (if mentioned)
- Why to call them NOW (specific trigger from the search result)
- Score 0-100 (the total of the four sub-scores below)
- Score breakdown (must sum to the total score):
  - capital_event (0-40): recapitalizations, credit facilities, JV capital, acquisitions, institutional partners
  - construction_stage (0-25): under construction, groundbreaking, permitting, first deliveries
  - expansion_velocity (0-20): multi-market growth, pipeline mentions, "expanding into", multiple projects
  - freshness (0-15): activity within last 14 days=15, last 30 days=10, last 90 days=5
- Score explanation: 2-4 bullet points explaining why it scored that way
- Insurance triggers: 0-4 labels from this list that apply:
  "Builder's Risk → Property conversion", "New lender covenants / insurance requirements",
  "Portfolio scale / blanket limits", "New state expansion", "JV / institutional capital event",
  "Refinance window / debt facility", "Lease-up stabilization shift"
- unit_band: classify the operator's typical project size as one of: "<40", "40-150", "150-400", "400-1000", "1000+"
- active_project_count_estimate: estimated number of active projects (integer, best guess from context)
- markets_active_estimate: estimated number of metros/markets they operate in (integer, best guess)
- swim_lane_fit_score (0-100): Start at 50, then apply modifiers:
  +25 if unit_band "40-150", +20 if "150-400", -15 if "<40", -10 if "400-1000", -25 if "1000+"
  +15 if active_project_count 2-5, +5 if 1, -10 if >10
  +10 if markets_active 2-4, +5 if 1, -5 if 5+
  Clamp result to 0-100
- competitive_difficulty: "Low" (niche/emerging operator), "Medium" (established regional), "High" (national/institutional)

Return ONLY valid JSON in this exact format:

{{
  "prospects": [
    {{
      "company": "Company Name",
      "executive": "Executive Name",
      "title": "CEO",
      "linkedin": "linkedin.com/in/profile",
      "city": "City",
      "state": "TX",
      "score": 85,
      "score_breakdown": {{
        "capital_event": 35,
        "construction_stage": 20,
        "expansion_velocity": 18,
        "freshness": 12
      }},
      "score_explanation": [
        "Recent $200M credit facility with institutional lender",
        "3 communities under construction across TX",
        "Expanding into AZ and FL markets",
        "Activity reported within last 2 weeks"
      ],
      "insurance_triggers": [
        "Builder's Risk → Property conversion",
        "JV / institutional capital event"
      ],
      "unit_band": "40-150",
      "active_project_count_estimate": 3,
      "markets_active_estimate": 2,
      "swim_lane_fit_score": 85,
      "competitive_difficulty": "Low",
      "tiv": "$50M-200M",
      "units": "200-500 units",
      "projectName": "Project Name",
      "projectStatus": "Under construction / Pre-leasing / Recently opened",
      "signals": ["Signal 1", "Signal 2", "Signal 3"],
      "whyNow": "Why call this prospect now"
    }}
  ]
}}

CRITICAL: Return ONLY the JSON object, no other text. Only include companies clearly related to BTR/SFR development."""

            print(f"Calling Claude API with web_search for {city}...")

            max_retries = 3
            message = None
            for attempt in range(max_retries + 1):
                try:
                    message = client.messages.create(
                        model="claude-sonnet-4-20250514",
                        max_tokens=4096,
                        tools=[
                            {
                                "type": "web_search_20250305",
                                "name": "web_search",
                                "max_uses": 5
                            }
                        ],
                        messages=[
                            {
                                "role": "user",
                                "content": search_prompt
                            }
                        ]
                    )
                    break
                except anthropic.RateLimitError:
                    if attempt < max_retries:
                        wait_time = 2 ** (attempt + 1)
                        print(f"Rate limited (attempt {attempt + 1}/{max_retries + 1}), waiting {wait_time}s...")
                        time.sleep(wait_time)
                    else:
                        return [], "API rate limit reached after retries. Please wait 1-2 minutes and try again."

        response_text = ""
        for block in message.content:
            if block.type == "text":
                response_text += block.text

        print(f"Claude response length: {len(response_text)} chars")

        if not response_text.strip():
            return [], "Claude returned an empty response. Try again."

        # Parse JSON from response
        json_start = response_text.find('{"prospects"')
        if json_start == -1:
            json_start = response_text.find('{  "prospects"')
        if json_start == -1:
            json_start = response_text.find('{\n')

        if json_start != -1:
            brace_count = 0
            json_end = json_start
            for i in range(json_start, len(response_text)):
                if response_text[i] == '{':
                    brace_count += 1
                elif response_text[i] == '}':
                    brace_count -= 1
                    if brace_count == 0:
                        json_end = i + 1
                        break

            json_str = response_text[json_start:json_end]
            try:
                data = json.loads(json_str)
                prospects = data.get('prospects', [])
                if prospects:
                    print(f"Successfully parsed {len(prospects)} prospects")
                    _search_cache[cache_key] = {
                        'prospects': prospects,
                        'timestamp': datetime.now()
                    }
                    return prospects, None
                else:
                    return [], "No BTR prospects found in search results. Try a different city or state."
            except json.JSONDecodeError as e:
                print(f"JSON parse error: {e}")
                return [], "Failed to parse AI response. Try searching again."
        else:
            print(f"No JSON found in response: {response_text[:500]}")
            return [], "AI response did not contain prospect data. Try searching again."

    except anthropic.AuthenticationError:
        return [], "Invalid ANTHROPIC_API_KEY. Check your API key in .env or Railway variables."
    except anthropic.APIConnectionError:
        return [], "Cannot connect to Claude API. Check your internet connection."
    except anthropic.RateLimitError:
        return [], "Claude API rate limit reached. Please wait 1-2 minutes and try again."
    except Exception as e:
        print(f"Search error: {str(e)}")
        return [], f"Search failed: {str(e)}"

def get_existing_companies(city=None):
    """Get list of company names already in the database, optionally filtered by city"""
    conn = _get_db_conn()
    c = conn.cursor()
    if city:
        c.execute('SELECT DISTINCT company FROM prospects WHERE LOWER(city) LIKE ? OR LOWER(state) LIKE ?',
                  (f'%{city.lower()}%', f'%{city.lower()}%'))
    else:
        c.execute('SELECT DISTINCT company FROM prospects')
    companies = [row[0] for row in c.fetchall()]
    conn.close()
    return companies

def save_prospects_to_db(prospects):
    """Save prospects to SQLite database"""
    conn = _get_db_conn()
    c = conn.cursor()
    
    saved_count = 0
    for p in prospects:
        try:
            c.execute('''
                INSERT OR IGNORE INTO prospects 
                (company, executive, title, linkedin, city, state, score, tiv, units, 
                 project_name, project_status, signals, why_now)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                p.get('company'),
                p.get('executive'),
                p.get('title'),
                p.get('linkedin'),
                p.get('city'),
                p.get('state', 'TX'),
                p.get('score'),
                p.get('tiv'),
                p.get('units'),
                p.get('projectName'),
                p.get('projectStatus'),
                json.dumps(p.get('signals', [])),
                p.get('whyNow')
            ))
            if c.rowcount > 0:
                saved_count += 1
        except Exception as e:
            print(f"Error saving prospect {p.get('company')}: {str(e)}")
            continue
    
    conn.commit()
    conn.close()
    return saved_count

def get_all_prospects_from_db():
    """Retrieve all prospects from database"""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT * FROM prospects ORDER BY score DESC, created_at DESC')
    
    prospects = []
    for row in c.fetchall():
        prospects.append({
            'id': row[0],
            'company': row[1],
            'executive': row[2],
            'title': row[3],
            'linkedin': row[4],
            'email': row[5],
            'phone': row[6],
            'city': row[7],
            'state': row[8],
            'score': row[9],
            'tiv': row[10],
            'units': row[11],
            'projectName': row[12],
            'projectStatus': row[13],
            'signals': json.loads(row[14]) if row[14] else [],
            'whyNow': row[15],
            'createdAt': row[16]
        })
    
    conn.close()
    return prospects

MASTER_CSV_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'btr-prospects-master.csv')

def generate_master_csv():
    """Regenerate the master CSV spreadsheet from all database prospects"""
    prospects = get_all_prospects_from_db()

    headers = [
        'Company', 'Executive', 'Title', 'LinkedIn', 'City', 'State',
        'Score', 'TIV', 'Units', 'Project Name', 'Project Status',
        'Signals', 'Why Call Now', 'Date Found'
    ]

    with open(MASTER_CSV_PATH, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for p in prospects:
            signals = ', '.join(p.get('signals', [])) if isinstance(p.get('signals'), list) else p.get('signals', '')
            writer.writerow([
                p.get('company', ''),
                p.get('executive', ''),
                p.get('title', ''),
                p.get('linkedin', ''),
                p.get('city', ''),
                p.get('state', ''),
                p.get('score', ''),
                p.get('tiv', ''),
                p.get('units', ''),
                p.get('projectName', ''),
                p.get('projectStatus', ''),
                signals,
                p.get('whyNow', ''),
                p.get('createdAt', '')
            ])

    print(f"Master CSV updated: {len(prospects)} prospects written to {MASTER_CSV_PATH}")
    return len(prospects)

# Generate master CSV on startup (in background to avoid blocking)
threading.Thread(target=generate_master_csv, daemon=True).start()

# --- Daily Discovery Engine ---

def run_daily_discovery(is_scheduled=False):
    """Main daily discovery orchestrator — event signal scanner.
    Uses discovery_engine.py with adapter architecture to fetch from multiple sources."""
    global _discovery_running
    _discovery_running = True
    try:
        config = DISCOVERY_CONFIG
        from discovery_engine import run_discovery_job

        print(f"[Discovery] Starting signal scan at {datetime.now().isoformat()}")

        results, digest, total_new, adapter_stats = run_discovery_job(config, is_scheduled=is_scheduled)

        # Save run record
        conn = _get_db_conn()
        c = conn.cursor()
        c.execute('''
            INSERT INTO discovery_runs (results_json, digest_text, city_count, total_new, status, adapter_stats)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            json.dumps(results),
            digest,
            len(config['cities']),
            total_new,
            'completed',
            json.dumps(adapter_stats)
        ))
        conn.commit()
        conn.close()

        # Deliver via webhook if configured
        if config['delivery_method'] == 'webhook' and config['webhook_url']:
            try:
                import urllib.request
                payload = json.dumps({
                    'type': 'daily_discovery',
                    'run_at': datetime.now().isoformat(),
                    'total_new': total_new,
                    'results': results,
                    'digest': digest,
                    'adapter_stats': adapter_stats
                }).encode('utf-8')
                req = urllib.request.Request(
                    config['webhook_url'],
                    data=payload,
                    headers={'Content-Type': 'application/json'}
                )
                urllib.request.urlopen(req, timeout=10)
                print(f"[Discovery] Webhook delivered to {config['webhook_url']}")
            except Exception as e:
                print(f"[Discovery] Webhook delivery failed: {e}")

        print(f"[Discovery] Run complete. {total_new} new signals across {len(config['cities'])} cities.")
        return results, digest

    except Exception as e:
        print(f"[Discovery] Run failed: {e}")
        traceback.print_exc()
        return {}, ""
    finally:
        _discovery_running = False

# ===================================================================
# BROKER API ROUTES (Deal Board)
# ===================================================================

@app.route('/api/broker/saved', methods=['GET'])
@require_auth
@require_any_role('admin', 'broker')
def api_broker_saved_list():
    """List broker's saved prospects (Deal Board)."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT id, prospect_id, company, notes, status, created_at, updated_at FROM broker_saved WHERE user_id = ? ORDER BY created_at DESC', (g.user['id'],))
    rows = c.fetchall()
    conn.close()
    items = []
    for r in rows:
        created = r[5]
        updated = r[6]
        if isinstance(created, datetime):
            created = created.isoformat()
        if isinstance(updated, datetime):
            updated = updated.isoformat()
        items.append({
            'id': r[0], 'prospect_id': r[1], 'company': r[2],
            'notes': r[3], 'status': r[4],
            'created_at': str(created or ''), 'updated_at': str(updated or ''),
        })
    return jsonify({'success': True, 'items': items})


@app.route('/api/broker/saved', methods=['POST'])
@require_auth
@require_any_role('admin', 'broker')
def api_broker_saved_create():
    """Save a prospect to broker's Deal Board."""
    data = request.json or {}
    prospect_id = data.get('prospect_id')
    company = data.get('company', '').strip()
    notes = data.get('notes', '').strip()

    if not company and not prospect_id:
        return jsonify({'success': False, 'message': 'prospect_id or company required'}), 400

    # If prospect_id given, look up company name
    if prospect_id and not company:
        conn = _get_db_conn()
        c = conn.cursor()
        c.execute('SELECT company FROM prospects WHERE id = ?', (prospect_id,))
        row = c.fetchone()
        conn.close()
        company = row[0] if row else 'Unknown'

    item_id = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()
    conn = _get_db_conn()
    c = conn.cursor()
    try:
        c.execute('INSERT INTO broker_saved (id, user_id, prospect_id, company, notes, status, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                  (item_id, g.user['id'], prospect_id, company, notes, 'saved', now, now))
        conn.commit()
    except _IntegrityError:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': 'Already saved'}), 409
    conn.close()
    return jsonify({'success': True, 'item': {
        'id': item_id, 'prospect_id': prospect_id, 'company': company,
        'notes': notes, 'status': 'saved', 'created_at': now, 'updated_at': now,
    }})


@app.route('/api/broker/saved/<item_id>', methods=['PATCH'])
@require_auth
@require_any_role('admin', 'broker')
def api_broker_saved_update(item_id):
    """Update a saved Deal Board item (notes, status)."""
    data = request.json or {}
    conn = _get_db_conn()
    c = conn.cursor()
    # Verify ownership
    c.execute('SELECT id FROM broker_saved WHERE id = ? AND user_id = ?', (item_id, g.user['id']))
    if not c.fetchone():
        conn.close()
        return jsonify({'success': False, 'message': 'Not found'}), 404

    updates = []
    params = []
    if 'notes' in data:
        updates.append('notes = ?')
        params.append(data['notes'])
    if 'status' in data:
        updates.append('status = ?')
        params.append(data['status'])
    if updates:
        updates.append('updated_at = ?')
        params.append(datetime.utcnow().isoformat())
        params.append(item_id)
        c.execute(f'UPDATE broker_saved SET {", ".join(updates)} WHERE id = ?', params)
        conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/broker/saved/<item_id>', methods=['DELETE'])
@require_auth
@require_any_role('admin', 'broker')
def api_broker_saved_delete(item_id):
    """Remove item from Deal Board."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('DELETE FROM broker_saved WHERE id = ? AND user_id = ?', (item_id, g.user['id']))
    deleted = c.rowcount
    conn.commit()
    conn.close()
    if not deleted:
        return jsonify({'success': False, 'message': 'Not found'}), 404
    return jsonify({'success': True})


@app.route('/api/broker/export-csv', methods=['GET'])
@require_auth
@require_any_role('admin', 'broker')
def api_broker_export_csv():
    """Export broker's saved Deal Board items as CSV."""
    import csv
    import io

    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('''SELECT bs.company, bs.notes, bs.status, bs.created_at,
                        p.executive, p.title, p.city, p.state, p.score, p.tiv, p.units,
                        p.signals, p.why_now, p.email, p.phone
                 FROM broker_saved bs
                 LEFT JOIN prospects p ON bs.prospect_id = p.id
                 WHERE bs.user_id = ?
                 ORDER BY bs.created_at DESC''', (g.user['id'],))
    rows = c.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Company', 'Notes', 'Status', 'Saved At',
                     'Executive', 'Title', 'City', 'State', 'Score', 'TIV', 'Units',
                     'Signals', 'Why Now', 'Email', 'Phone'])
    for r in rows:
        writer.writerow([str(v or '') for v in r])

    resp = make_response(output.getvalue())
    resp.headers['Content-Type'] = 'text/csv'
    resp.headers['Content-Disposition'] = 'attachment; filename=deal-board-export.csv'
    return resp


# ===================================================================
# UNDERWRITING SHEET API (Admin Only)
# ===================================================================

from underwriting_columns import UNDERWRITING_COLUMNS, COLUMN_KEYS, HEADER_MAP, COLUMN_TYPES

@app.route('/api/underwriting/communities', methods=['POST'])
@require_auth
@require_role('admin')
def api_uw_create_community():
    """Create a new underwriting community + initial row."""
    data = request.json or {}
    row_data = data.get('row', {})
    location_name = row_data.get('location_name') or data.get('location_name', '')
    community_key = data.get('community_key') or f"{location_name}_{uuid.uuid4().hex[:6]}"

    if not location_name:
        return jsonify({'ok': False, 'error': 'location_name is required'}), 400

    community_id = str(uuid.uuid4())
    row_id = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()

    conn = _get_db_conn()
    c = conn.cursor()
    try:
        c.execute('INSERT INTO underwriting_communities (id, community_key, location_name, created_at, updated_at) VALUES (?, ?, ?, ?, ?)',
                  (community_id, community_key, location_name, now, now))

        col_names = ', '.join(COLUMN_KEYS)
        placeholders = ', '.join(['?'] * len(COLUMN_KEYS))
        values = [str(row_data.get(k, '') or '') for k in COLUMN_KEYS]

        c.execute(f'INSERT INTO underwriting_rows (id, community_id, row_version, created_at, {col_names}) VALUES (?, ?, ?, ?, {placeholders})',
                  [row_id, community_id, 1, now] + values)
        conn.commit()
    except _IntegrityError:
        conn.rollback()
        conn.close()
        return jsonify({'ok': False, 'error': 'Community key already exists'}), 409
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'ok': False, 'error': str(e)}), 500
    conn.close()
    return jsonify({'ok': True, 'community_id': community_id, 'row_id': row_id})


@app.route('/api/underwriting/communities/<community_id>/add-units', methods=['POST'])
@require_auth
@require_role('admin')
def api_uw_add_units(community_id):
    """Add a new row version (add-on phase) to an existing community."""
    data = request.json or {}
    row_data = data.get('row', {})
    base_row_id = data.get('base_row_id')

    conn = _get_db_conn()
    c = conn.cursor()

    # Load base row
    if base_row_id:
        c.execute(f'SELECT row_version, {", ".join(COLUMN_KEYS)} FROM underwriting_rows WHERE id = ? AND community_id = ?', (base_row_id, community_id))
    else:
        c.execute(f'SELECT row_version, {", ".join(COLUMN_KEYS)} FROM underwriting_rows WHERE community_id = ? ORDER BY row_version DESC LIMIT 1', (community_id,))
    base = c.fetchone()
    if not base:
        conn.close()
        return jsonify({'ok': False, 'error': 'Community or base row not found'}), 404

    base_version = base[0]
    base_values = {COLUMN_KEYS[i]: (base[i + 1] or '') for i in range(len(COLUMN_KEYS))}

    # Merge: base values + user edits
    merged = {k: str(row_data.get(k, base_values.get(k, '')) or '') for k in COLUMN_KEYS}

    row_id = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()
    new_version = base_version + 1

    col_names = ', '.join(COLUMN_KEYS)
    placeholders = ', '.join(['?'] * len(COLUMN_KEYS))
    values = [merged[k] for k in COLUMN_KEYS]

    c.execute(f'INSERT INTO underwriting_rows (id, community_id, row_version, created_at, {col_names}) VALUES (?, ?, ?, ?, {placeholders})',
              [row_id, community_id, new_version, now] + values)
    # Update community timestamp
    c.execute('UPDATE underwriting_communities SET updated_at = ? WHERE id = ?', (now, community_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'row_id': row_id, 'row_version': new_version})


@app.route('/api/underwriting/rows/<row_id>', methods=['PATCH'])
@require_auth
@require_role('admin')
def api_uw_update_row(row_id):
    """Update fields on an existing underwriting row."""
    data = request.json or {}
    if not data:
        return jsonify({'ok': False, 'error': 'No fields to update'}), 400

    updates = []
    params = []
    for k in COLUMN_KEYS:
        if k in data:
            updates.append(f'{k} = ?')
            params.append(str(data[k]) if data[k] is not None else '')
    if not updates:
        return jsonify({'ok': False, 'error': 'No valid fields provided'}), 400

    params.append(row_id)
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute(f'UPDATE underwriting_rows SET {", ".join(updates)} WHERE id = ?', params)
    if c.rowcount == 0:
        conn.close()
        return jsonify({'ok': False, 'error': 'Row not found'}), 404
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/underwriting/rows', methods=['GET'])
@require_auth
@require_role('admin')
def api_uw_list_rows():
    """List underwriting rows. ?community_id=&latest=true|false"""
    community_id = request.args.get('community_id', '')
    latest_only = request.args.get('latest', 'true').lower() == 'true'

    conn = _get_db_conn()
    c = conn.cursor()

    col_select = ', '.join(f'r.{k}' for k in COLUMN_KEYS)

    if latest_only:
        # Subquery: max row_version per community
        sql = f'''SELECT r.id, r.community_id, r.row_version, r.created_at,
                         uc.community_key, uc.location_name as uc_name, {col_select}
                  FROM underwriting_rows r
                  JOIN underwriting_communities uc ON uc.id = r.community_id
                  INNER JOIN (
                      SELECT community_id, MAX(row_version) as max_v FROM underwriting_rows GROUP BY community_id
                  ) latest ON r.community_id = latest.community_id AND r.row_version = latest.max_v'''
        if community_id:
            sql += ' WHERE r.community_id = ?'
            c.execute(sql + ' ORDER BY uc.location_name', (community_id,))
        else:
            c.execute(sql + ' ORDER BY uc.location_name')
    else:
        sql = f'''SELECT r.id, r.community_id, r.row_version, r.created_at,
                         uc.community_key, uc.location_name as uc_name, {col_select}
                  FROM underwriting_rows r
                  JOIN underwriting_communities uc ON uc.id = r.community_id'''
        if community_id:
            sql += ' WHERE r.community_id = ?'
            c.execute(sql + ' ORDER BY uc.location_name, r.row_version', (community_id,))
        else:
            c.execute(sql + ' ORDER BY uc.location_name, r.row_version')

    rows_raw = c.fetchall()
    conn.close()

    rows = []
    for r in rows_raw:
        row_dict = {
            'id': r[0], 'community_id': r[1], 'row_version': r[2],
            'created_at': str(r[3] or ''),
            'community_key': r[4], '_community_name': r[5],
        }
        for i, k in enumerate(COLUMN_KEYS):
            row_dict[k] = r[6 + i] or ''
        rows.append(row_dict)

    return jsonify({'ok': True, 'data': rows, 'meta': {'count': len(rows)}})


@app.route('/api/underwriting/export', methods=['GET'])
@require_auth
@require_role('admin')
def api_uw_export():
    """Export underwriting rows as XLSX with exact spreadsheet columns."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, numbers
    from openpyxl.utils import get_column_letter

    mode = request.args.get('mode', 'latest')
    community_id = request.args.get('community_id', '')

    conn = _get_db_conn()
    c = conn.cursor()

    col_select = ', '.join(f'r.{k}' for k in COLUMN_KEYS)

    if mode == 'latest':
        sql = f'''SELECT r.row_version, {col_select}
                  FROM underwriting_rows r
                  JOIN underwriting_communities uc ON uc.id = r.community_id
                  INNER JOIN (
                      SELECT community_id, MAX(row_version) as max_v FROM underwriting_rows GROUP BY community_id
                  ) latest ON r.community_id = latest.community_id AND r.row_version = latest.max_v'''
    else:
        sql = f'''SELECT r.row_version, {col_select}
                  FROM underwriting_rows r
                  JOIN underwriting_communities uc ON uc.id = r.community_id'''

    if community_id:
        sql += ' WHERE r.community_id = ?'
        c.execute(sql + ' ORDER BY uc.location_name, r.row_version', (community_id,))
    else:
        c.execute(sql + ' ORDER BY uc.location_name, r.row_version')

    rows_raw = c.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Underwriting SOV'

    # Header row
    headers = [col['header'] for col in UNDERWRITING_COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    for r in rows_raw:
        row_vals = []
        for i, col in enumerate(UNDERWRITING_COLUMNS):
            val = r[i + 1] or ''  # +1 to skip row_version
            ctype = col['type']
            if val and ctype in ('currency', 'numeric'):
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass
            elif val and ctype == 'integer':
                try:
                    val = int(float(val))
                except (ValueError, TypeError):
                    pass
            elif val and ctype == 'percent':
                try:
                    val = float(val) / 100.0 if float(val) > 1 else float(val)
                except (ValueError, TypeError):
                    pass
            row_vals.append(val)
        ws.append(row_vals)

    # Formatting
    for col_idx, col in enumerate(UNDERWRITING_COLUMNS, 1):
        letter = get_column_letter(col_idx)
        ctype = col['type']
        # Column widths
        ws.column_dimensions[letter].width = max(14, min(len(col['header']) + 4, 30))
        # Number formats for data rows
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if ctype == 'currency':
                cell.number_format = '#,##0.00'
            elif ctype == 'percent':
                cell.number_format = '0.0%'
            elif ctype == 'date':
                cell.number_format = 'YYYY-MM-DD'

    # Freeze header row
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resp = make_response(output.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename=underwriting_sov_{mode}.xlsx'
    return resp


@app.route('/api/underwriting/columns', methods=['GET'])
@require_auth
@require_role('admin')
def api_uw_columns():
    """Return canonical column definitions for frontend grid/form rendering."""
    return jsonify({'ok': True, 'columns': UNDERWRITING_COLUMNS})


# API Routes

@app.route('/')
def index():
    """Serve the main HTML app"""
    return send_from_directory('static', 'index.html')

@app.route('/favicon.ico')
def favicon():
    """Prevent favicon 404 errors"""
    return '', 204

@app.route('/health')
def health():
    """Simple health check for Railway"""
    return jsonify({'status': 'ok'}), 200

@app.route('/api/health')
def api_health():
    """API health check endpoint with DB persistence verification."""
    api_key = os.getenv('ANTHROPIC_API_KEY')
    key_status = 'not set'
    if api_key and api_key != 'your_anthropic_api_key_here':
        key_status = f'configured (ends in ...{api_key[-4:]})'

    # DB type and table counts for persistence verification
    db_type = 'postgres' if _is_postgres() else 'sqlite'
    table_counts = {}
    try:
        conn = _get_db_conn()
        c = conn.cursor()
        for table in ['users', 'crm_leads', 'crm_companies', 'prospecting_runs',
                      'run_prospects', 'discovery_runs', 'government_signals']:
            try:
                c.execute(f'SELECT COUNT(*) FROM {table}')
                table_counts[table] = c.fetchone()[0]
            except Exception:
                table_counts[table] = -1
        conn.close()
    except Exception as e:
        table_counts = {'error': str(e)}

    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'api_key_status': key_status,
        'db_type': db_type,
        'db_table_counts': table_counts,
        'railway': bool(os.getenv('RAILWAY_ENVIRONMENT', '')),
    }), 200

@app.route('/api/search', methods=['POST'])
@require_auth
def api_search():
    """Search for new BTR prospects"""
    try:
        data = request.json
        city = data.get('city', 'Texas')
        limit = data.get('limit', 10)

        print(f"Searching for {limit} prospects in {city}...")

        # Search using Claude API - now returns (prospects, error_msg)
        prospects, error_msg = search_btr_prospects(city, limit)

        if error_msg:
            # On rate limit or API failure, return existing DB results so UI isn't empty
            db_prospects = get_all_prospects_from_db()
            return jsonify({
                'success': False,
                'message': error_msg,
                'prospects': db_prospects,
                'fromCache': True
            }), 200

        if not prospects:
            return jsonify({
                'success': False,
                'message': 'No prospects found. Try a different city or state.',
                'prospects': []
            }), 200

        # Save to database
        saved_count = save_prospects_to_db(prospects)

        # Auto-update master spreadsheet
        total = generate_master_csv()

        return jsonify({
            'success': True,
            'message': f'Found {len(prospects)} prospects, saved {saved_count} new ones. Master spreadsheet updated ({total} total).',
            'prospects': prospects,
            'savedCount': saved_count,
            'totalInSpreadsheet': total
        })

    except Exception as e:
        print(f"API Error: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Search failed: {str(e)}',
            'prospects': []
        }), 500

@app.route('/api/prospects', methods=['GET'])
@require_auth
def api_get_prospects():
    """Get all prospects from database"""
    try:
        prospects = get_all_prospects_from_db()
        return jsonify({
            'success': True,
            'prospects': prospects
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e),
            'prospects': []
        }), 500

@app.route('/api/prospects/<int:prospect_id>', methods=['DELETE'])
@require_auth
def api_delete_prospect(prospect_id):
    """Delete a prospect"""
    try:
        conn = _get_db_conn()
        c = conn.cursor()
        c.execute('DELETE FROM prospects WHERE id = ?', (prospect_id,))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Prospect deleted'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/api/export', methods=['GET'])
@require_auth
def api_export_csv():
    """Download the master CSV spreadsheet with all prospects"""
    try:
        # Always regenerate fresh from DB before download
        count = generate_master_csv()
        if count == 0:
            return jsonify({
                'success': False,
                'message': 'No prospects to export. Run a search first.'
            }), 200

        return send_file(
            MASTER_CSV_PATH,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'btr-prospects-master-{datetime.now().strftime("%Y-%m-%d")}.csv'
        )
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Export failed: {str(e)}'
        }), 500

@app.route('/api/email/generate', methods=['POST'])
@require_auth
@require_any_role('admin', 'producer')
def api_generate_email():
    """Generate personalized email for a prospect using EMAIL GEN SPEC"""
    try:
        data = request.json
        prospect = data.get('prospect', {})

        if not prospect or not prospect.get('company'):
            return jsonify({
                'success': False,
                'message': 'No prospect data provided.'
            }), 400

        # Email options (with defaults)
        email_purpose = data.get('emailPurpose', 'cold_outreach')
        tone = data.get('tone', 'professional_direct')
        offer = data.get('offer', '15_min_call')
        trigger_event = data.get('triggerEvent', '')

        # Extract name parts safely
        full_name = str(prospect.get('executive', '') or '').strip()
        name_parts = full_name.split() if full_name else []
        first_name = name_parts[0] if name_parts else 'there'
        last_name = name_parts[-1] if len(name_parts) > 1 else ''

        # Determine role-based angle from title
        title = str(prospect.get('title', '') or '').lower()
        if any(k in title for k in ['cfo', 'finance', 'capital', 'treasurer']):
            role_angle = "CFO/Finance: cost of risk -> DSCR/refi/exit impact"
        elif any(k in title for k in ['asset', 'portfolio']):
            role_angle = "Asset Management: portfolio consistency + claims outcomes + renewals stability"
        elif any(k in title for k in ['develop', 'construction', 'build', 'project']):
            role_angle = "Developer/Construction: structure early + construction-to-perm + cost control"
        elif any(k in title for k in ['operat', 'property', 'manage']):
            role_angle = "Operations: fewer surprises + smoother renewals + practical risk fixes"
        elif any(k in title for k in ['broker', 'agent']):
            role_angle = "Broker/AM: partnership + differentiated capacity + program fit"
        else:
            role_angle = "Developer/Construction: structure early + construction-to-perm + cost control"

        # Safely handle signals - could be list, JSON string, or plain string
        raw_signals = prospect.get('signals', [])
        if isinstance(raw_signals, str):
            try:
                raw_signals = json.loads(raw_signals)
            except (json.JSONDecodeError, ValueError):
                raw_signals = [raw_signals] if raw_signals else []
        if isinstance(raw_signals, list):
            signals = ', '.join(str(s) for s in raw_signals)
        else:
            signals = str(raw_signals)

        trigger = str(trigger_event or prospect.get('whyNow', '') or '')

        prompt = f"""You are an expert B2B email copywriter for commercial insurance. Follow the EMAIL GEN SPEC exactly.

RULES:
- 90-150 words total (HARD CAP 170)
- 2-4 short paragraphs, 1-2 sentences each
- Grade level: clear, plain English; no jargon unless industry-specific
- Ask exactly ONE question, and it must be the CTA
- No bullet lists
- Do NOT use any of these phrases: "Hope you're doing well", "hope this finds you well", "Just checking in", "Circling back" (unless follow_up), multiple CTAs, overconfident claims
- Include at most ONE specific detail reference about the company/project
- If you cannot verify a detail, omit it — NEVER invent facts
- No generic flattery ("impressive company", "love what you're doing")

Generate a personalized email with these inputs:

RECIPIENT:
- recipient_first_name: {first_name}
- recipient_last_name: {last_name}
- recipient_title: {prospect.get('title') or 'Executive'}
- company_name: {prospect.get('company', '')}
- industry: BTR (Build-to-Rent)
- geography: {prospect.get('city') or 'Texas'}, {prospect.get('state') or 'TX'}

EMAIL CONFIG:
- email_purpose: {email_purpose}
- tone: {tone}
- offer: {offer}
- value_prop: We specialize in protecting BTR portfolios — structuring insurance programs that safeguard NOI, stabilize renewal costs, and protect exit valuations from day one.

SENDER:
- sender_name: Max Hopkins
- sender_title: BTR Insurance Specialist
- sender_company: BTR Insurance
- sender_email: max@btrinsurance.com

CONTEXT:
- Project: {prospect.get('projectName') or 'N/A'}
- Project Status: {prospect.get('projectStatus') or 'N/A'}
- TIV: {prospect.get('tiv') or 'N/A'}
- Units: {prospect.get('units') or 'N/A'}
- Signals: {signals}
- Trigger Event / Why Now: {trigger}

ROLE-BASED ANGLE TO USE: {role_angle}

CTA RULES:
- If offer is 15_min_call: "Open to a quick 15-minute call this week?"
- If offer is share_resource: Offer to send a relevant resource
- If offer is quick_question: Ask one specific question about their situation

OUTPUT FORMAT — Return ONLY valid JSON, nothing else:
{{"subject": "3-6 word subject line", "body": "Hi {first_name},\\n\\n[paragraph 1]\\n\\n[paragraph 2]\\n\\nBest,\\nMax Hopkins\\nBTR Insurance Specialist\\nmax@btrinsurance.com"}}

QUALITY CHECKS before finalizing:
- Word count <= 170
- One question mark total
- One CTA only
- No forbidden phrases
- No unverified claims"""

        print(f"Email gen: generating for {prospect.get('company')} / {full_name}")

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=800,
            messages=[
                {"role": "user", "content": prompt}
            ]
        )

        email_text = message.content[0].text if message.content else ""
        print(f"Email gen raw response: {email_text[:200]}")

        # Try to parse JSON response
        try:
            # Strip any markdown code fences
            cleaned = email_text.strip()
            if cleaned.startswith('```'):
                # Remove opening fence (with optional language tag)
                first_newline = cleaned.find('\n')
                if first_newline != -1:
                    cleaned = cleaned[first_newline + 1:]
                else:
                    cleaned = cleaned[3:]
                # Remove closing fence
                if cleaned.rstrip().endswith('```'):
                    cleaned = cleaned.rstrip()[:-3]
                cleaned = cleaned.strip()

            email_json = json.loads(cleaned)
            subject = email_json.get('subject', '')
            body = email_json.get('body', '')

            print(f"Email gen success: subject='{subject[:50]}', body length={len(body)}")

            return jsonify({
                'success': True,
                'subject': subject,
                'body': body,
                'email': f"Subject: {subject}\n\n{body}" if subject else body
            })
        except (json.JSONDecodeError, ValueError) as parse_err:
            print(f"Email gen JSON parse failed: {parse_err}, returning raw text")
            # Fallback: try to extract subject line from raw text
            subject = ''
            body = email_text
            lines = email_text.strip().split('\n')
            for i, line in enumerate(lines):
                if line.upper().startswith('SUBJECT:'):
                    subject = line.split(':', 1)[1].strip()
                    body = '\n'.join(lines[i+1:]).strip()
                    break

            return jsonify({
                'success': True,
                'subject': subject,
                'body': body,
                'email': email_text
            })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e),
            'email': ''
        }), 500

# --- Async Prospecting Run Endpoints ---

@app.route('/api/prospecting/run', methods=['POST'])
@require_auth
def api_start_prospecting_run():
    """Start an async prospecting run. Returns immediately with a run_id."""
    try:
        data = request.json or {}
        raw_cities = data.get('cities', [])
        max_per_city = min(int(data.get('maxProspectsPerCity', 25)), 50)
        max_total = min(int(data.get('maxTotalProspects', 300)), 300)

        # Normalize cities: accept strings or {city, state} objects
        cities = []
        for c in raw_cities:
            if isinstance(c, str):
                cities.append({'city': c, 'state': ''})
            elif isinstance(c, dict):
                cities.append({'city': c.get('city', ''), 'state': c.get('state', '')})

        if not cities:
            return jsonify({'success': False, 'message': 'No cities provided.'}), 400

        run_id = str(uuid.uuid4())
        search_params = {
            'cities': cities,
            'maxProspectsPerCity': max_per_city,
            'maxTotalProspects': max_total,
        }

        # Create run record
        conn = _get_db_conn()
        c = conn.cursor()
        c.execute('''
            INSERT INTO prospecting_runs (id, status, search_params, created_at, updated_at)
            VALUES (?, 'pending', ?, ?, ?)
        ''', (run_id, json.dumps(search_params),
              datetime.utcnow().isoformat(), datetime.utcnow().isoformat()))
        conn.commit()
        conn.close()

        # Enqueue background job
        from queue_config import enqueue
        from jobs import execute_prospecting_run
        enqueue(execute_prospecting_run, run_id, search_params, job_timeout=600)

        return jsonify({
            'success': True,
            'run_id': run_id,
            'status': 'pending'
        })

    except Exception as e:
        print(f"[Prospecting] Start error: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/prospecting/run/<run_id>/status', methods=['GET'])
@require_auth
def api_prospecting_run_status(run_id):
    """Poll the status of a prospecting run."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT status, total_prospects, error, created_at, completed_at FROM prospecting_runs WHERE id = ?',
              (run_id,))
    row = c.fetchone()
    conn.close()

    if not row:
        return jsonify({'success': False, 'message': 'Run not found'}), 404

    return jsonify({
        'success': True,
        'run_id': run_id,
        'status': row[0],
        'total_prospects': row[1] or 0,
        'error': row[2],
        'created_at': row[3],
        'completed_at': row[4],
    })


@app.route('/api/prospecting/run/<run_id>/results', methods=['GET'])
@require_auth
def api_prospecting_run_results(run_id):
    """Get results for a prospecting run, sorted by score DESC, with pagination."""
    limit = min(int(request.args.get('limit', 50)), 200)
    offset = int(request.args.get('offset', 0))

    conn = _get_db_conn()
    c = conn.cursor()

    # Total count
    c.execute('SELECT COUNT(*) FROM run_prospects WHERE run_id = ?', (run_id,))
    total = c.fetchone()[0]

    # Fetch page
    c.execute('''
        SELECT id, company_name, city, state, score, tiv_estimate, deal_status,
               signals, why_call_now, executive, title, linkedin, units, project_name, created_at, score_meta
        FROM run_prospects
        WHERE run_id = ?
        ORDER BY score DESC
        LIMIT ? OFFSET ?
    ''', (run_id, limit, offset))

    prospects = []
    for row in c.fetchall():
        signals_raw = row[7]
        try:
            signals = json.loads(signals_raw) if signals_raw else []
        except (json.JSONDecodeError, TypeError):
            signals = [signals_raw] if signals_raw else []

        # Unpack score_meta if present
        score_meta = {}
        try:
            score_meta = json.loads(row[15]) if len(row) > 15 and row[15] else {}
        except (json.JSONDecodeError, TypeError, IndexError):
            pass

        prospect = {
            'id': row[0],
            'company': row[1],
            'city': row[2],
            'state': row[3],
            'score': row[4],
            'tiv': row[5],
            'projectStatus': row[6],
            'signals': signals,
            'whyNow': row[8],
            'executive': row[9],
            'title': row[10],
            'linkedin': row[11],
            'units': row[12],
            'projectName': row[13],
            'createdAt': row[14],
        }

        # Merge score_meta fields into prospect
        if score_meta:
            prospect['score_breakdown'] = score_meta.get('score_breakdown', {})
            prospect['score_explanation'] = score_meta.get('score_explanation', [])
            prospect['insurance_triggers'] = score_meta.get('insurance_triggers', [])
            prospect['unit_band'] = score_meta.get('unit_band', '')
            prospect['active_project_count_estimate'] = score_meta.get('active_project_count_estimate', 0)
            prospect['markets_active_estimate'] = score_meta.get('markets_active_estimate', 0)
            prospect['swim_lane_fit_score'] = score_meta.get('swim_lane_fit_score', 0)
            prospect['competitive_difficulty'] = score_meta.get('competitive_difficulty', '')

        prospects.append(prospect)

    conn.close()

    # Enrich with government signals (no scoring changes, no latency penalty on failure)
    prospects = enrich_prospects_with_gov_signals(prospects)

    return jsonify({
        'success': True,
        'total': total,
        'prospects': prospects,
    })


# --- Daily Discovery API Routes ---

@app.route('/api/discovery/config', methods=['GET'])
@require_auth
@require_any_role('admin', 'producer')
def api_discovery_config():
    """Get current discovery configuration"""
    return jsonify({'success': True, 'config': DISCOVERY_CONFIG})


@app.route('/api/discovery/config', methods=['PUT'])
@require_auth
@require_role('admin')
def api_update_discovery_config():
    """Update discovery configuration (runtime only)"""
    data = request.json
    if 'max_signals_per_day' in data:
        DISCOVERY_CONFIG['max_signals_per_day'] = int(data['max_signals_per_day'])
    if 'icp_keywords' in data:
        DISCOVERY_CONFIG['icp_keywords'] = data['icp_keywords']
    if 'delivery_method' in data:
        DISCOVERY_CONFIG['delivery_method'] = data['delivery_method']
    if 'webhook_url' in data:
        DISCOVERY_CONFIG['webhook_url'] = data['webhook_url']
    return jsonify({'success': True, 'config': DISCOVERY_CONFIG})


@app.route('/api/discovery/latest', methods=['GET'])
@require_auth
@require_any_role('admin', 'producer')
def api_discovery_latest():
    """Get the most recent discovery run"""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT * FROM discovery_runs ORDER BY run_at DESC LIMIT 1')
    row = c.fetchone()
    conn.close()
    if not row:
        return jsonify({'success': True, 'run': None, 'message': 'No discovery runs yet.'})
    adapter_stats = None
    try:
        adapter_stats = json.loads(row[7]) if len(row) > 7 and row[7] else None
    except (json.JSONDecodeError, TypeError, IndexError):
        pass
    return jsonify({
        'success': True,
        'run': {
            'id': row[0], 'run_at': row[1],
            'results': json.loads(row[2]) if row[2] else {},
            'digest': row[3], 'city_count': row[4],
            'total_new': row[5], 'status': row[6],
            'adapter_stats': adapter_stats
        }
    })


@app.route('/api/discovery/run', methods=['POST'])
@require_auth
@require_any_role('admin', 'producer')
def api_discovery_run():
    """Manually trigger a discovery run (background thread)"""
    global _discovery_running
    if _discovery_running:
        return jsonify({'success': False, 'message': 'A discovery run is already in progress.'})

    _discovery_running = True

    def run_bg():
        global _discovery_running
        try:
            run_daily_discovery(is_scheduled=False)
        except Exception as e:
            print(f"[Discovery] Background run failed: {e}")
            traceback.print_exc()
        finally:
            _discovery_running = False

    thread = threading.Thread(target=run_bg, daemon=True)
    thread.start()
    return jsonify({'success': True, 'message': 'Discovery run started. Results will appear shortly.'})


@app.route('/api/discovery/status', methods=['GET'])
@require_auth
@require_any_role('admin', 'producer')
def api_discovery_status():
    """Check if a discovery run is currently in progress"""
    return jsonify({'success': True, 'running': _discovery_running})


@app.route('/api/discovery/history', methods=['GET'])
@require_auth
@require_any_role('admin', 'producer')
def api_discovery_history():
    """Get past discovery run summaries"""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT id, run_at, city_count, total_new, status FROM discovery_runs ORDER BY run_at DESC LIMIT 20')
    runs = []
    for row in c.fetchall():
        runs.append({
            'id': row[0], 'run_at': row[1],
            'city_count': row[2], 'total_new': row[3], 'status': row[4]
        })
    conn.close()
    return jsonify({'success': True, 'runs': runs})


@app.route('/api/discovery/run/<int:run_id>', methods=['GET'])
@require_auth
@require_any_role('admin', 'producer')
def api_discovery_run_detail(run_id):
    """Get full results for a specific discovery run"""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT * FROM discovery_runs WHERE id = ?', (run_id,))
    row = c.fetchone()
    conn.close()
    if not row:
        return jsonify({'success': False, 'message': 'Run not found'}), 404
    adapter_stats = None
    try:
        adapter_stats = json.loads(row[7]) if len(row) > 7 and row[7] else None
    except (json.JSONDecodeError, TypeError, IndexError):
        pass
    return jsonify({
        'success': True,
        'run': {
            'id': row[0], 'run_at': row[1],
            'results': json.loads(row[2]) if row[2] else {},
            'digest': row[3], 'city_count': row[4],
            'total_new': row[5], 'status': row[6],
            'adapter_stats': adapter_stats
        }
    })


@app.route('/api/discovery/source-refresh', methods=['GET'])
@require_auth
@require_any_role('admin', 'producer')
def api_discovery_source_refresh():
    """Get last-refreshed timestamps for each source type"""
    from discovery_engine import get_source_refresh_times
    return jsonify({'success': True, 'sources': get_source_refresh_times()})


# ===================================================================
# STATEWIDE SEARCH + RANKINGS
# ===================================================================

STATEWIDE_STATES = {
    'TX': 'Texas',
    'AZ': 'Arizona',
    'GA': 'Georgia',
    'NC': 'North Carolina',
    'FL': 'Florida',
}

def _run_statewide_search(state_abbr):
    """
    Run statewide BTR search for a single state.
    Uses max 2 SerpAPI queries, 24h caching, Claude extraction (no web search tool).
    Returns (items_list, error_message) tuple.
    Each item: {event_type, city, state, company, units, summary, confidence, title, url, snippet, date}
    """
    from serpapi_client import cached_serpapi_search, get_cached, set_cached, SerpAPIError

    state_name = STATEWIDE_STATES.get(state_abbr, state_abbr)
    today = datetime.utcnow().strftime('%Y-%m-%d')

    # Check full-result cache first (Claude-processed output)
    result_cache_key = f"statewide_result:{state_abbr.lower()}:{today}"
    cached_result = get_cached(result_cache_key)
    if cached_result is not None:
        print(f"[Statewide] Cache hit for {state_abbr} processed results")
        return cached_result, None

    serp_key = os.getenv('SERPAPI_API_KEY', '')
    if not serp_key:
        return [], "SerpAPI key not configured. Statewide search requires SerpAPI."

    # Stage A: 2 SerpAPI queries (cached per-query for 24h via cached_serpapi_search)
    q1 = f'("build to rent" OR BTR OR "single-family rental") "{state_name}" (acquires OR acquisition OR sale OR sells OR JV OR recap OR refinancing OR "credit facility")'
    q2 = f'("build to rent" OR BTR OR "horizontal multifamily") "{state_name}" (groundbreaking OR "under construction" OR permit OR rezoning OR entitlement OR "planning commission")'

    all_candidates = []
    seen_links = set()

    for query in [q1, q2]:
        try:
            results = cached_serpapi_search(query, num=15, feature='statewide', city=state_abbr, state=state_abbr)
            for r in results:
                link = r.get('link', '')
                if link and link not in seen_links:
                    seen_links.add(link)
                    all_candidates.append(r)
        except SerpAPIError as e:
            print(f"[Statewide] SerpAPI error for {state_abbr}: {e}")
            # Continue with what we have

    if not all_candidates:
        return [], None

    # Stage B: Claude extraction (no web search tool)
    candidates_json = json.dumps([{
        'title': c.get('title', ''),
        'url': c.get('link', ''),
        'snippet': c.get('snippet', ''),
        'source': c.get('source', ''),
        'date': c.get('date', ''),
    } for c in all_candidates], indent=2)

    today_str = datetime.utcnow().strftime('%B %d, %Y')
    ninety_days_ago = (datetime.utcnow() - timedelta(days=90)).strftime('%B %Y')

    extraction_prompt = f"""You are an analyst for a Build-to-Rent (BTR) and Single-Family Rental (SFR) insurance brokerage.

Analyze these search results for BTR/SFR activity in {state_name} ({state_abbr}).

CANDIDATE RESULTS:
{candidates_json}

Today is {today_str}. Only include events from the last 90 days (since {ninety_days_ago}).

For EACH relevant result, extract:
- event_type: one of "acquisition", "sale", "groundbreaking", "permit", "rezoning", "financing", "JV", "construction", "other"
- city: infer from title/snippet. If unknown, set "Unknown"
- state: "{state_abbr}"
- company: the operator/developer/buyer involved (if identifiable)
- units: numeric estimate if mentioned, else null
- summary: one-sentence description of the event
- confidence: "high" (explicit BTR/SFR mention), "medium" (likely BTR context), "low" (possible but uncertain)
- title: the original title
- url: the original url
- date: the date from the result if available

IMPORTANT:
- Skip results that are clearly NOT about BTR/SFR real estate (e.g. other industries)
- One result may yield one item
- Do NOT fabricate data; only extract what is in the snippets/titles

Return ONLY valid JSON:
{{"items": [
  {{"event_type": "...", "city": "...", "state": "{state_abbr}", "company": "...", "units": null, "summary": "...", "confidence": "...", "title": "...", "url": "...", "date": "..."}}
]}}"""

    try:
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            messages=[{"role": "user", "content": extraction_prompt}]
        )
        response_text = message.content[0].text if message.content else ''

        # Parse JSON from response
        items = []
        try:
            # Try direct parse
            parsed = json.loads(response_text)
            items = parsed.get('items', [])
        except json.JSONDecodeError:
            # Find JSON block in response
            start = response_text.find('{"items"')
            if start == -1:
                start = response_text.find('```json')
                if start != -1:
                    start = response_text.find('{', start)
            if start != -1:
                depth = 0
                end = start
                for i in range(start, len(response_text)):
                    if response_text[i] == '{':
                        depth += 1
                    elif response_text[i] == '}':
                        depth -= 1
                        if depth == 0:
                            end = i + 1
                            break
                try:
                    parsed = json.loads(response_text[start:end])
                    items = parsed.get('items', [])
                except json.JSONDecodeError:
                    print(f"[Statewide] Failed to parse Claude response for {state_abbr}")

        # Cache processed results for 24h
        set_cached(result_cache_key, items)
        return items, None

    except Exception as e:
        print(f"[Statewide] Claude extraction error for {state_abbr}: {e}")
        traceback.print_exc()
        return [], str(e)


def _compute_activity_scores(items, days=7):
    """
    Compute activity scores grouped by city from statewide items.
    Returns dict: {city: {activity_score, signals_count, dominant_event_types}}
    """
    from collections import Counter
    cutoff = (datetime.utcnow() - timedelta(days=days)).isoformat()

    city_items = {}
    for item in items:
        city = item.get('city', 'Unknown')
        if city == 'Unknown':
            continue
        if city not in city_items:
            city_items[city] = []
        city_items[city].append(item)

    city_scores = {}
    for city, city_item_list in city_items.items():
        score = 0
        event_counts = Counter()

        for item in city_item_list:
            event_type = item.get('event_type', 'other')
            confidence = item.get('confidence', 'low')
            event_counts[event_type] += 1

            # Base point per signal
            base = 1
            # Confidence weight
            if confidence == 'high':
                base *= 2.0
            elif confidence == 'medium':
                base *= 1.5

            # Event type weight
            if event_type in ('acquisition', 'sale', 'financing', 'JV'):
                base *= 3.0  # Capital events weighted higher
            elif event_type in ('groundbreaking', 'construction', 'permit', 'rezoning'):
                base *= 2.0  # Construction activity
            else:
                base *= 1.0

            score += base

        # Dominant event types (top 3)
        dominant = [et for et, _ in event_counts.most_common(3)]

        city_scores[city] = {
            'city': city,
            'activity_score': round(score, 1),
            'signals_count': len(city_item_list),
            'dominant_event_types': dominant,
        }

    return city_scores


@app.route('/api/discovery/statewide/run', methods=['POST'])
@require_auth
@require_any_role('admin', 'producer')
def api_statewide_run():
    """Run statewide search for a single state. Returns items immediately (cached)."""
    data = request.json or {}
    state = data.get('state', '').upper()
    if state not in STATEWIDE_STATES:
        return jsonify({'success': False, 'message': f'State must be one of: {", ".join(STATEWIDE_STATES.keys())}'}), 400

    try:
        items, error = _run_statewide_search(state)
        if error:
            return jsonify({'success': False, 'message': error}), 500
        return jsonify({'success': True, 'state': state, 'items': items, 'count': len(items)})
    except Exception as e:
        print(f"[Statewide] Run error: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/discovery/state-summary', methods=['GET'])
@require_auth
@require_any_role('admin', 'producer')
def api_state_summary():
    """Get activity summary for a state: top cities (7d + 30d) + items."""
    from serpapi_client import get_cached
    state = request.args.get('state', '').upper()
    if state not in STATEWIDE_STATES:
        return jsonify({'success': False, 'message': 'Invalid state'}), 400

    today = datetime.utcnow().strftime('%Y-%m-%d')
    result_cache_key = f"statewide_result:{state.lower()}:{today}"
    items = get_cached(result_cache_key) or []

    # Compute top cities
    scores_7d = _compute_activity_scores(items, days=7)
    scores_30d = _compute_activity_scores(items, days=30)

    top_7d = sorted(scores_7d.values(), key=lambda x: x['activity_score'], reverse=True)[:3]
    top_30d = sorted(scores_30d.values(), key=lambda x: x['activity_score'], reverse=True)[:3]

    return jsonify({
        'success': True,
        'state': state,
        'top_cities_7d': top_7d,
        'top_cities_30d': top_30d,
        'items': items,
        'total_signals': len(items),
    })


@app.route('/api/discovery/state-rankings', methods=['GET'])
@require_auth
@require_any_role('admin', 'producer')
def api_state_rankings():
    """Get rankings across all 5 states for the last 7 days."""
    from serpapi_client import get_cached
    today = datetime.utcnow().strftime('%Y-%m-%d')

    rankings = []
    for state_abbr in STATEWIDE_STATES:
        result_cache_key = f"statewide_result:{state_abbr.lower()}:{today}"
        items = get_cached(result_cache_key) or []

        capital_count = 0
        construction_count = 0
        for item in items:
            et = item.get('event_type', '')
            if et in ('acquisition', 'sale', 'financing', 'JV'):
                capital_count += 1
            elif et in ('groundbreaking', 'construction', 'permit', 'rezoning'):
                construction_count += 1

        scores = _compute_activity_scores(items, days=7)
        total_score = sum(s['activity_score'] for s in scores.values())
        top_cities = sorted(scores.values(), key=lambda x: x['activity_score'], reverse=True)[:3]

        rankings.append({
            'state': state_abbr,
            'state_name': STATEWIDE_STATES[state_abbr],
            'state_activity_score': round(total_score, 1),
            'total_signals': len(items),
            'capital_events_count': capital_count,
            'construction_signals_count': construction_count,
            'top_cities': top_cities,
            'last_updated': today if items else None,
        })

    rankings.sort(key=lambda x: x['state_activity_score'], reverse=True)
    return jsonify({'success': True, 'rankings': rankings})


# ===================================================================
# SIGNAL WEIGHTING + MOMENTUM + CALL TIMING ENGINE
# ===================================================================

def get_signal_weight(topic_or_text):
    """
    Return signal weight 1-5 based on event/signal type.
    Financing/capital events: 5, Acquisition/disposition: 4,
    Construction: 3, Permits/rezoning: 2, Generic: 1.
    """
    t = (topic_or_text or '').lower()
    # Weight 4: Refinance / debt renewal (check BEFORE financing to avoid false match)
    if any(k in t for k in ('refinanc', 'debt facility', 'renewal', 'loan')):
        return 4
    # Weight 5: Financing / capital events
    if any(k in t for k in ('financing', 'credit facility', 'recap', 'recapitalization',
                             'jv', 'joint venture', 'preferred equity', 'capital',
                             'institutional', 'fund', 'credit')):
        return 5
    # Weight 4: Acquisitions / dispositions
    if any(k in t for k in ('acquisition', 'acquires', 'disposition', 'portfolio sale',
                             'sale', 'sells', 'purchase', 'bought')):
        return 4
    # Weight 3: Construction activity
    if any(k in t for k in ('groundbreaking', 'under construction', 'construction',
                             'starts', 'new_build', 'new build', 'breaking ground',
                             'delivered', 'completion')):
        return 3
    # Weight 2: Permits / entitlements
    if any(k in t for k in ('permit', 'rezoning', 'entitlement', 'planning',
                             'approval', 'zoning', 'permit_rezoning')):
        return 2
    # Weight 1: Generic BTR mention
    return 1


def materialize_weighted_signals(days=90):
    """
    Materialize weighted_signals table from discovery_runs + search_cache.
    Idempotent: clears and rebuilds for the window. No SerpAPI calls.
    """
    print("[WeightedSignals] Materializing weighted signals...")
    all_items = _gather_all_signals(days=days)
    if not all_items:
        print("[WeightedSignals] No signals to materialize.")
        return 0

    conn = _get_db_conn()
    c = conn.cursor()
    cutoff = (datetime.utcnow() - timedelta(days=days)).isoformat()
    c.execute('DELETE FROM weighted_signals WHERE published_at < ? OR published_at IS NULL', (cutoff,))

    inserted = 0
    for item in all_items:
        # Compute weight from both topic and textual content
        topic_weight = get_signal_weight(item['topic'])
        text_weight = get_signal_weight(item.get('summary', '') + ' ' + item.get('title', ''))
        weight = max(topic_weight, text_weight)

        sig_id = hashlib.md5(
            f"{item['state']}:{item['city']}:{item['topic']}:{item.get('entity_name','')}:{item['date_str']}"
            .encode()
        ).hexdigest()

        c.execute('''
            INSERT INTO weighted_signals
            (id, state, city, topic, signal_weight, entity_name, title, summary, source, source_type, confidence, published_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT (id) DO UPDATE SET
                signal_weight = EXCLUDED.signal_weight, entity_name = EXCLUDED.entity_name,
                title = EXCLUDED.title, summary = EXCLUDED.summary, source = EXCLUDED.source,
                source_type = EXCLUDED.source_type, confidence = EXCLUDED.confidence, published_at = EXCLUDED.published_at
        ''', (sig_id, item['state'], item['city'], item['topic'], weight,
              item.get('entity_name', ''), item.get('title', ''), item.get('summary', ''),
              item.get('source', ''), item.get('source_type', ''), item.get('confidence', 'medium'),
              item['date_str']))
        inserted += 1

    conn.commit()
    conn.close()
    print(f"[WeightedSignals] Materialized {inserted} weighted signals from {len(all_items)} raw items.")
    return inserted


def compute_market_momentum():
    """
    Daily job: compute market_momentum for each state+city in the 5 Sunbelt states.
    Reads from weighted_signals table. No SerpAPI calls.
    """
    print("[Momentum] Computing market momentum...")
    now = datetime.utcnow()
    window_end = now.strftime('%Y-%m-%d')
    cutoff_7d = (now - timedelta(days=7)).isoformat()
    cutoff_14d = (now - timedelta(days=14)).isoformat()
    cutoff_30d = (now - timedelta(days=30)).isoformat()

    conn = _get_db_conn()
    c = conn.cursor()

    # Get all unique (state, city) pairs from weighted_signals in the 5 states
    # Filter out NULL/empty/Unknown cities to keep rankings clean
    c.execute('''
        SELECT DISTINCT state, city FROM weighted_signals
        WHERE state IN ('TX','AZ','GA','NC','FL') AND published_at >= ?
        AND city IS NOT NULL AND city != '' AND city != 'Unknown'
    ''', (cutoff_30d,))
    markets = c.fetchall()

    inserted = 0
    for state, city in markets:
        # Unweighted counts
        c.execute('SELECT COUNT(*) FROM weighted_signals WHERE state=? AND city=? AND published_at>=?', (state, city, cutoff_7d))
        sig_7d = c.fetchone()[0]
        c.execute('SELECT COUNT(*) FROM weighted_signals WHERE state=? AND city=? AND published_at>=?', (state, city, cutoff_14d))
        sig_14d = c.fetchone()[0]
        c.execute('SELECT COUNT(*) FROM weighted_signals WHERE state=? AND city=? AND published_at>=?', (state, city, cutoff_30d))
        sig_30d = c.fetchone()[0]

        # Weighted counts
        c.execute('SELECT COALESCE(SUM(signal_weight),0) FROM weighted_signals WHERE state=? AND city=? AND published_at>=?', (state, city, cutoff_7d))
        wsig_7d = c.fetchone()[0]
        c.execute('SELECT COALESCE(SUM(signal_weight),0) FROM weighted_signals WHERE state=? AND city=? AND published_at>=?', (state, city, cutoff_14d))
        wsig_14d = c.fetchone()[0]
        c.execute('SELECT COALESCE(SUM(signal_weight),0) FROM weighted_signals WHERE state=? AND city=? AND published_at>=?', (state, city, cutoff_30d))
        wsig_30d = c.fetchone()[0]

        # Momentum score
        baseline = wsig_30d / 4.0
        ratio_7d = wsig_7d / max(1, baseline)
        ratio_14d = wsig_14d / max(1, baseline * 2)
        momentum_score = max(0, min(100, ratio_7d * 60 + ratio_14d * 40))

        if momentum_score >= 65:
            momentum_label = 'Accelerating'
        elif momentum_score >= 40:
            momentum_label = 'Stable'
        else:
            momentum_label = 'Cooling'

        mid = str(uuid.uuid4())
        c.execute('''
            INSERT INTO market_momentum
            (id, state, city, window_end_date, signals_7d, signals_14d, signals_30d,
             weighted_signals_7d, weighted_signals_14d, weighted_signals_30d,
             momentum_score, momentum_label, computed_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT (state, city, window_end_date) DO UPDATE SET
                id = EXCLUDED.id, signals_7d = EXCLUDED.signals_7d, signals_14d = EXCLUDED.signals_14d,
                signals_30d = EXCLUDED.signals_30d, weighted_signals_7d = EXCLUDED.weighted_signals_7d,
                weighted_signals_14d = EXCLUDED.weighted_signals_14d, weighted_signals_30d = EXCLUDED.weighted_signals_30d,
                momentum_score = EXCLUDED.momentum_score, momentum_label = EXCLUDED.momentum_label
        ''', (mid, state, city, window_end, sig_7d, sig_14d, sig_30d,
              wsig_7d, wsig_14d, wsig_30d,
              round(momentum_score, 1), momentum_label))
        inserted += 1

    conn.commit()
    conn.close()
    print(f"[Momentum] Computed momentum for {inserted} markets.")
    return inserted


def _get_momentum_for_city(state, city):
    """Lookup latest momentum_score for a city. Returns (score, label) or (50, 'Stable')."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('''
        SELECT momentum_score, momentum_label FROM market_momentum
        WHERE state=? AND city=? ORDER BY window_end_date DESC LIMIT 1
    ''', (state, city))
    row = c.fetchone()
    conn.close()
    if row:
        return row[0], row[1]
    return 50.0, 'Stable'


def _compute_freshness_score(state, city, company_name):
    """Compute freshness (0-100) based on most recent signal for this entity/market."""
    conn = _get_db_conn()
    c = conn.cursor()
    # Look for signals matching entity or city
    c.execute('''
        SELECT MAX(published_at) FROM weighted_signals
        WHERE state=? AND city=? AND (entity_name LIKE ? OR entity_name = '')
    ''', (state, city, f'%{company_name[:20]}%' if company_name else '%'))
    row = c.fetchone()
    conn.close()

    if not row or not row[0]:
        return 30
    try:
        pub_date = datetime.fromisoformat(row[0].replace('Z', '+00:00')) if 'T' in row[0] else datetime.strptime(row[0][:10], '%Y-%m-%d')
        days_ago = (datetime.utcnow() - pub_date.replace(tzinfo=None)).days
    except (ValueError, TypeError):
        return 30

    if days_ago <= 14:
        return 90
    elif days_ago <= 30:
        return 70
    elif days_ago <= 90:
        return 50
    return 30


def compute_lead_timing_scores(workspace_id=None):
    """
    Daily job: compute call timing scores for all prospects in run_prospects + CRM.
    Reads precomputed market_momentum. No SerpAPI calls.
    """
    print("[CallTiming] Computing lead timing scores...")
    conn = _get_db_conn()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Gather all prospects from recent prospecting runs (last 90 days)
    cutoff = (datetime.utcnow() - timedelta(days=90)).isoformat()
    c.execute('''
        SELECT rp.company_name, rp.city, rp.state, rp.score, rp.score_meta,
               rp.run_id
        FROM run_prospects rp
        JOIN prospecting_runs pr ON rp.run_id = pr.id
        WHERE pr.created_at >= ? AND rp.score > 0
    ''', (cutoff,))
    prospects = c.fetchall()

    # Also gather from the main prospects table
    c.execute('SELECT company, city, state, score FROM prospects WHERE score > 0')
    main_prospects = c.fetchall()

    conn.close()

    # Build deduped prospect list
    seen_keys = set()
    prospect_list = []

    for p in prospects:
        company = p['company_name']
        city = p['city'] or ''
        state = p['state'] or ''
        key = f"{company.lower().strip()}|{city.lower()}|{state.lower()}"
        if key in seen_keys:
            continue
        seen_keys.add(key)

        score_meta = {}
        if p['score_meta']:
            try:
                score_meta = json.loads(p['score_meta'])
            except (json.JSONDecodeError, TypeError):
                pass

        prospect_list.append({
            'company_name': company,
            'city': city,
            'state': state,
            'score': p['score'],
            'swim_lane_fit_score': score_meta.get('swim_lane_fit_score'),
            'competitive_difficulty': score_meta.get('competitive_difficulty', 'Medium'),
            'unit_band': score_meta.get('unit_band', ''),
            'prospect_key': key,
        })

    for p in main_prospects:
        company = p['company']
        city = p['city'] or ''
        state = p['state'] or ''
        key = f"{company.lower().strip()}|{city.lower()}|{state.lower()}"
        if key in seen_keys:
            continue
        seen_keys.add(key)
        prospect_list.append({
            'company_name': company,
            'city': city,
            'state': state,
            'score': p['score'],
            'swim_lane_fit_score': None,
            'competitive_difficulty': 'Medium',
            'unit_band': '',
            'prospect_key': key,
        })

    if not prospect_list:
        print("[CallTiming] No prospects to score.")
        return 0

    conn = _get_db_conn()
    c = conn.cursor()
    scored = 0

    for p in prospect_list:
        trigger_severity = p['score'] or 0

        # Swim lane fit
        swim_lane_fit = p.get('swim_lane_fit_score')
        if swim_lane_fit is None:
            # Estimate from unit_band
            ub = (p.get('unit_band') or '').lower()
            swim_lane_fit = 50
            if '40-150' in ub:
                swim_lane_fit = 75
            elif '150-400' in ub:
                swim_lane_fit = 70
            elif '<40' in ub:
                swim_lane_fit = 35
            elif '400-1000' in ub:
                swim_lane_fit = 40
            elif '1000' in ub:
                swim_lane_fit = 25

        # Engagement score proxy
        engagement = 60
        cd = (p.get('competitive_difficulty') or 'Medium')
        if cd == 'High':
            engagement -= 15
        elif cd == 'Low':
            engagement += 10
        ub = (p.get('unit_band') or '').lower()
        if '40-150' in ub or '150-400' in ub:
            engagement += 10
        engagement = max(0, min(100, engagement))

        # Market momentum
        momentum, momentum_label = _get_momentum_for_city(p['state'], p['city'])

        # Freshness
        freshness = _compute_freshness_score(p['state'], p['city'], p['company_name'])

        # Call timing formula
        call_timing = (
            0.35 * trigger_severity +
            0.20 * swim_lane_fit +
            0.20 * engagement +
            0.15 * momentum +
            0.10 * freshness
        )
        call_timing = max(0, min(100, round(call_timing, 1)))

        if call_timing >= 75:
            timing_label = 'Call Now'
        elif call_timing >= 55:
            timing_label = 'Work'
        else:
            timing_label = 'Watch'

        # Generate reasons
        reasons = []
        if trigger_severity >= 80:
            reasons.append(f"High trigger severity ({trigger_severity}/100)")
        elif trigger_severity >= 60:
            reasons.append(f"Moderate trigger severity ({trigger_severity}/100)")
        if swim_lane_fit >= 70:
            reasons.append(f"Strong swim-lane fit ({swim_lane_fit}/100)")
        if momentum >= 65:
            reasons.append(f"Accelerating market ({p['city']}, {p['state']})")
        elif momentum < 40:
            reasons.append(f"Cooling market conditions in {p['city']}")
        if freshness >= 70:
            reasons.append("Recent activity signals detected")
        if cd == 'Low':
            reasons.append("Low competitive difficulty - easier engagement")
        elif cd == 'High':
            reasons.append("High competitive difficulty - needs differentiated approach")
        if not reasons:
            reasons.append(f"Composite score: {call_timing}")

        ws_id = workspace_id or 'default'
        lid = str(uuid.uuid4())
        c.execute('''
            INSERT INTO lead_timing_scores
            (id, workspace_id, prospect_key, company_name, state, city,
             trigger_severity, swim_lane_fit, engagement_score,
             market_momentum_score, freshness_score, call_timing_score,
             timing_label, reasons, computed_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT (workspace_id, prospect_key) DO UPDATE SET
                id = EXCLUDED.id, company_name = EXCLUDED.company_name, state = EXCLUDED.state, city = EXCLUDED.city,
                trigger_severity = EXCLUDED.trigger_severity, swim_lane_fit = EXCLUDED.swim_lane_fit,
                engagement_score = EXCLUDED.engagement_score, market_momentum_score = EXCLUDED.market_momentum_score,
                freshness_score = EXCLUDED.freshness_score, call_timing_score = EXCLUDED.call_timing_score,
                timing_label = EXCLUDED.timing_label, reasons = EXCLUDED.reasons
        ''', (lid, ws_id, p['prospect_key'], p['company_name'], p['state'], p['city'],
              trigger_severity, swim_lane_fit, engagement, momentum, freshness,
              call_timing, timing_label, json.dumps(reasons)))
        scored += 1

    conn.commit()
    conn.close()
    print(f"[CallTiming] Scored {scored} prospects.")
    return scored


def run_daily_optimization():
    """
    Master daily job: materialize signals → compute momentum → compute timing.
    Runs after trend detection (8:00am PT). No SerpAPI calls.
    """
    print("[Optimization] Starting daily optimization pipeline...")
    try:
        materialize_weighted_signals(days=90)
        compute_market_momentum()
        compute_lead_timing_scores()
        print("[Optimization] Daily optimization complete.")
    except Exception as e:
        print(f"[Optimization] Error: {e}")
        traceback.print_exc()


# ===================================================================
# TREND DETECTION ENGINE + WEEKLY BRIEF
# ===================================================================

def _gather_all_signals(days=30):
    """
    Gather all signals from discovery_runs and statewide search_cache.
    Returns list of normalized items: {state, city, topic, date_str, source}
    No SerpAPI calls — reads only existing cached/stored data.
    """
    from serpapi_client import get_cached
    cutoff = (datetime.utcnow() - timedelta(days=days)).isoformat()
    all_items = []

    # Source 1: discovery_runs.results_json (daily discovery signals)
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT results_json, run_at FROM discovery_runs WHERE run_at >= ? AND status = ?', (cutoff, 'completed'))
    for row in c.fetchall():
        try:
            results = json.loads(row[0]) if row[0] else {}
            run_date = row[1]
            for location, data in results.items():
                for sig in (data.get('signals') or []):
                    topic = sig.get('signal_type', 'other')
                    if topic == 'not_relevant':
                        continue
                    all_items.append({
                        'state': sig.get('state', ''),
                        'city': sig.get('city', 'Unknown'),
                        'topic': topic,
                        'date_str': run_date,
                        'source': 'discovery',
                        'title': sig.get('title', ''),
                        'url': sig.get('url', ''),
                        'summary': sig.get('summary', ''),
                        'entity_name': sig.get('entity_name', ''),
                        'confidence': sig.get('confidence', 'medium'),
                        'source_type': sig.get('source_type', 'news'),
                    })
        except (json.JSONDecodeError, TypeError):
            continue
    conn.close()

    # Source 2: statewide search_cache (statewide_result:* keys)
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute("SELECT cache_key, payload_json, created_at FROM search_cache WHERE cache_key LIKE 'statewide_result:%'")
    for row in c.fetchall():
        try:
            items = json.loads(row[1]) if row[1] else []
            cache_date = row[2]
            for item in items:
                topic = item.get('event_type', 'other')
                all_items.append({
                    'state': item.get('state', ''),
                    'city': item.get('city', 'Unknown'),
                    'topic': topic,
                    'date_str': cache_date,
                    'source': 'statewide',
                    'title': item.get('title', ''),
                    'url': item.get('url', ''),
                    'summary': item.get('summary', ''),
                    'entity_name': item.get('company', ''),
                    'confidence': item.get('confidence', 'medium'),
                    'source_type': 'news',
                })
        except (json.JSONDecodeError, TypeError):
            continue
    conn.close()

    return all_items


def run_trend_detection():
    """
    Daily trend detection job. Aggregates signals, computes trends, stores results.
    No SerpAPI calls. Operates only on existing data.
    """
    print("[TrendDetection] Starting trend detection...")
    all_items = _gather_all_signals(days=30)
    if not all_items:
        print("[TrendDetection] No signals to analyze.")
        return

    now = datetime.utcnow()
    cutoff_7d = (now - timedelta(days=7)).isoformat()

    # Group by (state, city, topic)
    from collections import defaultdict
    groups = defaultdict(lambda: {'items_7d': 0, 'items_30d': 0})
    for item in all_items:
        key = (item['state'], item['city'], item['topic'])
        groups[key]['items_30d'] += 1
        if item['date_str'] and item['date_str'] >= cutoff_7d:
            groups[key]['items_7d'] += 1

    # Compute trends and insert
    conn = _get_db_conn()
    c = conn.cursor()
    computed_at = now.strftime('%Y-%m-%d')
    inserted = 0

    for (state, city, topic), counts in groups.items():
        count_7d = counts['items_7d']
        count_30d = counts['items_30d']
        baseline_avg = count_30d / 4.0 if count_30d > 0 else 0
        trend_ratio = count_7d / baseline_avg if baseline_avg > 0 else 0

        if count_7d >= 3 and trend_ratio >= 1.8:
            # Classify
            if trend_ratio >= 2.5:
                classification = 'Accelerating'
            elif count_7d > count_30d / 2:
                classification = 'Peaking'
            else:
                classification = 'Emerging'

            trend_id = str(uuid.uuid4())
            c.execute('''
                INSERT INTO trend_signals (id, state, city, topic, count_7d, count_30d, trend_ratio, classification, computed_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT (state, city, topic, computed_at) DO UPDATE SET
                    id = EXCLUDED.id, count_7d = EXCLUDED.count_7d, count_30d = EXCLUDED.count_30d,
                    trend_ratio = EXCLUDED.trend_ratio, classification = EXCLUDED.classification
            ''', (trend_id, state, city, topic, count_7d, count_30d, round(trend_ratio, 2), classification, computed_at))
            inserted += 1

    conn.commit()
    conn.close()
    print(f"[TrendDetection] Completed. {inserted} trend signals detected from {len(all_items)} total items.")


def generate_weekly_brief():
    """
    Weekly Sunbelt Intelligence Brief generation.
    Pulls trend_signals + recent signals, generates report via Claude.
    No SerpAPI calls.
    """
    print("[WeeklyBrief] Generating Sunbelt Intelligence Brief...")
    now = datetime.utcnow()
    week_end = now.strftime('%Y-%m-%d')
    week_start = (now - timedelta(days=7)).strftime('%Y-%m-%d')

    # 1. Get trend signals from last 7 days
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT state, city, topic, count_7d, count_30d, trend_ratio, classification FROM trend_signals WHERE computed_at >= ? ORDER BY trend_ratio DESC', (week_start,))
    trends = [{'state': r[0], 'city': r[1], 'topic': r[2], 'count_7d': r[3], 'count_30d': r[4], 'trend_ratio': r[5], 'classification': r[6]} for r in c.fetchall()]
    conn.close()

    # 2. Get recent signal summaries
    all_items = _gather_all_signals(days=7)
    # State activity counts
    from collections import Counter
    state_counts = Counter(item['state'] for item in all_items if item['state'])
    capital_events = [i for i in all_items if i['topic'] in ('acquisition', 'sale', 'financing', 'JV', 'recapitalization')]
    construction_events = [i for i in all_items if i['topic'] in ('groundbreaking', 'construction', 'permit', 'rezoning', 'new_build', 'under_construction', 'permit_rezoning')]
    refinancing_events = [i for i in all_items if i['topic'] in ('financing', 'recapitalization', 'refinancing')]

    # Build context for Claude
    context = {
        'week': f"{week_start} to {week_end}",
        'states': ['TX', 'AZ', 'GA', 'NC', 'FL'],
        'state_signal_counts': dict(state_counts),
        'total_signals': len(all_items),
        'trend_signals': trends[:20],
        'capital_events_sample': [{'company': e.get('entity_name', ''), 'city': e['city'], 'state': e['state'], 'topic': e['topic'], 'summary': e.get('summary', e.get('title', ''))} for e in capital_events[:15]],
        'construction_sample': [{'entity': e.get('entity_name', ''), 'city': e['city'], 'state': e['state'], 'topic': e['topic'], 'summary': e.get('summary', e.get('title', ''))} for e in construction_events[:15]],
        'refinancing_sample': [{'entity': e.get('entity_name', ''), 'city': e['city'], 'state': e['state'], 'summary': e.get('summary', e.get('title', ''))} for e in refinancing_events[:10]],
    }

    prompt = f"""You are a senior market analyst for a Build-to-Rent (BTR) and Single-Family Rental (SFR) insurance brokerage covering the Sunbelt states.

Generate the Weekly Sunbelt Risk Intelligence Brief for {context['week']}.

DATA:
{json.dumps(context, indent=2)}

Write a structured report with these EXACT sections (use markdown headers):

## Most Active State
Identify the most active state by signal volume and explain why (specific deals, permits, trends).

## Top 3 Cities by Capital Movement
Rank the top 3 cities seeing the most acquisition/financing/JV activity. Include company names and deal details where available.

## Financing Themes Emerging
What financing patterns are emerging? (credit facilities, institutional capital, JV formations, etc.)

## Construction Acceleration Signals
Where is construction activity accelerating? (groundbreakings, permits, rezoning approvals)

## Refinancing Signals
Note any refinancing or recapitalization signals that indicate portfolio repositioning.

## Insurance Inflection Implications
Tie findings to insurance opportunities for 40-400 unit BTR operators. What coverage needs are emerging? (Builder's Risk transitions, portfolio scale triggers, new state expansions)

## 3 Strategic Conversation Angles
Provide 3 specific, actionable conversation starters a producer could use this week when calling BTR operators.

TONE: Professional, analytical, data-driven. Not salesy. Reference specific companies, cities, and deal details from the data.
If the data is sparse, note that and provide analysis based on what's available.

Return the brief as markdown text."""

    try:
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}]
        )
        brief_text = message.content[0].text if message.content else ''
    except Exception as e:
        print(f"[WeeklyBrief] Claude error: {e}")
        brief_text = f"Brief generation failed: {str(e)}"

    # Store
    brief_id = str(uuid.uuid4())
    brief_data = {
        'text': brief_text,
        'stats': {
            'total_signals': context['total_signals'],
            'state_counts': context['state_signal_counts'],
            'trend_count': len(trends),
            'capital_events': len(capital_events),
            'construction_events': len(construction_events),
        }
    }

    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('INSERT INTO weekly_briefs (id, brief_json, week_start, week_end) VALUES (?, ?, ?, ?)',
              (brief_id, json.dumps(brief_data), week_start, week_end))
    conn.commit()
    conn.close()
    print(f"[WeeklyBrief] Brief generated and stored (id={brief_id})")
    return brief_id


# --- Intelligence API Endpoints ---

@app.route('/api/intelligence/trends', methods=['GET'])
@require_auth
def api_intelligence_trends():
    """Get trend signals with optional state filter."""
    state = request.args.get('state', '').upper()
    topic = request.args.get('topic', '')
    days = min(int(request.args.get('days', 7)), 90)
    cutoff = (datetime.utcnow() - timedelta(days=days)).strftime('%Y-%m-%d')

    conn = _get_db_conn()
    c = conn.cursor()
    query = 'SELECT id, state, city, topic, count_7d, count_30d, trend_ratio, classification, computed_at FROM trend_signals WHERE computed_at >= ?'
    params = [cutoff]

    if state:
        query += ' AND state = ?'
        params.append(state)
    if topic:
        query += ' AND topic = ?'
        params.append(topic)

    query += ' ORDER BY trend_ratio DESC'
    c.execute(query, params)

    trends = []
    for r in c.fetchall():
        trends.append({
            'id': r[0], 'state': r[1], 'city': r[2], 'topic': r[3],
            'count_7d': r[4], 'count_30d': r[5], 'trend_ratio': r[6],
            'classification': r[7], 'computed_at': r[8],
        })
    conn.close()
    return jsonify({'success': True, 'trends': trends})


@app.route('/api/intelligence/briefs', methods=['GET'])
@require_auth
def api_intelligence_briefs():
    """List weekly briefs (latest first)."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT id, generated_at, week_start, week_end FROM weekly_briefs ORDER BY generated_at DESC LIMIT 20')
    briefs = [{'id': r[0], 'generated_at': r[1], 'week_start': r[2], 'week_end': r[3]} for r in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'briefs': briefs})


@app.route('/api/intelligence/briefs/latest', methods=['GET'])
@require_auth
def api_intelligence_briefs_latest():
    """Get the latest weekly brief with full content."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT id, brief_json, generated_at, week_start, week_end FROM weekly_briefs ORDER BY generated_at DESC LIMIT 1')
    row = c.fetchone()
    conn.close()
    if not row:
        return jsonify({'success': True, 'brief': None})
    return jsonify({
        'success': True,
        'brief': {
            'id': row[0],
            'content': json.loads(row[1]) if row[1] else {},
            'generated_at': row[2],
            'week_start': row[3],
            'week_end': row[4],
        }
    })


@app.route('/api/intelligence/briefs/<brief_id>', methods=['GET'])
@require_auth
def api_intelligence_brief_detail(brief_id):
    """Get a specific weekly brief."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('SELECT id, brief_json, generated_at, week_start, week_end FROM weekly_briefs WHERE id = ?', (brief_id,))
    row = c.fetchone()
    conn.close()
    if not row:
        return jsonify({'success': False, 'message': 'Brief not found'}), 404
    return jsonify({
        'success': True,
        'brief': {
            'id': row[0],
            'content': json.loads(row[1]) if row[1] else {},
            'generated_at': row[2],
            'week_start': row[3],
            'week_end': row[4],
        }
    })


@app.route('/api/intelligence/briefs/generate', methods=['POST'])
@require_auth
@require_role('admin')
def api_intelligence_generate_brief():
    """Admin: manually trigger weekly brief generation."""
    try:
        # Run trend detection first
        run_trend_detection()
        brief_id = generate_weekly_brief()
        return jsonify({'success': True, 'brief_id': brief_id})
    except Exception as e:
        print(f"[Intelligence] Brief generation error: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/intelligence/trends/run', methods=['POST'])
@require_auth
@require_role('admin')
def api_intelligence_run_trends():
    """Admin: manually trigger trend detection."""
    try:
        run_trend_detection()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/intelligence/state-rankings', methods=['GET'])
@require_auth
def api_intelligence_state_rankings():
    """Enhanced state rankings with trend strength included."""
    from serpapi_client import get_cached
    today = datetime.utcnow().strftime('%Y-%m-%d')

    # Get trend counts per state
    conn = _get_db_conn()
    c = conn.cursor()
    cutoff_7d = (datetime.utcnow() - timedelta(days=7)).strftime('%Y-%m-%d')
    c.execute('SELECT state, COUNT(*), MAX(trend_ratio) FROM trend_signals WHERE computed_at >= ? GROUP BY state', (cutoff_7d,))
    trend_data = {r[0]: {'trend_count': r[1], 'max_trend_ratio': r[2]} for r in c.fetchall()}
    conn.close()

    rankings = []
    for state_abbr in STATEWIDE_STATES:
        result_cache_key = f"statewide_result:{state_abbr.lower()}:{today}"
        items = get_cached(result_cache_key) or []

        capital_count = sum(1 for i in items if i.get('event_type') in ('acquisition', 'sale', 'financing', 'JV'))
        construction_count = sum(1 for i in items if i.get('event_type') in ('groundbreaking', 'construction', 'permit', 'rezoning'))

        scores = _compute_activity_scores(items, days=7)
        total_score = sum(s['activity_score'] for s in scores.values())
        top_cities = sorted(scores.values(), key=lambda x: x['activity_score'], reverse=True)[:3]

        td = trend_data.get(state_abbr, {})
        rankings.append({
            'state': state_abbr,
            'state_name': STATEWIDE_STATES[state_abbr],
            'state_activity_score': round(total_score, 1),
            'total_signals': len(items),
            'capital_events_count': capital_count,
            'construction_signals_count': construction_count,
            'top_cities': top_cities,
            'trend_count': td.get('trend_count', 0),
            'max_trend_ratio': td.get('max_trend_ratio', 0),
            'last_updated': today if items else None,
        })

    rankings.sort(key=lambda x: x['state_activity_score'], reverse=True)
    return jsonify({'success': True, 'rankings': rankings})


# --- Momentum & Call Timing API Endpoints ---

@app.route('/api/intelligence/momentum', methods=['GET'])
@require_auth
def api_intelligence_momentum():
    """Get market momentum data. Optional state filter."""
    state = request.args.get('state', '').upper()
    conn = _get_db_conn()
    c = conn.cursor()
    if state:
        c.execute('''
            SELECT state, city, momentum_score, momentum_label, signals_7d, signals_30d,
                   weighted_signals_7d, weighted_signals_30d, window_end_date
            FROM market_momentum WHERE state = ?
            ORDER BY momentum_score DESC
        ''', (state,))
    else:
        c.execute('''
            SELECT state, city, momentum_score, momentum_label, signals_7d, signals_30d,
                   weighted_signals_7d, weighted_signals_30d, window_end_date
            FROM market_momentum
            ORDER BY momentum_score DESC LIMIT 50
        ''')
    rows = c.fetchall()
    conn.close()
    markets = [{
        'state': r[0], 'city': r[1], 'momentum_score': r[2], 'momentum_label': r[3],
        'signals_7d': r[4], 'signals_30d': r[5],
        'weighted_signals_7d': r[6], 'weighted_signals_30d': r[7],
        'window_end_date': r[8]
    } for r in rows]
    return jsonify({'success': True, 'markets': markets})


@app.route('/api/intelligence/momentum/top', methods=['GET'])
@require_auth
def api_intelligence_momentum_top():
    """Get top 10 cities by momentum_score."""
    conn = _get_db_conn()
    c = conn.cursor()
    c.execute('''
        SELECT state, city, momentum_score, momentum_label, weighted_signals_7d, weighted_signals_30d
        FROM market_momentum
        ORDER BY momentum_score DESC LIMIT 10
    ''')
    rows = c.fetchall()
    conn.close()
    cities = [{
        'state': r[0], 'city': r[1], 'momentum_score': r[2], 'momentum_label': r[3],
        'weighted_signals_7d': r[4], 'weighted_signals_30d': r[5]
    } for r in rows]
    return jsonify({'success': True, 'cities': cities})


@app.route('/api/intelligence/call-timing', methods=['GET'])
@require_auth
def api_intelligence_call_timing():
    """Get call timing scores. Optional filters: label, state, limit."""
    label = request.args.get('label', '')
    state = request.args.get('state', '').upper()
    limit = min(int(request.args.get('limit', 50)), 200)

    conn = _get_db_conn()
    c = conn.cursor()
    query = 'SELECT * FROM lead_timing_scores WHERE 1=1'
    params = []
    if label:
        query += ' AND timing_label = ?'
        params.append(label)
    if state:
        query += ' AND state = ?'
        params.append(state)
    query += ' ORDER BY call_timing_score DESC LIMIT ?'
    params.append(limit)

    c.execute(query, params)
    cols = [desc[0] for desc in c.description]
    rows = c.fetchall()
    conn.close()

    scores = []
    for r in rows:
        d = dict(zip(cols, r))
        d['reasons'] = json.loads(d['reasons']) if d.get('reasons') else []
        scores.append(d)
    return jsonify({'success': True, 'scores': scores})


@app.route('/api/intelligence/call-timing/lookup', methods=['GET'])
@require_auth
def api_intelligence_call_timing_lookup():
    """Lookup call timing for a specific prospect_key (or bulk via comma-separated keys)."""
    keys = request.args.get('keys', '')
    if not keys:
        return jsonify({'success': True, 'scores': {}})

    key_list = [k.strip() for k in keys.split(',') if k.strip()]
    conn = _get_db_conn()
    c = conn.cursor()
    placeholders = ','.join('?' * len(key_list))
    c.execute(f'''
        SELECT prospect_key, call_timing_score, timing_label, reasons,
               trigger_severity, swim_lane_fit, engagement_score,
               market_momentum_score, freshness_score
        FROM lead_timing_scores WHERE prospect_key IN ({placeholders})
    ''', key_list)
    rows = c.fetchall()
    conn.close()

    result = {}
    for r in rows:
        result[r[0]] = {
            'call_timing_score': r[1], 'timing_label': r[2],
            'reasons': json.loads(r[3]) if r[3] else [],
            'trigger_severity': r[4], 'swim_lane_fit': r[5],
            'engagement_score': r[6], 'market_momentum_score': r[7],
            'freshness_score': r[8]
        }
    return jsonify({'success': True, 'scores': result})


@app.route('/api/intelligence/optimization/run', methods=['POST'])
@require_auth
@require_role('admin')
def api_intelligence_run_optimization():
    """Admin: manually trigger the full optimization pipeline."""
    try:
        run_daily_optimization()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/intelligence/backfill-weights', methods=['POST'])
@require_auth
@require_role('admin')
def api_intelligence_backfill_weights():
    """Admin: one-time backfill of weighted_signals for last 90 days."""
    try:
        count = materialize_weighted_signals(days=90)
        return jsonify({'success': True, 'materialized': count})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# --- Scoped momentum API (spec: /api/intel/) ---

@app.route('/api/intel/momentum/top-cities', methods=['GET'])
@require_auth
def api_intel_momentum_top_cities():
    """
    Top cities by momentum_score for the latest window_end_date.
    Params: ?state=TX&limit=10
    """
    state = request.args.get('state', '').upper()
    limit = min(int(request.args.get('limit', 10)), 50)

    conn = _get_db_conn()
    c = conn.cursor()

    # Get latest window_end_date
    c.execute('SELECT MAX(window_end_date) FROM market_momentum')
    row = c.fetchone()
    latest_date = row[0] if row and row[0] else None

    if not latest_date:
        conn.close()
        return jsonify({'success': True, 'cities': [], 'window_end_date': None})

    query = '''
        SELECT state, city, momentum_score, momentum_label,
               signals_7d, signals_14d, signals_30d,
               weighted_signals_7d, weighted_signals_14d, weighted_signals_30d,
               window_end_date
        FROM market_momentum
        WHERE window_end_date = ?
    '''
    params = [latest_date]

    if state:
        query += ' AND state = ?'
        params.append(state)

    query += ' ORDER BY momentum_score DESC LIMIT ?'
    params.append(limit)

    c.execute(query, params)
    rows = c.fetchall()
    conn.close()

    cities = [{
        'state': r[0], 'city': r[1], 'momentum_score': r[2], 'momentum_label': r[3],
        'signals_7d': r[4], 'signals_14d': r[5], 'signals_30d': r[6],
        'weighted_signals_7d': r[7], 'weighted_signals_14d': r[8], 'weighted_signals_30d': r[9],
        'window_end_date': r[10]
    } for r in rows]
    return jsonify({'success': True, 'cities': cities, 'window_end_date': latest_date})


@app.route('/api/intel/momentum/state-summary', methods=['GET'])
@require_auth
def api_intel_momentum_state_summary():
    """
    Latest-day summary per state (TX/AZ/GA/NC/FL):
    total weighted_signals_7d and avg momentum_score.
    """
    conn = _get_db_conn()
    c = conn.cursor()

    # Get latest window_end_date
    c.execute('SELECT MAX(window_end_date) FROM market_momentum')
    row = c.fetchone()
    latest_date = row[0] if row and row[0] else None

    if not latest_date:
        conn.close()
        return jsonify({'success': True, 'states': [], 'window_end_date': None})

    c.execute('''
        SELECT state,
               SUM(weighted_signals_7d) as total_weighted_7d,
               SUM(weighted_signals_30d) as total_weighted_30d,
               SUM(signals_7d) as total_signals_7d,
               SUM(signals_30d) as total_signals_30d,
               ROUND(AVG(momentum_score), 1) as avg_momentum,
               COUNT(*) as city_count
        FROM market_momentum
        WHERE window_end_date = ? AND state IN ('TX','AZ','GA','NC','FL')
        GROUP BY state
        ORDER BY total_weighted_7d DESC
    ''', (latest_date,))
    rows = c.fetchall()
    conn.close()

    states = [{
        'state': r[0],
        'weighted_signals_7d': r[1],
        'weighted_signals_30d': r[2],
        'signals_7d': r[3],
        'signals_30d': r[4],
        'avg_momentum_score': r[5],
        'city_count': r[6],
    } for r in rows]
    return jsonify({'success': True, 'states': states, 'window_end_date': latest_date})


# ===================================================================
# SUNBELT INTELLIGENCE — Phase 1 (Discovery-only endpoints)
# All metrics computed live from _gather_all_signals() — no joins to
# prospect search, pipeline, or pre-materialized tables.
# ===================================================================

def _sunbelt_meta(params, counts, generated_at=None):
    """Standard meta block for sunbelt responses."""
    return {
        'generated_at': generated_at or datetime.utcnow().isoformat(),
        'params': params,
        'source': 'discovery_only',
        'counts': counts,
    }


@app.route('/api/sunbelt/weekly', methods=['GET'])
@require_auth
def api_sunbelt_weekly():
    """Weekly brief computed live from discovery signals."""
    import time
    t0 = time.time()
    try:
        window_days = min(int(request.args.get('windowDays', 7)), 90)
    except (ValueError, TypeError):
        window_days = 7

    try:
        signals = _gather_all_signals(days=window_days)
        elapsed = round(time.time() - t0, 3)
        app.logger.info(f'[sunbelt/weekly] {len(signals)} signals gathered in {elapsed}s (window={window_days}d)')

        if not signals:
            return jsonify({
                'ok': True,
                'data': {
                    'top_markets': [], 'top_topics': [], 'highlights': [],
                    'brief_text': 'No discovery signals in selected window.',
                },
                'meta': _sunbelt_meta({'windowDays': window_days}, {'total_signals': 0}),
            })

        # Top markets by weighted signal count
        market_counts = {}
        topic_counts = {}
        for s in signals:
            city = s.get('city') or None
            state = s.get('state') or None
            key = f"{city or 'Unknown'}, {state or '??'}"
            w = get_signal_weight(s.get('topic', '') + ' ' + s.get('title', ''))
            market_counts[key] = market_counts.get(key, 0) + w
            topic = s.get('topic', 'other')
            topic_counts[topic] = topic_counts.get(topic, 0) + 1

        top_markets = sorted(market_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        top_topics = sorted(topic_counts.items(), key=lambda x: x[1], reverse=True)[:10]

        # Highlights: top 10 signals by weight
        scored = []
        for s in signals:
            w = get_signal_weight(s.get('topic', '') + ' ' + s.get('title', ''))
            scored.append((w, s))
        scored.sort(key=lambda x: x[0], reverse=True)

        highlights = []
        for w, s in scored[:10]:
            highlights.append({
                'title': s.get('title') or s.get('summary', '')[:80] or 'Untitled',
                'url': s.get('url') or None,
                'city': s.get('city') or None,
                'state': s.get('state') or None,
                'topic': s.get('topic', 'other'),
                'date': s.get('date_str', ''),
                'weight': w,
            })

        # Brief text summary
        top_mkt_names = [m[0] for m in top_markets[:5]]
        top_topic_names = [t[0] for t in top_topics[:5]]
        brief_text = (
            f"Weekly Sunbelt Intelligence ({window_days}-day window): "
            f"{len(signals)} discovery signals across {len(market_counts)} markets. "
            f"Top markets: {', '.join(top_mkt_names)}. "
            f"Dominant topics: {', '.join(top_topic_names)}."
        )

        return jsonify({
            'ok': True,
            'data': {
                'top_markets': [{'market': m[0], 'weighted_signals': m[1]} for m in top_markets],
                'top_topics': [{'topic': t[0], 'count': t[1]} for t in top_topics],
                'highlights': highlights,
                'brief_text': brief_text,
            },
            'meta': _sunbelt_meta({'windowDays': window_days},
                                  {'total_signals': len(signals), 'markets': len(market_counts), 'topics': len(topic_counts)}),
        })
    except Exception as e:
        app.logger.error(f'[sunbelt/weekly] Error: {e}')
        return jsonify({'ok': False, 'error': 'Failed to generate weekly brief', 'details': str(e)}), 500


@app.route('/api/sunbelt/momentum', methods=['GET'])
@require_auth
def api_sunbelt_momentum():
    """Momentum index computed live from discovery signals with defensive normalization."""
    import time
    t0 = time.time()
    try:
        window_days = min(int(request.args.get('windowDays', 7)), 90)
        baseline_days = min(int(request.args.get('baselineDays', 30)), 180)
    except (ValueError, TypeError):
        window_days, baseline_days = 7, 30

    try:
        signals = _gather_all_signals(days=baseline_days)
        elapsed = round(time.time() - t0, 3)
        app.logger.info(f'[sunbelt/momentum] {len(signals)} signals gathered in {elapsed}s (window={window_days}d, baseline={baseline_days}d)')

        if not signals:
            return jsonify({
                'ok': True,
                'data': {'markets': [], 'message': 'No discovery signals in selected window.'},
                'meta': _sunbelt_meta({'windowDays': window_days, 'baselineDays': baseline_days}, {'total_signals': 0}),
            })

        window_cutoff = (datetime.utcnow() - timedelta(days=window_days)).isoformat()

        # Aggregate by city
        city_data = {}  # key: "city|state" -> {count_window, count_baseline, weighted_window, weighted_baseline}
        for s in signals:
            city = s.get('city') or None
            state = s.get('state') or None
            if not city or not state:
                continue
            key = f"{city}|{state}"
            if key not in city_data:
                city_data[key] = {'city': city, 'state': state, 'count_window': 0, 'count_baseline': 0,
                                  'weighted_window': 0, 'weighted_baseline': 0}
            w = get_signal_weight(s.get('topic', '') + ' ' + s.get('title', ''))
            d = city_data[key]
            d['count_baseline'] += 1
            d['weighted_baseline'] += w
            if s.get('date_str', '') >= window_cutoff:
                d['count_window'] += 1
                d['weighted_window'] += w

        # Compute raw scores with defensive math
        raw_scores = []
        for key, d in city_data.items():
            ratio = d['count_window'] / max(1, d['count_baseline'])
            raw = (d['count_window'] * 2) + (ratio * 10)
            raw_scores.append((key, d, raw, ratio))

        # Normalize 0-100 across all cities
        if not raw_scores:
            return jsonify({
                'ok': True,
                'data': {'markets': [], 'message': 'No city-level signals found.'},
                'meta': _sunbelt_meta({'windowDays': window_days, 'baselineDays': baseline_days},
                                      {'total_signals': len(signals)}),
            })

        max_raw = max(r[2] for r in raw_scores)
        min_raw = min(r[2] for r in raw_scores)

        markets = []
        for key, d, raw, ratio in raw_scores:
            if max_raw == min_raw:
                score = 50.0
            else:
                score = round(((raw - min_raw) / (max_raw - min_raw)) * 100, 1)

            if ratio >= 0.5:
                label = 'accelerating'
            elif ratio >= 0.2:
                label = 'steady'
            else:
                label = 'cooling'

            markets.append({
                'city': d['city'],
                'state': d['state'],
                'momentum_score': score,
                'momentum_label': label,
                'signals_window': d['count_window'],
                'signals_baseline': d['count_baseline'],
                'weighted_window': d['weighted_window'],
                'weighted_baseline': d['weighted_baseline'],
                'ratio': round(ratio, 4),
            })

        markets.sort(key=lambda x: x['momentum_score'], reverse=True)
        markets = markets[:50]

        return jsonify({
            'ok': True,
            'data': {'markets': markets},
            'meta': _sunbelt_meta({'windowDays': window_days, 'baselineDays': baseline_days},
                                  {'total_signals': len(signals), 'cities_scored': len(raw_scores)}),
        })
    except Exception as e:
        app.logger.error(f'[sunbelt/momentum] Error: {e}')
        return jsonify({'ok': False, 'error': 'Failed to compute momentum', 'details': str(e)}), 500


@app.route('/api/sunbelt/trends', methods=['GET'])
@require_auth
def api_sunbelt_trends():
    """Trend signals computed live from discovery data with example items."""
    import time
    t0 = time.time()
    try:
        window_days = min(int(request.args.get('windowDays', 7)), 90)
        baseline_days = min(int(request.args.get('baselineDays', 30)), 180)
    except (ValueError, TypeError):
        window_days, baseline_days = 7, 30
    topic_filter = request.args.get('topic', '').strip().lower()

    try:
        signals = _gather_all_signals(days=baseline_days)
        elapsed = round(time.time() - t0, 3)
        app.logger.info(f'[sunbelt/trends] {len(signals)} signals gathered in {elapsed}s (window={window_days}d, baseline={baseline_days}d)')

        if not signals:
            return jsonify({
                'ok': True,
                'data': {'trends': [], 'message': 'No discovery signals in selected window.'},
                'meta': _sunbelt_meta({'windowDays': window_days, 'baselineDays': baseline_days, 'topic': topic_filter or None},
                                      {'total_signals': 0}),
            })

        window_cutoff = (datetime.utcnow() - timedelta(days=window_days)).isoformat()

        # Group by (state, city, topic)
        groups = {}  # key -> {count_window, count_baseline, items_window}
        for s in signals:
            state = s.get('state') or None
            city = s.get('city') or None
            topic = s.get('topic', 'other')
            if topic_filter and topic.lower() != topic_filter:
                continue
            key = f"{state or ''}|{city or ''}|{topic}"
            if key not in groups:
                groups[key] = {'state': state, 'city': city, 'topic': topic,
                               'count_window': 0, 'count_baseline': 0, 'items_window': []}
            g = groups[key]
            g['count_baseline'] += 1
            if s.get('date_str', '') >= window_cutoff:
                g['count_window'] += 1
                if len(g['items_window']) < 5:
                    g['items_window'].append({
                        'title': s.get('title') or s.get('summary', '')[:80] or 'Untitled',
                        'source_url': s.get('url') or None,
                        'city': city,
                        'state': state,
                        'date': s.get('date_str', ''),
                    })

        # Compute trend ratio and classify
        trends = []
        for key, g in groups.items():
            if g['count_window'] < 2:
                continue
            baseline_weekly = g['count_baseline'] / max(1, baseline_days / 7.0)
            trend_ratio = g['count_window'] / max(1, baseline_weekly) if baseline_weekly > 0 else g['count_window']

            if trend_ratio >= 2.5:
                classification = 'Accelerating'
            elif trend_ratio >= 1.5:
                classification = 'Emerging'
            elif g['count_window'] > g['count_baseline'] / 2:
                classification = 'Peaking'
            else:
                classification = 'Cooling'

            trends.append({
                'state': g['state'],
                'city': g['city'],
                'topic': g['topic'],
                'count_window': g['count_window'],
                'count_baseline': g['count_baseline'],
                'trend_ratio': round(trend_ratio, 2),
                'classification': classification,
                'examples': g['items_window'],
            })

        trends.sort(key=lambda x: x['trend_ratio'], reverse=True)

        return jsonify({
            'ok': True,
            'data': {'trends': trends},
            'meta': _sunbelt_meta({'windowDays': window_days, 'baselineDays': baseline_days, 'topic': topic_filter or None},
                                  {'total_signals': len(signals), 'trends_found': len(trends)}),
        })
    except Exception as e:
        app.logger.error(f'[sunbelt/trends] Error: {e}')
        return jsonify({'ok': False, 'error': 'Failed to compute trends', 'details': str(e)}), 500


@app.route('/api/sunbelt/state-rankings', methods=['GET'])
@require_auth
def api_sunbelt_state_rankings():
    """State rankings computed live from discovery signals. Caps: 50 states, 10 cities, 10 topics."""
    import time
    t0 = time.time()
    try:
        window_days = min(int(request.args.get('windowDays', 30)), 180)
    except (ValueError, TypeError):
        window_days = 30

    try:
        signals = _gather_all_signals(days=window_days)
        elapsed = round(time.time() - t0, 3)
        app.logger.info(f'[sunbelt/state-rankings] {len(signals)} signals gathered in {elapsed}s (window={window_days}d)')

        if not signals:
            return jsonify({
                'ok': True,
                'data': {'rankings': [], 'message': 'No discovery signals in selected window.'},
                'meta': _sunbelt_meta({'windowDays': window_days}, {'total_signals': 0}),
            })

        # Aggregate by state
        state_data = {}  # state -> {signals, cities, topics, capital_count, construction_count}
        for s in signals:
            state = s.get('state') or None
            if not state:
                continue
            if state not in state_data:
                state_data[state] = {'state': state, 'total': 0, 'weighted': 0,
                                     'cities': {}, 'topics': {}, 'capital': 0, 'construction': 0}
            sd = state_data[state]
            w = get_signal_weight(s.get('topic', '') + ' ' + s.get('title', ''))
            sd['total'] += 1
            sd['weighted'] += w

            city = s.get('city') or 'Unknown'
            sd['cities'][city] = sd['cities'].get(city, 0) + w

            topic = s.get('topic', 'other')
            sd['topics'][topic] = sd['topics'].get(topic, 0) + 1

            if topic in ('acquisition', 'sale', 'financing', 'jv', 'joint_venture', 'capital'):
                sd['capital'] += 1
            if topic in ('groundbreaking', 'construction', 'permit', 'rezoning', 'new_build', 'permit_rezoning'):
                sd['construction'] += 1

        # Build rankings
        rankings = []
        for state, sd in state_data.items():
            top_cities = sorted(sd['cities'].items(), key=lambda x: x[1], reverse=True)[:10]
            top_topics = sorted(sd['topics'].items(), key=lambda x: x[1], reverse=True)[:10]
            rankings.append({
                'state': state,
                'state_name': STATEWIDE_STATES.get(state, state),
                'total_signals': sd['total'],
                'weighted_signals': sd['weighted'],
                'capital_events': sd['capital'],
                'construction_signals': sd['construction'],
                'top_cities': [{'city': c[0], 'weighted_score': c[1]} for c in top_cities],
                'top_topics': [{'topic': t[0], 'count': t[1]} for t in top_topics],
            })

        rankings.sort(key=lambda x: x['weighted_signals'], reverse=True)
        rankings = rankings[:50]

        return jsonify({
            'ok': True,
            'data': {'rankings': rankings},
            'meta': _sunbelt_meta({'windowDays': window_days},
                                  {'total_signals': len(signals), 'states_ranked': len(rankings)}),
        })
    except Exception as e:
        app.logger.error(f'[sunbelt/state-rankings] Error: {e}')
        return jsonify({'ok': False, 'error': 'Failed to compute state rankings', 'details': str(e)}), 500


# --- Sunbelt AI Sparknotes ---

_TAB_WINDOW_DEFAULTS = {'weekly': 7, 'momentum': 7, 'trends': 7, 'state_rankings': 30}

def _build_sparknotes_items(tab, window_days):
    """Build the item list for sparknotes from the same data the UI renders."""
    signals = _gather_all_signals(days=window_days)
    if not signals:
        return []

    if tab == 'weekly':
        # Top highlights by weight (same logic as /api/sunbelt/weekly)
        scored = []
        for s in signals:
            w = get_signal_weight(s.get('topic', '') + ' ' + s.get('title', ''))
            scored.append((w, s))
        scored.sort(key=lambda x: x[0], reverse=True)
        return [{
            'id': str(i),
            'title': s.get('title') or s.get('summary', '')[:80] or 'Untitled',
            'snippet_or_summary': s.get('summary') or s.get('title') or '',
            'topic': s.get('topic', 'other'),
            'city': s.get('city') or None,
            'state': s.get('state') or None,
            'source_url': s.get('url') or None,
            'created_at': s.get('date_str', ''),
        } for i, (w, s) in enumerate(scored[:15])]

    elif tab == 'momentum':
        # Top markets by momentum (same logic as /api/sunbelt/momentum)
        window_cutoff = (datetime.utcnow() - timedelta(days=window_days)).isoformat()
        city_data = {}
        for s in signals:
            city = s.get('city') or None
            state = s.get('state') or None
            if not city or not state:
                continue
            key = f"{city}|{state}"
            if key not in city_data:
                city_data[key] = {'city': city, 'state': state, 'count_window': 0, 'count_baseline': 0,
                                  'titles': [], 'topics': set()}
            d = city_data[key]
            d['count_baseline'] += 1
            d['topics'].add(s.get('topic', 'other'))
            if s.get('date_str', '') >= window_cutoff:
                d['count_window'] += 1
                if len(d['titles']) < 3:
                    d['titles'].append(s.get('title') or s.get('summary', '')[:80] or 'Untitled')
        ranked = sorted(city_data.values(), key=lambda d: d['count_window'], reverse=True)[:15]
        return [{
            'id': str(i),
            'title': f"{d['city']}, {d['state']} — {d['count_window']} signals in {window_days}d",
            'snippet_or_summary': '; '.join(d['titles']) if d['titles'] else 'Market activity detected',
            'topic': ', '.join(sorted(d['topics'])[:3]),
            'city': d['city'],
            'state': d['state'],
            'source_url': None,
            'created_at': '',
        } for i, d in enumerate(ranked)]

    elif tab == 'trends':
        # Trend groups (same logic as /api/sunbelt/trends)
        baseline_days = max(window_days * 4, 30)
        all_sigs = _gather_all_signals(days=baseline_days)
        window_cutoff = (datetime.utcnow() - timedelta(days=window_days)).isoformat()
        groups = {}
        for s in all_sigs:
            state = s.get('state') or None
            city = s.get('city') or None
            topic = s.get('topic', 'other')
            key = f"{state or ''}|{city or ''}|{topic}"
            if key not in groups:
                groups[key] = {'state': state, 'city': city, 'topic': topic,
                               'count_window': 0, 'count_baseline': 0, 'title': ''}
            g = groups[key]
            g['count_baseline'] += 1
            if s.get('date_str', '') >= window_cutoff:
                g['count_window'] += 1
                if not g['title']:
                    g['title'] = s.get('title') or s.get('summary', '')[:80] or 'Untitled'
        trends = []
        for key, g in groups.items():
            if g['count_window'] < 2:
                continue
            baseline_weekly = g['count_baseline'] / max(1, baseline_days / 7.0)
            ratio = g['count_window'] / max(1, baseline_weekly) if baseline_weekly > 0 else g['count_window']
            if ratio >= 2.5:
                classification = 'Accelerating'
            elif ratio >= 1.5:
                classification = 'Emerging'
            elif g['count_window'] > g['count_baseline'] / 2:
                classification = 'Peaking'
            else:
                classification = 'Cooling'
            trends.append({**g, 'ratio': ratio, 'classification': classification})
        trends.sort(key=lambda x: x['ratio'], reverse=True)
        return [{
            'id': str(i),
            'title': f"{t['city'] or 'Unknown'}, {t['state'] or '??'} — {t['topic']} ({t['classification']})",
            'snippet_or_summary': f"{t['count_window']} signals in {window_days}d vs {t['count_baseline']} baseline. Ratio: {t['ratio']:.1f}x. {t['title']}",
            'topic': t['topic'],
            'city': t['city'],
            'state': t['state'],
            'source_url': None,
            'created_at': '',
        } for i, t in enumerate(trends[:15])]

    elif tab == 'state_rankings':
        # State rankings (same logic as /api/sunbelt/state-rankings)
        state_data = {}
        for s in signals:
            state = s.get('state') or None
            if not state:
                continue
            if state not in state_data:
                state_data[state] = {'state': state, 'total': 0, 'weighted': 0,
                                     'cities': {}, 'topics': {}, 'capital': 0, 'construction': 0}
            sd = state_data[state]
            w = get_signal_weight(s.get('topic', '') + ' ' + s.get('title', ''))
            sd['total'] += 1
            sd['weighted'] += w
            city = s.get('city') or 'Unknown'
            sd['cities'][city] = sd['cities'].get(city, 0) + w
            topic = s.get('topic', 'other')
            sd['topics'][topic] = sd['topics'].get(topic, 0) + 1
            if topic in ('acquisition', 'sale', 'financing', 'jv', 'joint_venture', 'capital'):
                sd['capital'] += 1
            if topic in ('groundbreaking', 'construction', 'permit', 'rezoning', 'new_build', 'permit_rezoning'):
                sd['construction'] += 1
        rankings = sorted(state_data.values(), key=lambda d: d['weighted'], reverse=True)[:15]
        return [{
            'id': str(i),
            'title': f"{STATEWIDE_STATES.get(sd['state'], sd['state'])} — {sd['total']} signals, weight {sd['weighted']}",
            'snippet_or_summary': f"Capital events: {sd['capital']}, Construction: {sd['construction']}. Top cities: {', '.join(c for c, _ in sorted(sd['cities'].items(), key=lambda x: x[1], reverse=True)[:3])}",
            'topic': ', '.join(t for t, _ in sorted(sd['topics'].items(), key=lambda x: x[1], reverse=True)[:3]),
            'city': None,
            'state': sd['state'],
            'source_url': None,
            'created_at': '',
        } for i, sd in enumerate(rankings)]

    return []


@app.route('/api/sunbelt/sparknotes', methods=['POST'])
@require_auth
def api_sunbelt_sparknotes():
    """Generate AI Sparknotes summary for Sunbelt tab items."""
    try:
        return _do_sparknotes()
    except Exception as e:
        app.logger.error(f'[sparknotes] Unhandled error: {e}\n{traceback.format_exc()}')
        return jsonify({'ok': False, 'error': f'Internal error: {str(e)}'}), 500


def _do_sparknotes():
    body = request.get_json(silent=True) or {}
    tab = body.get('tab', 'weekly')
    if tab not in _TAB_WINDOW_DEFAULTS:
        return jsonify({'ok': False, 'error': f'Invalid tab: {tab}. Must be one of: {", ".join(_TAB_WINDOW_DEFAULTS.keys())}'}), 400

    try:
        window_days = min(int(body.get('windowDays', _TAB_WINDOW_DEFAULTS[tab])), 180)
    except (ValueError, TypeError):
        window_days = _TAB_WINDOW_DEFAULTS[tab]

    date_bucket = datetime.utcnow().strftime('%Y-%m-%d')

    # Check cache (use separate connection so failures don't poison later queries)
    try:
        cache_conn = _get_db_conn()
        cache_c = cache_conn.cursor()
        cache_c.execute('SELECT payload_json, created_at FROM sunbelt_summaries WHERE tab = ? AND window_days = ? AND date_bucket = ?',
                        (tab, window_days, date_bucket))
        cached = cache_c.fetchone()
        cache_conn.close()
        if cached:
            payload = json.loads(cached[0])
            gen_at = cached[1]
            if isinstance(gen_at, datetime):
                gen_at = gen_at.isoformat()
            return jsonify({
                'ok': True,
                'data': payload,
                'meta': {'cached': True, 'generated_at': str(gen_at)},
            })
    except Exception as e:
        app.logger.debug(f'[sparknotes] Cache check skipped: {e}')
        try:
            cache_conn.close()
        except Exception:
            pass

    # Build items for the prompt
    try:
        items = _build_sparknotes_items(tab, window_days)
    except Exception as e:
        app.logger.error(f'[sparknotes] Failed to build items for tab={tab}: {e}')
        return jsonify({'ok': False, 'error': 'Failed to gather signal data', 'details': str(e)}), 500

    if not items:
        return jsonify({
            'ok': False,
            'error': f'No signals found for tab "{tab}" in the last {window_days} days. Run discovery first.',
        }), 404

    # Build LLM prompt
    prompt = f"""You are a senior market analyst for a Build-to-Rent (BTR) and Single-Family Rental (SFR) insurance brokerage.

Analyze the following {len(items)} intelligence items from the "{tab}" view ({window_days}-day window) and produce a structured Sparknotes summary.

ITEMS:
{json.dumps(items, indent=2)}

Return ONLY valid JSON (no markdown fencing, no explanation) matching this EXACT schema:

{{
  "executive_summary": "3-7 bullet points as a single string, each bullet on its own line starting with '- '. Concise, brokerage tone. Focus on actionable intelligence for BTR/SFR insurance producers.",
  "key_themes": ["theme1", "theme2", "...up to 6 themes"],
  "sparknotes_by_item": [
    {{
      "id": "<matching item id>",
      "one_liner": "1 sentence distilling the item",
      "bullets": ["key point 1", "key point 2"],
      "why_it_matters": "1 sentence on insurance/brokerage relevance",
      "suggested_next_step": "1 sentence actionable next step for a producer"
    }}
  ]
}}

RULES:
- sparknotes_by_item MUST have one entry per input item, in the same order, with matching "id" values.
- executive_summary should synthesize themes across ALL items, not just repeat them.
- key_themes should be short labels (2-5 words each).
- Keep bullet points under 25 words each.
- Tone: professional, analytical, data-driven. Not salesy.
- If data is sparse, note that and provide analysis based on what's available."""

    try:
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = message.content[0].text if message.content else ''
    except Exception as e:
        app.logger.error(f'[sparknotes] LLM call failed: {e}')
        return jsonify({'ok': False, 'error': 'AI summary generation failed. Please try again.', 'details': str(e)}), 500

    # Parse JSON from response (strip markdown fences if present)
    raw = raw.strip()
    if raw.startswith('```'):
        raw = raw.split('\n', 1)[1] if '\n' in raw else raw[3:]
        if raw.endswith('```'):
            raw = raw[:-3]
        raw = raw.strip()

    try:
        payload = json.loads(raw)
    except json.JSONDecodeError as e:
        app.logger.error(f'[sparknotes] JSON parse failed: {e}\nRaw: {raw[:500]}')
        return jsonify({'ok': False, 'error': 'AI returned invalid format. Please try again.'}), 500

    # Validate required keys
    if not isinstance(payload.get('executive_summary'), str):
        payload['executive_summary'] = ''
    if not isinstance(payload.get('key_themes'), list):
        payload['key_themes'] = []
    if not isinstance(payload.get('sparknotes_by_item'), list):
        payload['sparknotes_by_item'] = []

    generated_at = datetime.utcnow().isoformat()

    # Cache the result (separate connection — non-critical)
    try:
        conn = _get_db_conn()
        c = conn.cursor()
        summary_id = str(uuid.uuid4())
        c.execute('INSERT INTO sunbelt_summaries (id, tab, window_days, date_bucket, payload_json, created_at) VALUES (?, ?, ?, ?, ?, ?)',
                  (summary_id, tab, window_days, date_bucket, json.dumps(payload), generated_at))
        conn.commit()
        conn.close()
    except _IntegrityError:
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
    except Exception as e:
        app.logger.warning(f'[sparknotes] Cache write failed: {e}')
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass

    return jsonify({
        'ok': True,
        'data': payload,
        'meta': {'cached': False, 'generated_at': generated_at},
    })


# --- Government Signal Enrichment (Phase 1) ---

# City-specific gov data source configs: what to search per city
_GOV_SIGNAL_CITIES = [
    {'city': 'Phoenix', 'state': 'AZ'},
    {'city': 'Dallas', 'state': 'TX'},
    {'city': 'Atlanta', 'state': 'GA'},
    {'city': 'Charlotte', 'state': 'NC'},
]

_GOV_SIGNAL_QUERIES = {
    'permit': '{city} {state} building permit multifamily residential',
    'zoning': '{city} {state} rezoning multifamily residential development',
    'deed': '{city} {state} deed transfer multifamily apartment',
    'mortgage': '{city} {state} commercial mortgage multifamily',
    'llc': '{city} {state} LLC formation real estate development',
    'ucc': '{city} {state} UCC filing real estate construction',
}


def refresh_government_signals():
    """
    Background job: populate government_signals from public data.
    Searches for gov-related signals per configured city using existing search cache / SerpAPI.
    Rate-limited, deduplicates by source_url, caps at 50 items per city per signal_type.
    Runs daily. Stubbed sources that fail return 0 without crashing.
    """
    import time as _time
    import requests as _req
    from datetime import date

    app.logger.info('[gov-signals] Starting refresh_government_signals()')
    total_inserted = 0
    total_skipped = 0
    city_stats = {}

    for city_conf in _GOV_SIGNAL_CITIES:
        city = city_conf['city']
        state = city_conf['state']
        city_key = f"{city}, {state}"
        city_count = 0

        for sig_type, query_tmpl in _GOV_SIGNAL_QUERIES.items():
            try:
                query = query_tmpl.format(city=city, state=state)

                # Use SerpAPI if available, otherwise skip
                serpapi_key = os.getenv('SERPAPI_API_KEY', '')
                if not serpapi_key:
                    app.logger.debug(f'[gov-signals] No SERPAPI_API_KEY, skipping {sig_type} for {city_key}')
                    continue

                # Rate limit: 1 second between SerpAPI calls
                _time.sleep(1.0)

                resp = _req.get('https://serpapi.com/search.json', params={
                    'api_key': serpapi_key,
                    'engine': 'google',
                    'q': query,
                    'num': 10,
                    'tbs': 'qdr:m',  # past month
                }, timeout=15)

                if resp.status_code != 200:
                    app.logger.warning(f'[gov-signals] SerpAPI {resp.status_code} for {sig_type}/{city_key}')
                    continue

                data = resp.json()
                organic = data.get('organic_results', [])

                conn = _get_db_conn()
                c = conn.cursor()

                # Count existing for this city+type to enforce cap of 50
                c.execute('SELECT COUNT(*) FROM government_signals WHERE city = ? AND state = ? AND signal_type = ?',
                          (city, state, sig_type))
                existing_count = c.fetchone()[0]
                remaining_cap = max(0, 50 - existing_count)

                inserted_this_batch = 0
                for item in organic[:remaining_cap]:
                    source_url = item.get('link', '')
                    if not source_url:
                        continue

                    # Deduplicate by source_url
                    c.execute('SELECT 1 FROM government_signals WHERE source_url = ?', (source_url,))
                    if c.fetchone():
                        total_skipped += 1
                        continue

                    title = item.get('title', '')
                    snippet = item.get('snippet', '')
                    source_name = item.get('source', item.get('displayed_link', 'web'))

                    # Extract operator name heuristic: look for known patterns
                    operator_name = None
                    for kw in ('LLC', 'Inc', 'Corp', 'LP', 'Trust', 'Partners', 'Group', 'Holdings', 'Development', 'Homes', 'Properties'):
                        for word_chunk in (title + ' ' + snippet).split(','):
                            if kw.lower() in word_chunk.lower() and len(word_chunk.strip()) < 80:
                                candidate = word_chunk.strip()
                                if len(candidate) > 3 and len(candidate) < 80:
                                    operator_name = candidate
                                    break
                        if operator_name:
                            break

                    # Extract amount heuristic: look for $X patterns
                    amount = None
                    import re as _re
                    amt_match = _re.search(r'\$[\d,.]+\s*(?:million|M|billion|B|K)?', title + ' ' + snippet, _re.IGNORECASE)
                    if amt_match:
                        amt_str = amt_match.group(0).replace('$', '').replace(',', '').strip()
                        multiplier = 1
                        if 'billion' in amt_str.lower() or amt_str.upper().endswith('B'):
                            multiplier = 1_000_000_000
                        elif 'million' in amt_str.lower() or amt_str.upper().endswith('M'):
                            multiplier = 1_000_000
                        elif amt_str.upper().endswith('K'):
                            multiplier = 1_000
                        num_str = _re.sub(r'[a-zA-Z\s]', '', amt_str)
                        try:
                            amount = float(num_str) * multiplier
                        except (ValueError, TypeError):
                            amount = None

                    filing_date = date.today().isoformat()
                    sig_id = str(uuid.uuid4())

                    c.execute('''INSERT INTO government_signals
                                 (id, city, state, signal_type, operator_name, operator_aliases,
                                  project_name, amount, filing_date, source_url, source_name,
                                  summary, raw_payload)
                                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                              (sig_id, city, state, sig_type, operator_name, None,
                               None, amount, filing_date, source_url, source_name,
                               (snippet or title)[:500], json.dumps(item)))

                    inserted_this_batch += 1
                    city_count += 1
                    total_inserted += 1

                conn.commit()
                conn.close()

                app.logger.info(f'[gov-signals] {sig_type}/{city_key}: inserted {inserted_this_batch}, skipped dupes')

            except Exception as e:
                app.logger.error(f'[gov-signals] Error fetching {sig_type} for {city_key}: {e}')
                continue

        city_stats[city_key] = city_count

    app.logger.info(f'[gov-signals] Refresh complete: {total_inserted} inserted, {total_skipped} skipped. Per-city: {city_stats}')
    return {'inserted': total_inserted, 'skipped': total_skipped, 'city_stats': city_stats}


def _scheduled_gov_signals():
    """Wrapper for scheduled government signal refresh."""
    try:
        from queue_config import enqueue
        enqueue(refresh_government_signals, job_timeout=600)
    except Exception as e:
        print(f"[Scheduler] Gov signals refresh error: {e}")


@app.route('/api/government-signals', methods=['GET'])
@require_auth
def api_government_signals():
    """Query government signals for a city/state. Returns strict ok/error shape."""
    city = request.args.get('city', '').strip()
    state = request.args.get('state', '').strip().upper()
    try:
        days = min(int(request.args.get('days', 90)), 365)
    except (ValueError, TypeError):
        days = 90

    if not city or not state:
        return jsonify({'ok': False, 'error': 'city and state parameters are required'}), 400

    try:
        cutoff = (datetime.utcnow() - timedelta(days=days)).strftime('%Y-%m-%d')
        conn = _get_db_conn()
        c = conn.cursor()
        c.execute('''SELECT id, city, state, signal_type, operator_name, operator_aliases,
                            project_name, amount, filing_date, source_url, source_name,
                            summary, created_at
                     FROM government_signals
                     WHERE LOWER(city) = LOWER(?) AND state = ? AND filing_date >= ?
                     ORDER BY filing_date DESC
                     LIMIT 200''',
                  (city, state, cutoff))

        signals = []
        for row in c.fetchall():
            aliases = []
            try:
                aliases = json.loads(row[5]) if row[5] else []
            except (json.JSONDecodeError, TypeError):
                pass
            signals.append({
                'id': row[0],
                'city': row[1],
                'state': row[2],
                'signal_type': row[3],
                'operator_name': row[4],
                'operator_aliases': aliases,
                'project_name': row[6],
                'amount': row[7],
                'filing_date': row[8],
                'source_url': row[9],
                'source_name': row[10],
                'summary': row[11],
                'created_at': row[12],
            })
        conn.close()

        app.logger.info(f'[gov-signals-api] city={city}, state={state}, days={days}, count={len(signals)}')

        return jsonify({
            'ok': True,
            'data': signals,
            'meta': {'count': len(signals), 'generated_at': datetime.utcnow().isoformat(),
                     'params': {'city': city, 'state': state, 'days': days}},
        })
    except Exception as e:
        app.logger.error(f'[gov-signals-api] Error: {e}')
        return jsonify({'ok': False, 'error': 'Failed to query government signals', 'details': str(e)}), 500


def _normalize_name(name):
    """Normalize a company name for fuzzy matching: lowercase, strip punctuation, common suffixes."""
    import re as _re
    if not name:
        return ''
    n = name.lower().strip()
    n = _re.sub(r'[^a-z0-9\s]', '', n)
    for suffix in ('llc', 'inc', 'corp', 'lp', 'ltd', 'co', 'company', 'group', 'holdings', 'properties', 'development', 'homes', 'partners'):
        n = _re.sub(r'\b' + suffix + r'\b', '', n)
    return ' '.join(n.split())


def _token_overlap_score(a, b):
    """Simple token overlap ratio between two normalized strings."""
    tokens_a = set(a.split())
    tokens_b = set(b.split())
    if not tokens_a or not tokens_b:
        return 0.0
    overlap = tokens_a & tokens_b
    return len(overlap) / min(len(tokens_a), len(tokens_b))


def enrich_prospects_with_gov_signals(prospects):
    """
    Given a list of prospect dicts (with 'company', 'city', 'state'),
    attach 'government_activity' array to each. Single DB query + in-memory matching.
    Never fails the caller — returns prospects unchanged on error.
    """
    import time as _time
    t0 = _time.time()

    try:
        # Collect unique (city, state) pairs
        city_state_pairs = set()
        for p in prospects:
            city = (p.get('city') or '').strip()
            state = (p.get('state') or '').strip().upper()
            if city and state:
                city_state_pairs.add((city.lower(), state))

        if not city_state_pairs:
            for p in prospects:
                p['government_activity'] = []
            return prospects

        # Single DB query for all relevant signals (last 90 days)
        cutoff = (datetime.utcnow() - timedelta(days=90)).strftime('%Y-%m-%d')
        conn = _get_db_conn()
        c = conn.cursor()

        # Build OR clauses for each city/state pair
        clauses = []
        params = []
        for city_lower, state in city_state_pairs:
            clauses.append('(LOWER(city) = ? AND state = ?)')
            params.extend([city_lower, state])
        params.append(cutoff)

        query = f'''SELECT city, state, signal_type, operator_name, operator_aliases,
                           filing_date, summary, amount, source_url, source_name
                    FROM government_signals
                    WHERE ({' OR '.join(clauses)}) AND filing_date >= ?
                    ORDER BY filing_date DESC'''
        c.execute(query, params)

        # Index signals by (city_lower, state)
        signals_by_location = {}
        for row in c.fetchall():
            key = (row[0].lower(), row[1])
            if key not in signals_by_location:
                signals_by_location[key] = []
            aliases = []
            try:
                aliases = json.loads(row[4]) if row[4] else []
            except (json.JSONDecodeError, TypeError):
                pass
            signals_by_location[key].append({
                'signal_type': row[2],
                'operator_name': row[3],
                'operator_aliases': aliases,
                'filing_date': row[5],
                'summary': row[6],
                'amount': row[7],
                'source_url': row[8],
                'source_name': row[9],
            })
        conn.close()

        enriched_count = 0
        for p in prospects:
            city = (p.get('city') or '').strip()
            state = (p.get('state') or '').strip().upper()
            company = p.get('company', '')
            key = (city.lower(), state)

            gov_activity = []
            location_signals = signals_by_location.get(key, [])

            if location_signals and company:
                norm_company = _normalize_name(company)
                for sig in location_signals:
                    matched = False
                    # Exact match on operator_name
                    if sig['operator_name'] and _normalize_name(sig['operator_name']) == norm_company:
                        matched = True
                    # Match any alias
                    if not matched and sig['operator_aliases']:
                        for alias in sig['operator_aliases']:
                            if _normalize_name(alias) == norm_company:
                                matched = True
                                break
                    # Conservative fuzzy: token overlap >= 0.6
                    if not matched and sig['operator_name']:
                        if _token_overlap_score(norm_company, _normalize_name(sig['operator_name'])) >= 0.6:
                            matched = True

                    if matched:
                        gov_activity.append({
                            'signal_type': sig['signal_type'],
                            'filing_date': sig['filing_date'],
                            'summary': sig['summary'],
                            'amount': sig['amount'],
                            'source_url': sig['source_url'],
                            'source_name': sig['source_name'],
                        })

            # Also attach unmatched location signals as location-level context
            # if no operator match found, still show location signals (capped at 5)
            if not gov_activity and location_signals:
                for sig in location_signals[:5]:
                    gov_activity.append({
                        'signal_type': sig['signal_type'],
                        'filing_date': sig['filing_date'],
                        'summary': sig['summary'],
                        'amount': sig['amount'],
                        'source_url': sig['source_url'],
                        'source_name': sig['source_name'],
                    })

            p['government_activity'] = gov_activity
            if gov_activity:
                enriched_count += 1

        elapsed = round(_time.time() - t0, 3)
        app.logger.info(f'[gov-enrich] {enriched_count}/{len(prospects)} prospects enriched with gov signals in {elapsed}s')

    except Exception as e:
        app.logger.error(f'[gov-enrich] Error enriching prospects: {e}')
        # Never fail — return prospects without gov data
        for p in prospects:
            if 'government_activity' not in p:
                p['government_activity'] = []

    return prospects


# --- Scheduler Setup ---
def _scheduled_discovery():
    """Wrapper that enqueues the scheduled discovery job (runs ALL adapters including permits)."""
    from queue_config import enqueue
    enqueue(run_daily_discovery, True, job_timeout=600)  # is_scheduled=True

def _scheduled_trend_detection():
    """Daily trend detection (runs after discovery, at 7:30am PT)."""
    try:
        run_trend_detection()
    except Exception as e:
        print(f"[Scheduler] Trend detection error: {e}")

def _scheduled_weekly_brief():
    """Monday weekly brief generation (7:00am PT)."""
    try:
        run_trend_detection()  # Refresh trends first
        generate_weekly_brief()
    except Exception as e:
        print(f"[Scheduler] Weekly brief error: {e}")

def _scheduled_optimization():
    """Daily optimization: weighted signals → momentum → call timing (8:00am PT)."""
    try:
        run_daily_optimization()
    except Exception as e:
        print(f"[Scheduler] Optimization error: {e}")

def _scheduled_permit_feed():
    """Scheduled permit-feed ingestion (5:15 AM PT, before discovery)."""
    try:
        from permit_feed.ingest_job import run_ingest
        print("[Scheduler] Starting permit feed ingestion…")
        items = run_ingest()
        print(f"[Scheduler] Permit feed done: {len(items)} new signals")
    except Exception as e:
        print(f"[Scheduler] Permit feed error (non-fatal): {e}")

_scheduler = BackgroundScheduler(daemon=True)
_scheduler.add_job(
    _scheduled_permit_feed,
    CronTrigger(hour=5, minute=15, timezone=pytz.timezone('America/Los_Angeles')),
    id='daily_permit_feed',
    name='Daily Permit Feed Ingestion',
    replace_existing=True
)
_scheduler.add_job(
    _scheduled_discovery,
    CronTrigger(
        hour=DISCOVERY_CONFIG['schedule_hour'],
        minute=DISCOVERY_CONFIG['schedule_minute'],
        timezone=pytz.timezone(DISCOVERY_CONFIG['timezone'])
    ),
    id='daily_discovery',
    name='Daily BTR Discovery',
    replace_existing=True
)
_scheduler.add_job(
    _scheduled_trend_detection,
    CronTrigger(hour=7, minute=30, timezone=pytz.timezone('America/Los_Angeles')),
    id='daily_trends',
    name='Daily Trend Detection',
    replace_existing=True
)
_scheduler.add_job(
    _scheduled_weekly_brief,
    CronTrigger(day_of_week='mon', hour=7, minute=0, timezone=pytz.timezone('America/Los_Angeles')),
    id='weekly_brief',
    name='Weekly Sunbelt Brief',
    replace_existing=True
)
_scheduler.add_job(
    _scheduled_optimization,
    CronTrigger(hour=6, minute=45, timezone=pytz.timezone('America/Los_Angeles')),
    id='daily_optimization',
    name='Daily Signal Optimization',
    replace_existing=True
)
_scheduler.add_job(
    _scheduled_gov_signals,
    CronTrigger(hour=5, minute=30, timezone=pytz.timezone('America/Los_Angeles')),
    id='daily_gov_signals',
    name='Daily Government Signals Refresh',
    replace_existing=True
)
_scheduler.start()
print(f"[Scheduler] Daily discovery scheduled for {DISCOVERY_CONFIG['schedule_hour']}:{DISCOVERY_CONFIG['schedule_minute']:02d} AM {DISCOVERY_CONFIG['timezone']}")
print("[Scheduler] Daily signal optimization at 6:45 AM PT")
print("[Scheduler] Daily trend detection at 7:30 AM PT")
print("[Scheduler] Weekly Sunbelt Brief every Monday 7:00 AM PT")
print("[Scheduler] Daily government signals refresh at 5:30 AM PT")
print("[Scheduler] Daily permit feed ingestion at 5:15 AM PT")


if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)

