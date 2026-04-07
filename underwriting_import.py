"""
Underwriting spreadsheet import logic.

Parses XLSX/CSV files, maps headers to canonical columns,
validates rows, and creates communities + underwriting_rows.
"""

import csv
import io
import re
import uuid
from datetime import datetime

from underwriting_columns import (
    UNDERWRITING_COLUMNS, COLUMN_KEYS, KEY_MAP, COLUMN_TYPES,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MAX_FILE_SIZE = 25 * 1024 * 1024  # 25 MB
MAX_ROWS = 5000

REQUIRED_FIELDS = [
    col["key"] for col in UNDERWRITING_COLUMNS if col.get("required")
]  # location_name, city, state_province_district, number_of_units

# Minimum required for row to NOT be skipped
REQUIRED_MIN = ["location_name", "city", "state_province_district"]

CURRENCY_RE = re.compile(r'[\$,\s]')
PERCENT_RE = re.compile(r'[%\s]')


# ---------------------------------------------------------------------------
# Header mapping
# ---------------------------------------------------------------------------

def _normalize_header(h: str) -> str:
    """Normalize header for loose matching: lowercase, collapse whitespace, strip."""
    h = h.strip()
    h = re.sub(r'\s+', ' ', h)
    return h.lower()


def map_headers(file_headers: list[str], strict: bool) -> tuple[dict, list[str]]:
    """
    Map spreadsheet headers to canonical column keys.

    Returns:
        (mapping, errors)
        mapping: {col_index: canonical_key}
        errors: list of error messages
    """
    mapping = {}
    errors = []

    # Build lookup tables
    exact_lookup = {col["header"]: col["key"] for col in UNDERWRITING_COLUMNS}
    normalized_lookup = {_normalize_header(col["header"]): col["key"] for col in UNDERWRITING_COLUMNS}

    matched_keys = set()

    for idx, header in enumerate(file_headers):
        if not header or not header.strip():
            continue

        # Try exact match first
        if header in exact_lookup:
            key = exact_lookup[header]
            if key not in matched_keys:
                mapping[idx] = key
                matched_keys.add(key)
            continue

        if strict:
            # In strict mode, only exact matches allowed
            errors.append(f"Column {idx + 1} header '{header}' does not match any canonical header (strict mode)")
            continue

        # Merge mode: try normalized match
        norm = _normalize_header(header)
        if norm in normalized_lookup:
            key = normalized_lookup[norm]
            if key not in matched_keys:
                mapping[idx] = key
                matched_keys.add(key)
            continue

        # Not matched — that's ok in merge mode, just skip the column
        pass

    if strict:
        # Check all canonical headers are present
        all_keys = set(COLUMN_KEYS)
        missing = all_keys - matched_keys
        if missing:
            missing_headers = [
                col["header"] for col in UNDERWRITING_COLUMNS if col["key"] in missing
            ]
            errors.append(f"Missing required headers in strict mode: {', '.join(missing_headers)}")

    return mapping, errors


# ---------------------------------------------------------------------------
# Type coercion
# ---------------------------------------------------------------------------

def _parse_currency(val: str) -> float | None:
    """Parse currency string: remove $, commas, whitespace -> float."""
    if not val or not val.strip():
        return None
    cleaned = CURRENCY_RE.sub('', val.strip())
    # Handle parentheses for negatives: (1,234) -> -1234
    if cleaned.startswith('(') and cleaned.endswith(')'):
        cleaned = '-' + cleaned[1:-1]
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _parse_percent(val: str) -> float | None:
    """Parse percent string: '85%' -> 85 (stored as the numeric value, matching DB convention)."""
    if not val or not val.strip():
        return None
    cleaned = PERCENT_RE.sub('', val.strip())
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _parse_numeric(val: str) -> float | None:
    if not val or not val.strip():
        return None
    cleaned = CURRENCY_RE.sub('', val.strip())
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _parse_integer(val: str) -> int | None:
    if not val or not val.strip():
        return None
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None


def _parse_date(val: str) -> tuple[str | None, str | None]:
    """
    Parse date string. Returns (parsed_value, warning_message).
    Accepts MM/DD/YYYY, DD/MM/YYYY, YYYY-MM-DD, etc.
    """
    if not val or not val.strip():
        return None, None

    val = val.strip()

    # Try ISO format first: YYYY-MM-DD
    iso_match = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', val)
    if iso_match:
        return val, None

    # Try MM/DD/YYYY or DD/MM/YYYY
    slash_match = re.match(r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$', val)
    if slash_match:
        a, b, year = int(slash_match.group(1)), int(slash_match.group(2)), slash_match.group(3)
        warning = None
        # If a > 12, it must be DD/MM
        if a > 12 and b <= 12:
            return f"{year}-{b:02d}-{a:02d}", None
        # If b > 12, it must be MM/DD
        if b > 12 and a <= 12:
            return f"{year}-{a:02d}-{b:02d}", None
        # Ambiguous: assume MM/DD/YYYY (US convention)
        if a <= 12 and b <= 12:
            warning = f"Ambiguous date '{val}' — interpreted as MM/DD/YYYY"
            return f"{year}-{a:02d}-{b:02d}", warning
        return None, f"Invalid date: '{val}'"

    # Return as-is if we can't parse
    return val, f"Unparseable date format: '{val}'"


def coerce_value(val, col_key: str) -> tuple[str | None, list[str]]:
    """
    Coerce a raw cell value to the appropriate type for storage.

    All values stored as TEXT in DB. Returns (coerced_string, warnings).
    """
    warnings = []
    col_type = COLUMN_TYPES.get(col_key, 'text')

    if val is None:
        return None, warnings

    # Convert non-string values
    raw = str(val).strip() if val is not None else ''
    if not raw:
        return None, warnings

    if col_type == 'currency':
        parsed = _parse_currency(raw)
        if parsed is not None:
            return str(parsed), warnings
        warnings.append(f"Could not parse currency value '{raw}' for {col_key}")
        return raw, warnings

    elif col_type == 'numeric':
        parsed = _parse_numeric(raw)
        if parsed is not None:
            return str(parsed), warnings
        warnings.append(f"Could not parse numeric value '{raw}' for {col_key}")
        return raw, warnings

    elif col_type == 'integer':
        parsed = _parse_integer(raw)
        if parsed is not None:
            return str(parsed), warnings
        warnings.append(f"Could not parse integer value '{raw}' for {col_key}")
        return raw, warnings

    elif col_type == 'percent':
        parsed = _parse_percent(raw)
        if parsed is not None:
            return str(parsed), warnings
        warnings.append(f"Could not parse percent value '{raw}' for {col_key}")
        return raw, warnings

    elif col_type == 'date':
        parsed, date_warning = _parse_date(raw)
        if date_warning:
            warnings.append(date_warning)
        return parsed, warnings

    else:
        return raw, warnings


# ---------------------------------------------------------------------------
# Community key generation
# ---------------------------------------------------------------------------

def _make_community_key(row_data: dict) -> str:
    """Generate a deterministic community key from location fields."""
    parts = [
        row_data.get('location_name', '') or '',
        row_data.get('city', '') or '',
        row_data.get('state_province_district', '') or '',
        row_data.get('location_street_address', '') or '',
        row_data.get('postal_zip_code', '') or '',
    ]
    # If location_id exists, prepend it for stability
    loc_id = row_data.get('location_id', '') or ''
    if loc_id:
        parts.insert(0, loc_id)

    # Normalize: lowercase, strip, collapse spaces
    normalized = '|'.join(
        re.sub(r'\s+', ' ', p.strip().lower()) for p in parts
    )
    return normalized


# ---------------------------------------------------------------------------
# File parsing
# ---------------------------------------------------------------------------

def parse_xlsx(file_bytes: bytes, sheet_name: str | None = None) -> tuple[list[str], list[list], str | None]:
    """
    Parse XLSX file bytes.
    Returns (headers, data_rows, error).
    Each data_row is a list of cell values.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return [], [], f"Failed to open XLSX file: {str(e)}"

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows_iter = ws.iter_rows()

    # First row = headers
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], [], "Spreadsheet is empty (no header row)"

    headers = [str(cell.value).strip() if cell.value is not None else '' for cell in header_row]

    data_rows = []
    for row in rows_iter:
        if len(data_rows) >= MAX_ROWS:
            break
        values = [cell.value for cell in row]
        # Skip completely empty rows
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
            continue
        data_rows.append(values)

    wb.close()
    return headers, data_rows, None


def parse_csv(file_bytes: bytes) -> tuple[list[str], list[list], str | None]:
    """
    Parse CSV file bytes.
    Returns (headers, data_rows, error).
    """
    try:
        text = file_bytes.decode('utf-8-sig')  # Handle BOM
    except UnicodeDecodeError:
        try:
            text = file_bytes.decode('latin-1')
        except UnicodeDecodeError:
            return [], [], "Could not decode CSV file (tried UTF-8 and Latin-1)"

    reader = csv.reader(io.StringIO(text))
    try:
        headers = next(reader)
    except StopIteration:
        return [], [], "CSV file is empty (no header row)"

    headers = [h.strip() for h in headers]

    data_rows = []
    for row in reader:
        if len(data_rows) >= MAX_ROWS:
            break
        if all(not v.strip() for v in row):
            continue
        data_rows.append(row)

    return headers, data_rows, None


# ---------------------------------------------------------------------------
# Main import function
# ---------------------------------------------------------------------------

def process_import(
    file_bytes: bytes,
    filename: str,
    dry_run: bool = False,
    strict_headers: bool = False,
    mode: str = "merge",
    db_connector=None,
) -> dict:
    """
    Process an underwriting spreadsheet import.

    Args:
        file_bytes: raw file content
        filename: original filename (for extension detection)
        dry_run: if True, validate only, do not write to DB
        strict_headers: if True, require exact header match
        mode: "merge" or "strict" (strict also implies strict_headers)
        db_connector: callable returning a DB connection

    Returns:
        {
            ok: bool,
            data: {
                imported_rows: int,
                created_communities: int,
                matched_communities: int,
                skipped_rows: int,
                errors: [{rowIndex, message, field?}],
                warnings: [{rowIndex, message}],
            }
        }
    """
    errors = []
    warnings = []
    result = {
        'imported_rows': 0,
        'created_communities': 0,
        'matched_communities': 0,
        'skipped_rows': 0,
        'errors': errors,
        'warnings': warnings,
    }

    # File size check
    if len(file_bytes) > MAX_FILE_SIZE:
        return {
            'ok': False,
            'error': f'File too large ({len(file_bytes) / 1024 / 1024:.1f}MB). Maximum is {MAX_FILE_SIZE / 1024 / 1024:.0f}MB.',
            'data': result,
        }

    if mode == 'strict':
        strict_headers = True

    # Parse file
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext == 'xlsx':
        headers, data_rows, parse_error = parse_xlsx(file_bytes)
    elif ext == 'csv':
        headers, data_rows, parse_error = parse_csv(file_bytes)
    else:
        return {
            'ok': False,
            'error': f'Unsupported file type: .{ext}. Please upload .xlsx or .csv.',
            'data': result,
        }

    if parse_error:
        return {'ok': False, 'error': parse_error, 'data': result}

    if not data_rows:
        return {
            'ok': False,
            'error': 'No data rows found in file (only headers detected).',
            'data': result,
        }

    if len(data_rows) > MAX_ROWS:
        warnings.append({
            'rowIndex': 0,
            'message': f'File has more than {MAX_ROWS} rows. Only the first {MAX_ROWS} will be processed.',
        })

    # Map headers
    header_mapping, header_errors = map_headers(headers, strict_headers)
    if header_errors:
        if strict_headers:
            for err in header_errors:
                errors.append({'rowIndex': 0, 'message': err})
            return {
                'ok': False,
                'error': 'Header mapping failed in strict mode. See errors for details.',
                'data': result,
            }
        else:
            for err in header_errors:
                warnings.append({'rowIndex': 0, 'message': err})

    if not header_mapping:
        return {
            'ok': False,
            'error': 'No columns could be mapped to canonical headers. Check that column headers match the template.',
            'data': result,
        }

    # Check that minimum required fields are mappable
    mapped_keys = set(header_mapping.values())
    missing_required = [k for k in REQUIRED_MIN if k not in mapped_keys]
    if missing_required:
        missing_headers = [
            col["header"] for col in UNDERWRITING_COLUMNS if col["key"] in missing_required
        ]
        return {
            'ok': False,
            'error': f'Required columns not found in file: {", ".join(missing_headers)}',
            'data': result,
        }

    # Process rows
    parsed_rows = []
    for row_idx, raw_row in enumerate(data_rows):
        spreadsheet_row_num = row_idx + 2  # +1 for header, +1 for 1-based
        row_data = {}
        row_warnings = []

        for col_idx, col_key in header_mapping.items():
            raw_val = raw_row[col_idx] if col_idx < len(raw_row) else None
            coerced, w = coerce_value(raw_val, col_key)
            row_data[col_key] = coerced
            for warning_msg in w:
                row_warnings.append({
                    'rowIndex': spreadsheet_row_num,
                    'message': warning_msg,
                    'field': col_key,
                })

        # Check required fields
        missing = []
        for req_key in REQUIRED_MIN:
            val = row_data.get(req_key)
            if not val or not str(val).strip():
                missing.append(req_key)

        if missing:
            missing_names = [
                col["header"] for col in UNDERWRITING_COLUMNS if col["key"] in missing
            ]
            errors.append({
                'rowIndex': spreadsheet_row_num,
                'message': f'Missing required fields: {", ".join(missing_names)}',
            })
            result['skipped_rows'] += 1
            continue

        warnings.extend(row_warnings)

        # Generate community key
        community_key = _make_community_key(row_data)
        parsed_rows.append({
            'community_key': community_key,
            'data': row_data,
            'row_index': spreadsheet_row_num,
        })

    if dry_run:
        # In dry-run, simulate community matching using the keys
        seen_keys = {}
        for pr in parsed_rows:
            ck = pr['community_key']
            if ck not in seen_keys:
                seen_keys[ck] = True
                result['created_communities'] += 1  # Approximate
            result['imported_rows'] += 1

        return {'ok': True, 'data': result}

    # ---- Write to DB ----
    if not db_connector:
        return {
            'ok': False,
            'error': 'No database connector provided.',
            'data': result,
        }

    conn = db_connector()
    c = conn.cursor()

    try:
        # Cache existing communities by key
        c.execute('SELECT id, community_key FROM underwriting_communities')
        existing = {row[1]: row[0] for row in c.fetchall()}

        # Cache max row_version per community
        c.execute('SELECT community_id, MAX(row_version) FROM underwriting_rows GROUP BY community_id')
        max_versions = {row[0]: row[1] for row in c.fetchall()}

        # Track communities created in this import batch
        batch_created = {}

        col_names = ', '.join(COLUMN_KEYS)
        placeholders = ', '.join(['?'] * len(COLUMN_KEYS))

        for pr in parsed_rows:
            ck = pr['community_key']
            row_data = pr['data']
            row_index = pr['row_index']

            try:
                # Find or create community
                community_id = existing.get(ck) or batch_created.get(ck)

                if community_id:
                    result['matched_communities'] += 1
                else:
                    # Create new community
                    community_id = str(uuid.uuid4())
                    now = datetime.utcnow().isoformat()
                    location_name = row_data.get('location_name', '')
                    c.execute(
                        'INSERT INTO underwriting_communities (id, community_key, location_name, created_at, updated_at) VALUES (?, ?, ?, ?, ?)',
                        (community_id, ck, location_name, now, now)
                    )
                    batch_created[ck] = community_id
                    existing[ck] = community_id
                    result['created_communities'] += 1

                # Determine row_version
                current_max = max_versions.get(community_id, 0)
                new_version = current_max + 1
                max_versions[community_id] = new_version

                # Build row values
                row_id = str(uuid.uuid4())
                now = datetime.utcnow().isoformat()
                values = [str(row_data.get(k, '') or '') for k in COLUMN_KEYS]

                c.execute(
                    f'INSERT INTO underwriting_rows (id, community_id, row_version, created_at, {col_names}) VALUES (?, ?, ?, ?, {placeholders})',
                    [row_id, community_id, new_version, now] + values
                )

                result['imported_rows'] += 1

            except Exception as e:
                errors.append({
                    'rowIndex': row_index,
                    'message': f'Database error: {str(e)}',
                })
                result['skipped_rows'] += 1
                continue

        conn.commit()

    except Exception as e:
        conn.rollback()
        conn.close()
        return {
            'ok': False,
            'error': f'Database transaction failed: {str(e)}',
            'data': result,
        }

    conn.close()

    return {'ok': True, 'data': result}
